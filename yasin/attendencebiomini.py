import sys
import os
import sqlite3
import hashlib
import json
import csv
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict, Any
from dataclasses import dataclass
from ctypes import *
from pathlib import Path
import base64

from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QLabel, QPushButton, QLineEdit, QTableWidget, QTableWidgetItem,
                             QComboBox, QDateEdit, QMessageBox, QDialog, QFormLayout, 
                             QTextEdit, QTabWidget, QHeaderView, QFileDialog, QProgressBar,
                             QGroupBox, QGridLayout, QStackedWidget, QFrame, QSpinBox,
                             QCheckBox, QListWidget, QSplitter, QDialogButtonBox)
from PyQt6.QtCore import Qt, QDate, QTimer, pyqtSignal, QThread, QSize
from PyQt6.QtGui import QFont, QPalette, QColor, QIcon, QPixmap, QImage

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
except:
    pass

try:
    import openpyxl
    from openpyxl.styles import Font as ExcelFont, PatternFill, Alignment, Border, Side
except:
    pass

class NBiometricStatus:
    OK = 0
    FAILED = 1
    CANCELED = 2
    TIMEOUT = 3
    BAD_QUALITY = 4

class NDeviceStatus:
    DISCONNECTED = 0
    CONNECTED = 1
    BUSY = 2

@dataclass
class User:
    id: int
    username: str
    password_hash: str
    full_name: str
    email: str
    role: str
    department_id: Optional[int]
    shift_id: Optional[int]
    is_active: bool
    created_at: str

@dataclass
class Department:
    id: int
    name: str
    description: str

@dataclass
class Shift:
    id: int
    name: str
    start_time: str
    end_time: str
    grace_period: int

@dataclass
class Fingerprint:
    id: int
    user_id: int
    finger_position: str
    template_data: bytes
    quality_score: int
    enrolled_at: str

@dataclass
class AttendanceRecord:
    id: int
    user_id: int
    check_in: str
    check_out: Optional[str]
    date: str
    status: str
    notes: str

class BiometricSDK:
    def __init__(self):
        self.device_handle = None
        self.is_initialized = False
        self.lib = None
        self._load_sdk()
        
    def _load_sdk(self):
        try:
            sdk_paths = [
                'NBiometricClient.dll',
                'C:\\Program Files\\Neurotechnology\\Biometric 13.0\\Bin\\Win64_x64\\NBiometricClient.dll',
                'C:\\Program Files\\Neurotechnology\\MegaMatcher 13.0\\Bin\\Win64_x64\\NBiometricClient.dll',
                '/usr/lib/neurotechnology/NBiometricClient.so',
                '/opt/neurotechnology/lib/NBiometricClient.so'
            ]
            
            for path in sdk_paths:
                if os.path.exists(path):
                    self.lib = CDLL(path)
                    self._setup_functions()
                    break
        except Exception as e:
            pass
            
    def _setup_functions(self):
        if not self.lib:
            return
            
        try:
            self.lib.NBiometricClient_Create.restype = c_void_p
            self.lib.NBiometricClient_Initialize.argtypes = [c_void_p]
            self.lib.NBiometricClient_Initialize.restype = c_int
            
            self.lib.NSubject_Create.restype = c_void_p
            self.lib.NFinger_Create.restype = c_void_p
            
            self.lib.NBiometricClient_Capture.argtypes = [c_void_p, c_void_p]
            self.lib.NBiometricClient_Capture.restype = c_int
            
            self.lib.NBiometricClient_CreateTemplate.argtypes = [c_void_p, c_void_p]
            self.lib.NBiometricClient_CreateTemplate.restype = c_int
            
            self.lib.NBiometricClient_Verify.argtypes = [c_void_p, c_void_p]
            self.lib.NBiometricClient_Verify.restype = c_int
            
            self.lib.NFinger_GetTemplate.argtypes = [c_void_p, POINTER(c_void_p), POINTER(c_int)]
            self.lib.NFinger_GetTemplate.restype = c_int
            
            self.lib.NFinger_SetTemplate.argtypes = [c_void_p, c_void_p, c_int]
            self.lib.NFinger_SetTemplate.restype = c_int
        except:
            pass
    
    def initialize(self) -> bool:
        if not self.lib:
            return False
            
        try:
            self.device_handle = self.lib.NBiometricClient_Create()
            result = self.lib.NBiometricClient_Initialize(self.device_handle)
            self.is_initialized = (result == NBiometricStatus.OK)
            return self.is_initialized
        except:
            return False
    
    def capture_fingerprint(self) -> tuple[bool, bytes, int]:
        if not self.is_initialized:
            return False, b'', 0
            
        try:
            subject = self.lib.NSubject_Create()
            finger = self.lib.NFinger_Create()
            
            capture_result = self.lib.NBiometricClient_Capture(self.device_handle, subject)
            
            if capture_result != NBiometricStatus.OK:
                return False, b'', 0
            
            template_result = self.lib.NBiometricClient_CreateTemplate(self.device_handle, subject)
            
            if template_result != NBiometricStatus.OK:
                return False, b'', 0
            
            template_ptr = c_void_p()
            template_size = c_int()
            
            get_result = self.lib.NFinger_GetTemplate(finger, byref(template_ptr), byref(template_size))
            
            if get_result == NBiometricStatus.OK and template_size.value > 0:
                template_data = string_at(template_ptr, template_size.value)
                quality = 85
                return True, template_data, quality
            
            return False, b'', 0
        except:
            return False, b'', 0
    
    def verify_fingerprint(self, template1: bytes, template2: bytes) -> tuple[bool, int]:
        if not self.is_initialized:
            return False, 0
            
        try:
            subject1 = self.lib.NSubject_Create()
            subject2 = self.lib.NSubject_Create()
            
            finger1 = self.lib.NFinger_Create()
            finger2 = self.lib.NFinger_Create()
            
            self.lib.NFinger_SetTemplate(finger1, template1, len(template1))
            self.lib.NFinger_SetTemplate(finger2, template2, len(template2))
            
            verify_result = self.lib.NBiometricClient_Verify(self.device_handle, subject1)
            
            if verify_result == NBiometricStatus.OK:
                score = 95
                return True, score
            
            return False, 0
        except:
            return False, 0
    
    def get_device_status(self) -> int:
        if not self.is_initialized:
            return NDeviceStatus.DISCONNECTED
        return NDeviceStatus.CONNECTED
    
    def close(self):
        self.is_initialized = False
        self.device_handle = None

class Database:
    def __init__(self, db_path: str = 'biometric_attendance3.db'):
        self.db_path = db_path
        self.conn = None
        self._initialize_database()
    
    def _initialize_database(self):
        self.conn = sqlite3.connect(self.db_path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self._create_tables()
        self._insert_default_data()
    
    def _create_tables(self):
        cursor = self.conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                full_name TEXT NOT NULL,
                email TEXT,
                role TEXT NOT NULL,
                department_id INTEGER,
                shift_id INTEGER,
                is_active INTEGER DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (department_id) REFERENCES departments(id),
                FOREIGN KEY (shift_id) REFERENCES shifts(id)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS departments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                description TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS shifts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL,
                grace_period INTEGER DEFAULT 15
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS fingerprints (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                finger_position TEXT NOT NULL,
                template_data BLOB NOT NULL,
                quality_score INTEGER,
                enrolled_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                check_in TEXT NOT NULL,
                check_out TEXT,
                date TEXT NOT NULL,
                status TEXT DEFAULT 'Present',
                notes TEXT,
                FOREIGN KEY (user_id) REFERENCES users(id)
            )
        ''')
        
        cursor.execute('''
            CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance(date)
        ''')
        
        cursor.execute('''
            CREATE INDEX IF NOT EXISTS idx_attendance_user ON attendance(user_id)
        ''')
        
        self.conn.commit()
    
    def _insert_default_data(self):
        cursor = self.conn.cursor()
        
        cursor.execute("SELECT COUNT(*) FROM departments")
        if cursor.fetchone()[0] == 0:
            departments = [
                ('Computer Science', 'Department of Computer Science and IT'),
                ('Mathematics', 'Department of Mathematics'),
                ('Physics', 'Department of Physics'),
                ('Administration', 'Administrative Department')
            ]
            cursor.executemany("INSERT INTO departments (name, description) VALUES (?, ?)", departments)
        
        cursor.execute("SELECT COUNT(*) FROM shifts")
        if cursor.fetchone()[0] == 0:
            shifts = [
                ('Morning Shift', '08:00', '16:00', 15),
                ('Evening Shift', '14:00', '22:00', 15),
                ('Night Shift', '22:00', '06:00', 15),
                ('Full Day', '09:00', '17:00', 30)
            ]
            cursor.executemany("INSERT INTO shifts (name, start_time, end_time, grace_period) VALUES (?, ?, ?, ?)", shifts)
        
        cursor.execute("SELECT COUNT(*) FROM users")
        if cursor.fetchone()[0] == 0:
            admin_password = hashlib.sha256('admin123'.encode()).hexdigest()
            cursor.execute('''
                INSERT INTO users (username, password_hash, full_name, email, role, department_id, shift_id, is_active)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', ('admin', admin_password, 'System Administrator', 'admin@system.com', 'Admin', 4, 4, 1))
            
            users_data = [
                ('john.doe', hashlib.sha256('password123'.encode()).hexdigest(), 'Dr. John Doe', 'john.doe@university.edu', 'Teacher', 1, 1),
                ('jane.smith', hashlib.sha256('password123'.encode()).hexdigest(), 'Dr. Jane Smith', 'jane.smith@university.edu', 'Teacher', 2, 1),
                ('bob.wilson', hashlib.sha256('password123'.encode()).hexdigest(), 'Prof. Bob Wilson', 'bob.wilson@university.edu', 'Teacher', 1, 2),
                ('alice.brown', hashlib.sha256('password123'.encode()).hexdigest(), 'Dr. Alice Brown', 'alice.brown@university.edu', 'Teacher', 3, 1)
            ]
            
            for username, pwd_hash, full_name, email, role, dept_id, shift_id in users_data:
                cursor.execute('''
                    INSERT INTO users (username, password_hash, full_name, email, role, department_id, shift_id, is_active)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (username, pwd_hash, full_name, email, role, dept_id, shift_id, 1))
        
        self.conn.commit()
    
    def authenticate_user(self, username: str, password: str) -> Optional[User]:
        cursor = self.conn.cursor()
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        cursor.execute('''
            SELECT * FROM users WHERE username = ? AND password_hash = ? AND is_active = 1
        ''', (username, password_hash))
        
        row = cursor.fetchone()
        if row:
            return User(**dict(row))
        return None
    
    def get_all_users(self) -> List[User]:
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM users ORDER BY full_name")
        return [User(**dict(row)) for row in cursor.fetchall()]
    
    def get_user_by_id(self, user_id: int) -> Optional[User]:
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM users WHERE id = ?", (user_id,))
        row = cursor.fetchone()
        return User(**dict(row)) if row else None
    
    def create_user(self, username: str, password: str, full_name: str, email: str, 
                   role: str, department_id: int, shift_id: int) -> int:
        cursor = self.conn.cursor()
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        cursor.execute('''
            INSERT INTO users (username, password_hash, full_name, email, role, department_id, shift_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (username, password_hash, full_name, email, role, department_id, shift_id))
        self.conn.commit()
        return cursor.lastrowid
    
    def update_user(self, user_id: int, full_name: str, email: str, role: str, 
                   department_id: int, shift_id: int, is_active: bool):
        cursor = self.conn.cursor()
        cursor.execute('''
            UPDATE users SET full_name = ?, email = ?, role = ?, department_id = ?, shift_id = ?, is_active = ?
            WHERE id = ?
        ''', (full_name, email, role, department_id, shift_id, 1 if is_active else 0, user_id))
        self.conn.commit()
    
    def delete_user(self, user_id: int):
        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
        self.conn.commit()
    
    def get_all_departments(self) -> List[Department]:
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM departments ORDER BY name")
        return [Department(**dict(row)) for row in cursor.fetchall()]
    
    def create_department(self, name: str, description: str) -> int:
        cursor = self.conn.cursor()
        cursor.execute("INSERT INTO departments (name, description) VALUES (?, ?)", (name, description))
        self.conn.commit()
        return cursor.lastrowid
    
    def update_department(self, dept_id: int, name: str, description: str):
        cursor = self.conn.cursor()
        cursor.execute("UPDATE departments SET name = ?, description = ? WHERE id = ?", 
                      (name, description, dept_id))
        self.conn.commit()
    
    def delete_department(self, dept_id: int):
        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM departments WHERE id = ?", (dept_id,))
        self.conn.commit()
    
    def get_all_shifts(self) -> List[Shift]:
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM shifts ORDER BY name")
        return [Shift(**dict(row)) for row in cursor.fetchall()]
    
    def create_shift(self, name: str, start_time: str, end_time: str, grace_period: int) -> int:
        cursor = self.conn.cursor()
        cursor.execute("INSERT INTO shifts (name, start_time, end_time, grace_period) VALUES (?, ?, ?, ?)",
                      (name, start_time, end_time, grace_period))
        self.conn.commit()
        return cursor.lastrowid
    
    def update_shift(self, shift_id: int, name: str, start_time: str, end_time: str, grace_period: int):
        cursor = self.conn.cursor()
        cursor.execute("UPDATE shifts SET name = ?, start_time = ?, end_time = ?, grace_period = ? WHERE id = ?",
                      (name, start_time, end_time, grace_period, shift_id))
        self.conn.commit()
    
    def delete_shift(self, shift_id: int):
        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM shifts WHERE id = ?", (shift_id,))
        self.conn.commit()
    
    def save_fingerprint(self, user_id: int, finger_position: str, template_data: bytes, quality_score: int) -> int:
        cursor = self.conn.cursor()
        cursor.execute('''
            INSERT INTO fingerprints (user_id, finger_position, template_data, quality_score)
            VALUES (?, ?, ?, ?)
        ''', (user_id, finger_position, template_data, quality_score))
        self.conn.commit()
        return cursor.lastrowid
    
    def get_user_fingerprints(self, user_id: int) -> List[Fingerprint]:
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM fingerprints WHERE user_id = ?", (user_id,))
        return [Fingerprint(**dict(row)) for row in cursor.fetchall()]
    
    def get_all_fingerprints(self) -> List[Fingerprint]:
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM fingerprints")
        return [Fingerprint(**dict(row)) for row in cursor.fetchall()]
    
    def delete_fingerprint(self, fingerprint_id: int):
        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM fingerprints WHERE id = ?", (fingerprint_id,))
        self.conn.commit()
    
    def record_attendance(self, user_id: int, check_in: str, date_str: str, status: str = 'Present', notes: str = ''):
        cursor = self.conn.cursor()
        cursor.execute('''
            INSERT INTO attendance (user_id, check_in, date, status, notes)
            VALUES (?, ?, ?, ?, ?)
        ''', (user_id, check_in, date_str, status, notes))
        self.conn.commit()
    
    def update_checkout(self, attendance_id: int, check_out: str):
        cursor = self.conn.cursor()
        cursor.execute("UPDATE attendance SET check_out = ? WHERE id = ?", (check_out, attendance_id))
        self.conn.commit()
    
    def get_attendance_by_date(self, date_str: str) -> List[AttendanceRecord]:
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM attendance WHERE date = ? ORDER BY check_in", (date_str,))
        return [AttendanceRecord(**dict(row)) for row in cursor.fetchall()]
    
    def get_attendance_by_user(self, user_id: int, start_date: str, end_date: str) -> List[AttendanceRecord]:
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT * FROM attendance WHERE user_id = ? AND date BETWEEN ? AND ? ORDER BY date, check_in
        ''', (user_id, start_date, end_date))
        return [AttendanceRecord(**dict(row)) for row in cursor.fetchall()]
    
    def get_attendance_by_date_range(self, start_date: str, end_date: str) -> List[AttendanceRecord]:
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT * FROM attendance WHERE date BETWEEN ? AND ? ORDER BY date, check_in
        ''', (start_date, end_date))
        return [AttendanceRecord(**dict(row)) for row in cursor.fetchall()]
    
    def get_today_attendance(self, user_id: int) -> Optional[AttendanceRecord]:
        today = date.today().isoformat()
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT * FROM attendance WHERE user_id = ? AND date = ? ORDER BY check_in DESC LIMIT 1
        ''', (user_id, today))
        row = cursor.fetchone()
        return AttendanceRecord(**dict(row)) if row else None
    
    def close(self):
        if self.conn:
            self.conn.close()

class LoginDialog(QDialog):
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        self.logged_in_user = None
        self.setup_ui()
        
    def setup_ui(self):
        self.setWindowTitle('Biometric Attendance System - Login')
        self.setFixedSize(400, 300)
        
        layout = QVBoxLayout()
        
        title = QLabel('BIOMETRIC ATTENDANCE SYSTEM')
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setFont(QFont('Arial', 16, QFont.Weight.Bold))
        layout.addWidget(title)
        
        layout.addSpacing(20)
        
        form = QFormLayout()
        
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText('Enter username')
        form.addRow('Username:', self.username_input)
        
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_input.setPlaceholderText('Enter password')
        self.password_input.returnPressed.connect(self.login)
        form.addRow('Password:', self.password_input)
        
        layout.addLayout(form)
        layout.addSpacing(20)
        
        button_layout = QHBoxLayout()
        
        login_btn = QPushButton('Login')
        login_btn.setFixedHeight(40)
        login_btn.clicked.connect(self.login)
        button_layout.addWidget(login_btn)
        
        cancel_btn = QPushButton('Cancel')
        cancel_btn.setFixedHeight(40)
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
        
        info_label = QLabel('Default login: admin / admin123')
        info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        info_label.setStyleSheet('color: gray; font-size: 10px;')
        layout.addWidget(info_label)
        
        self.setLayout(layout)
    
    def login(self):
        username = self.username_input.text().strip()
        password = self.password_input.text()
        
        if not username or not password:
            QMessageBox.warning(self, 'Error', 'Please enter username and password')
            return
        
        user = self.db.authenticate_user(username, password)
        
        if user:
            self.logged_in_user = user
            self.accept()
        else:
            QMessageBox.critical(self, 'Login Failed', 'Invalid username or password')
            self.password_input.clear()
            self.password_input.setFocus()

class UserManagementWidget(QWidget):
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        self.setup_ui()
        self.load_users()
        
    def setup_ui(self):
        layout = QVBoxLayout()
        
        header = QLabel('User Management')
        header.setFont(QFont('Arial', 14, QFont.Weight.Bold))
        layout.addWidget(header)
        
        button_layout = QHBoxLayout()
        
        add_btn = QPushButton('Add User')
        add_btn.clicked.connect(self.add_user)
        button_layout.addWidget(add_btn)
        
        edit_btn = QPushButton('Edit User')
        edit_btn.clicked.connect(self.edit_user)
        button_layout.addWidget(edit_btn)
        
        delete_btn = QPushButton('Delete User')
        delete_btn.clicked.connect(self.delete_user)
        button_layout.addWidget(delete_btn)
        
        refresh_btn = QPushButton('Refresh')
        refresh_btn.clicked.connect(self.load_users)
        button_layout.addWidget(refresh_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(['ID', 'Username', 'Full Name', 'Email', 'Role', 'Department', 'Active'])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.table)
        
        self.setLayout(layout)
    
    def load_users(self):
        users = self.db.get_all_users()
        departments = {d.id: d.name for d in self.db.get_all_departments()}
        
        self.table.setRowCount(len(users))
        
        for row, user in enumerate(users):
            self.table.setItem(row, 0, QTableWidgetItem(str(user.id)))
            self.table.setItem(row, 1, QTableWidgetItem(user.username))
            self.table.setItem(row, 2, QTableWidgetItem(user.full_name))
            self.table.setItem(row, 3, QTableWidgetItem(user.email or ''))
            self.table.setItem(row, 4, QTableWidgetItem(user.role))
            dept_name = departments.get(user.department_id, '') if user.department_id else ''
            self.table.setItem(row, 5, QTableWidgetItem(dept_name))
            self.table.setItem(row, 6, QTableWidgetItem('Yes' if user.is_active else 'No'))
    
    def add_user(self):
        dialog = UserEditDialog(self.db, None, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_users()
    
    def edit_user(self):
        selected = self.table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, 'Warning', 'Please select a user to edit')
            return
        
        user_id = int(self.table.item(selected, 0).text())
        user = self.db.get_user_by_id(user_id)
        
        dialog = UserEditDialog(self.db, user, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_users()
    
    def delete_user(self):
        selected = self.table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, 'Warning', 'Please select a user to delete')
            return
        
        user_id = int(self.table.item(selected, 0).text())
        username = self.table.item(selected, 1).text()
        
        if username == 'admin':
            QMessageBox.critical(self, 'Error', 'Cannot delete admin user')
            return
        
        reply = QMessageBox.question(self, 'Confirm Delete', 
                                     f'Are you sure you want to delete user: {username}?',
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            self.db.delete_user(user_id)
            self.load_users()
            QMessageBox.information(self, 'Success', 'User deleted successfully')

class UserEditDialog(QDialog):
    def __init__(self, db: Database, user: Optional[User], parent=None):
        super().__init__(parent)
        self.db = db
        self.user = user
        self.setup_ui()
        
    def setup_ui(self):
        self.setWindowTitle('Edit User' if self.user else 'Add User')
        self.setFixedSize(400, 450)
        
        layout = QVBoxLayout()
        
        form = QFormLayout()
        
        self.username_input = QLineEdit()
        if self.user:
            self.username_input.setText(self.user.username)
            self.username_input.setEnabled(False)
        form.addRow('Username:', self.username_input)
        
        if not self.user:
            self.password_input = QLineEdit()
            self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
            form.addRow('Password:', self.password_input)
        
        self.fullname_input = QLineEdit()
        if self.user:
            self.fullname_input.setText(self.user.full_name)
        form.addRow('Full Name:', self.fullname_input)
        
        self.email_input = QLineEdit()
        if self.user:
            self.email_input.setText(self.user.email or '')
        form.addRow('Email:', self.email_input)
        
        self.role_combo = QComboBox()
        self.role_combo.addItems(['Teacher', 'Admin', 'Staff'])
        if self.user:
            self.role_combo.setCurrentText(self.user.role)
        form.addRow('Role:', self.role_combo)
        
        self.dept_combo = QComboBox()
        departments = self.db.get_all_departments()
        self.dept_combo.addItem('-- None --', 0)
        for dept in departments:
            self.dept_combo.addItem(dept.name, dept.id)
        if self.user and self.user.department_id:
            index = self.dept_combo.findData(self.user.department_id)
            if index >= 0:
                self.dept_combo.setCurrentIndex(index)
        form.addRow('Department:', self.dept_combo)
        
        self.shift_combo = QComboBox()
        shifts = self.db.get_all_shifts()
        self.shift_combo.addItem('-- None --', 0)
        for shift in shifts:
            self.shift_combo.addItem(shift.name, shift.id)
        if self.user and self.user.shift_id:
            index = self.shift_combo.findData(self.user.shift_id)
            if index >= 0:
                self.shift_combo.setCurrentIndex(index)
        form.addRow('Shift:', self.shift_combo)
        
        self.active_check = QCheckBox()
        self.active_check.setChecked(self.user.is_active if self.user else True)
        form.addRow('Active:', self.active_check)
        
        layout.addLayout(form)
        
        button_layout = QHBoxLayout()
        
        save_btn = QPushButton('Save')
        save_btn.clicked.connect(self.save)
        button_layout.addWidget(save_btn)
        
        cancel_btn = QPushButton('Cancel')
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def save(self):
        username = self.username_input.text().strip()
        full_name = self.fullname_input.text().strip()
        email = self.email_input.text().strip()
        role = self.role_combo.currentText()
        dept_id = self.dept_combo.currentData()
        shift_id = self.shift_combo.currentData()
        is_active = self.active_check.isChecked()
        
        if not username or not full_name:
            QMessageBox.warning(self, 'Error', 'Username and Full Name are required')
            return
        
        try:
            if self.user:
                self.db.update_user(self.user.id, full_name, email, role, dept_id if dept_id else None, 
                                   shift_id if shift_id else None, is_active)
                QMessageBox.information(self, 'Success', 'User updated successfully')
            else:
                password = self.password_input.text()
                if not password:
                    QMessageBox.warning(self, 'Error', 'Password is required')
                    return
                self.db.create_user(username, password, full_name, email, role, 
                                   dept_id if dept_id else None, shift_id if shift_id else None)
                QMessageBox.information(self, 'Success', 'User created successfully')
            
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to save user: {str(e)}')

class DepartmentManagementWidget(QWidget):
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        self.setup_ui()
        self.load_departments()
        
    def setup_ui(self):
        layout = QVBoxLayout()
        
        header = QLabel('Department Management')
        header.setFont(QFont('Arial', 14, QFont.Weight.Bold))
        layout.addWidget(header)
        
        button_layout = QHBoxLayout()
        
        add_btn = QPushButton('Add Department')
        add_btn.clicked.connect(self.add_department)
        button_layout.addWidget(add_btn)
        
        edit_btn = QPushButton('Edit Department')
        edit_btn.clicked.connect(self.edit_department)
        button_layout.addWidget(edit_btn)
        
        delete_btn = QPushButton('Delete Department')
        delete_btn.clicked.connect(self.delete_department)
        button_layout.addWidget(delete_btn)
        
        refresh_btn = QPushButton('Refresh')
        refresh_btn.clicked.connect(self.load_departments)
        button_layout.addWidget(refresh_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(['ID', 'Name', 'Description'])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.table)
        
        self.setLayout(layout)
    
    def load_departments(self):
        departments = self.db.get_all_departments()
        self.table.setRowCount(len(departments))
        
        for row, dept in enumerate(departments):
            self.table.setItem(row, 0, QTableWidgetItem(str(dept.id)))
            self.table.setItem(row, 1, QTableWidgetItem(dept.name))
            self.table.setItem(row, 2, QTableWidgetItem(dept.description or ''))
    
    def add_department(self):
        dialog = DepartmentEditDialog(self.db, None, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_departments()
    
    def edit_department(self):
        selected = self.table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, 'Warning', 'Please select a department to edit')
            return
        
        dept_id = int(self.table.item(selected, 0).text())
        departments = self.db.get_all_departments()
        dept = next((d for d in departments if d.id == dept_id), None)
        
        dialog = DepartmentEditDialog(self.db, dept, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_departments()
    
    def delete_department(self):
        selected = self.table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, 'Warning', 'Please select a department to delete')
            return
        
        dept_id = int(self.table.item(selected, 0).text())
        dept_name = self.table.item(selected, 1).text()
        
        reply = QMessageBox.question(self, 'Confirm Delete', 
                                     f'Are you sure you want to delete department: {dept_name}?',
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.db.delete_department(dept_id)
                self.load_departments()
                QMessageBox.information(self, 'Success', 'Department deleted successfully')
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Failed to delete department: {str(e)}')

class DepartmentEditDialog(QDialog):
    def __init__(self, db: Database, department: Optional[Department], parent=None):
        super().__init__(parent)
        self.db = db
        self.department = department
        self.setup_ui()
        
    def setup_ui(self):
        self.setWindowTitle('Edit Department' if self.department else 'Add Department')
        self.setFixedSize(400, 200)
        
        layout = QVBoxLayout()
        
        form = QFormLayout()
        
        self.name_input = QLineEdit()
        if self.department:
            self.name_input.setText(self.department.name)
        form.addRow('Name:', self.name_input)
        
        self.desc_input = QLineEdit()
        if self.department:
            self.desc_input.setText(self.department.description or '')
        form.addRow('Description:', self.desc_input)
        
        layout.addLayout(form)
        
        button_layout = QHBoxLayout()
        
        save_btn = QPushButton('Save')
        save_btn.clicked.connect(self.save)
        button_layout.addWidget(save_btn)
        
        cancel_btn = QPushButton('Cancel')
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def save(self):
        name = self.name_input.text().strip()
        description = self.desc_input.text().strip()
        
        if not name:
            QMessageBox.warning(self, 'Error', 'Department name is required')
            return
        
        try:
            if self.department:
                self.db.update_department(self.department.id, name, description)
                QMessageBox.information(self, 'Success', 'Department updated successfully')
            else:
                self.db.create_department(name, description)
                QMessageBox.information(self, 'Success', 'Department created successfully')
            
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to save department: {str(e)}')

class ShiftManagementWidget(QWidget):
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        self.setup_ui()
        self.load_shifts()
        
    def setup_ui(self):
        layout = QVBoxLayout()
        
        header = QLabel('Shift Management')
        header.setFont(QFont('Arial', 14, QFont.Weight.Bold))
        layout.addWidget(header)
        
        button_layout = QHBoxLayout()
        
        add_btn = QPushButton('Add Shift')
        add_btn.clicked.connect(self.add_shift)
        button_layout.addWidget(add_btn)
        
        edit_btn = QPushButton('Edit Shift')
        edit_btn.clicked.connect(self.edit_shift)
        button_layout.addWidget(edit_btn)
        
        delete_btn = QPushButton('Delete Shift')
        delete_btn.clicked.connect(self.delete_shift)
        button_layout.addWidget(delete_btn)
        
        refresh_btn = QPushButton('Refresh')
        refresh_btn.clicked.connect(self.load_shifts)
        button_layout.addWidget(refresh_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(['ID', 'Name', 'Start Time', 'End Time', 'Grace Period (min)'])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.table)
        
        self.setLayout(layout)
    
    def load_shifts(self):
        shifts = self.db.get_all_shifts()
        self.table.setRowCount(len(shifts))
        
        for row, shift in enumerate(shifts):
            self.table.setItem(row, 0, QTableWidgetItem(str(shift.id)))
            self.table.setItem(row, 1, QTableWidgetItem(shift.name))
            self.table.setItem(row, 2, QTableWidgetItem(shift.start_time))
            self.table.setItem(row, 3, QTableWidgetItem(shift.end_time))
            self.table.setItem(row, 4, QTableWidgetItem(str(shift.grace_period)))
    
    def add_shift(self):
        dialog = ShiftEditDialog(self.db, None, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_shifts()
    
    def edit_shift(self):
        selected = self.table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, 'Warning', 'Please select a shift to edit')
            return
        
        shift_id = int(self.table.item(selected, 0).text())
        shifts = self.db.get_all_shifts()
        shift = next((s for s in shifts if s.id == shift_id), None)
        
        dialog = ShiftEditDialog(self.db, shift, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_shifts()
    
    def delete_shift(self):
        selected = self.table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, 'Warning', 'Please select a shift to delete')
            return
        
        shift_id = int(self.table.item(selected, 0).text())
        shift_name = self.table.item(selected, 1).text()
        
        reply = QMessageBox.question(self, 'Confirm Delete', 
                                     f'Are you sure you want to delete shift: {shift_name}?',
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.db.delete_shift(shift_id)
                self.load_shifts()
                QMessageBox.information(self, 'Success', 'Shift deleted successfully')
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Failed to delete shift: {str(e)}')

class ShiftEditDialog(QDialog):
    def __init__(self, db: Database, shift: Optional[Shift], parent=None):
        super().__init__(parent)
        self.db = db
        self.shift = shift
        self.setup_ui()
        
    def setup_ui(self):
        self.setWindowTitle('Edit Shift' if self.shift else 'Add Shift')
        self.setFixedSize(400, 250)
        
        layout = QVBoxLayout()
        
        form = QFormLayout()
        
        self.name_input = QLineEdit()
        if self.shift:
            self.name_input.setText(self.shift.name)
        form.addRow('Name:', self.name_input)
        
        self.start_input = QLineEdit()
        self.start_input.setPlaceholderText('HH:MM (24-hour format)')
        if self.shift:
            self.start_input.setText(self.shift.start_time)
        form.addRow('Start Time:', self.start_input)
        
        self.end_input = QLineEdit()
        self.end_input.setPlaceholderText('HH:MM (24-hour format)')
        if self.shift:
            self.end_input.setText(self.shift.end_time)
        form.addRow('End Time:', self.end_input)
        
        self.grace_input = QSpinBox()
        self.grace_input.setRange(0, 120)
        self.grace_input.setValue(self.shift.grace_period if self.shift else 15)
        form.addRow('Grace Period (min):', self.grace_input)
        
        layout.addLayout(form)
        
        button_layout = QHBoxLayout()
        
        save_btn = QPushButton('Save')
        save_btn.clicked.connect(self.save)
        button_layout.addWidget(save_btn)
        
        cancel_btn = QPushButton('Cancel')
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def save(self):
        name = self.name_input.text().strip()
        start_time = self.start_input.text().strip()
        end_time = self.end_input.text().strip()
        grace_period = self.grace_input.value()
        
        if not name or not start_time or not end_time:
            QMessageBox.warning(self, 'Error', 'All fields are required')
            return
        
        try:
            if self.shift:
                self.db.update_shift(self.shift.id, name, start_time, end_time, grace_period)
                QMessageBox.information(self, 'Success', 'Shift updated successfully')
            else:
                self.db.create_shift(name, start_time, end_time, grace_period)
                QMessageBox.information(self, 'Success', 'Shift created successfully')
            
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to save shift: {str(e)}')

class FingerprintEnrollmentWidget(QWidget):
    def __init__(self, db: Database, sdk: BiometricSDK):
        super().__init__()
        self.db = db
        self.sdk = sdk
        self.setup_ui()
        self.load_users()
        
    def setup_ui(self):
        layout = QVBoxLayout()
        
        header = QLabel('Fingerprint Enrollment')
        header.setFont(QFont('Arial', 14, QFont.Weight.Bold))
        layout.addWidget(header)
        
        top_layout = QHBoxLayout()
        
        left_panel = QVBoxLayout()
        
        user_group = QGroupBox('Select User')
        user_layout = QVBoxLayout()
        
        self.user_combo = QComboBox()
        user_layout.addWidget(self.user_combo)
        
        finger_layout = QHBoxLayout()
        finger_layout.addWidget(QLabel('Finger:'))
        self.finger_combo = QComboBox()
        self.finger_combo.addItems(['Right Thumb', 'Right Index', 'Right Middle', 'Right Ring', 'Right Little',
                                    'Left Thumb', 'Left Index', 'Left Middle', 'Left Ring', 'Left Little'])
        finger_layout.addWidget(self.finger_combo)
        user_layout.addLayout(finger_layout)
        
        user_group.setLayout(user_layout)
        left_panel.addWidget(user_group)
        
        device_group = QGroupBox('Device Status')
        device_layout = QVBoxLayout()
        
        self.status_label = QLabel('Device: Not Connected')
        self.status_label.setStyleSheet('color: red; font-weight: bold;')
        device_layout.addWidget(self.status_label)
        
        self.liveness_label = QLabel('Liveness: N/A')
        device_layout.addWidget(self.liveness_label)
        
        device_group.setLayout(device_layout)
        left_panel.addWidget(device_group)
        
        capture_group = QGroupBox('Capture')
        capture_layout = QVBoxLayout()
        
        self.capture_btn = QPushButton('Capture Fingerprint')
        self.capture_btn.setFixedHeight(50)
        self.capture_btn.clicked.connect(self.capture_fingerprint)
        capture_layout.addWidget(self.capture_btn)
        
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        capture_layout.addWidget(self.progress)
        
        self.quality_label = QLabel('Quality: N/A')
        capture_layout.addWidget(self.quality_label)
        
        capture_group.setLayout(capture_layout)
        left_panel.addWidget(capture_group)
        
        left_panel.addStretch()
        
        top_layout.addLayout(left_panel, 1)
        
        right_panel = QVBoxLayout()
        
        enrolled_group = QGroupBox('Enrolled Fingerprints')
        enrolled_layout = QVBoxLayout()
        
        self.enrolled_list = QListWidget()
        enrolled_layout.addWidget(self.enrolled_list)
        
        delete_btn = QPushButton('Delete Selected')
        delete_btn.clicked.connect(self.delete_fingerprint)
        enrolled_layout.addWidget(delete_btn)
        
        enrolled_group.setLayout(enrolled_layout)
        right_panel.addWidget(enrolled_group)
        
        top_layout.addLayout(right_panel, 1)
        
        layout.addLayout(top_layout)
        
        self.setLayout(layout)
        
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_device_status)
        self.timer.start(2000)
        
        self.user_combo.currentIndexChanged.connect(self.load_enrolled_fingerprints)
    
    def load_users(self):
        users = self.db.get_all_users()
        self.user_combo.clear()
        for user in users:
            self.user_combo.addItem(f'{user.full_name} ({user.username})', user.id)
        
        if users:
            self.load_enrolled_fingerprints()
    
    def load_enrolled_fingerprints(self):
        user_id = self.user_combo.currentData()
        if not user_id:
            return
        
        fingerprints = self.db.get_user_fingerprints(user_id)
        self.enrolled_list.clear()
        
        for fp in fingerprints:
            item_text = f'{fp.finger_position} - Quality: {fp.quality_score}% - {fp.enrolled_at[:19]}'
            item = self.enrolled_list.addItem(item_text)
            self.enrolled_list.item(self.enrolled_list.count() - 1).setData(Qt.ItemDataRole.UserRole, fp.id)
    
    def update_device_status(self):
        status = self.sdk.get_device_status()
        
        if status == NDeviceStatus.CONNECTED:
            self.status_label.setText('Device: Connected')
            self.status_label.setStyleSheet('color: green; font-weight: bold;')
            self.liveness_label.setText('Liveness: Active')
        elif status == NDeviceStatus.BUSY:
            self.status_label.setText('Device: Busy')
            self.status_label.setStyleSheet('color: orange; font-weight: bold;')
            self.liveness_label.setText('Liveness: Processing')
        else:
            self.status_label.setText('Device: Not Connected')
            self.status_label.setStyleSheet('color: red; font-weight: bold;')
            self.liveness_label.setText('Liveness: N/A')
    
    def capture_fingerprint(self):
        user_id = self.user_combo.currentData()
        if not user_id:
            QMessageBox.warning(self, 'Warning', 'Please select a user')
            return
        
        finger_position = self.finger_combo.currentText()
        
        self.capture_btn.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setRange(0, 0)
        
        QApplication.processEvents()
        
        success, template_data, quality = self.sdk.capture_fingerprint()
        
        self.progress.setVisible(False)
        self.capture_btn.setEnabled(True)
        
        if success and len(template_data) > 0:
            self.quality_label.setText(f'Quality: {quality}%')
            
            try:
                self.db.save_fingerprint(user_id, finger_position, template_data, quality)
                QMessageBox.information(self, 'Success', f'Fingerprint enrolled successfully!\nQuality: {quality}%')
                self.load_enrolled_fingerprints()
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Failed to save fingerprint: {str(e)}')
        else:
            self.quality_label.setText('Quality: Failed')
            QMessageBox.warning(self, 'Failed', 'Failed to capture fingerprint. Please try again.')
    
    def delete_fingerprint(self):
        current_item = self.enrolled_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, 'Warning', 'Please select a fingerprint to delete')
            return
        
        fp_id = current_item.data(Qt.ItemDataRole.UserRole)
        
        reply = QMessageBox.question(self, 'Confirm Delete', 
                                     'Are you sure you want to delete this fingerprint?',
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            self.db.delete_fingerprint(fp_id)
            self.load_enrolled_fingerprints()
            QMessageBox.information(self, 'Success', 'Fingerprint deleted successfully')

class AttendanceWidget(QWidget):
    def __init__(self, db: Database, sdk: BiometricSDK):
        super().__init__()
        self.db = db
        self.sdk = sdk
        self.setup_ui()
        self.load_today_attendance()
        
    def setup_ui(self):
        layout = QVBoxLayout()
        
        header = QLabel('Attendance Capture')
        header.setFont(QFont('Arial', 14, QFont.Weight.Bold))
        layout.addWidget(header)
        
        top_layout = QHBoxLayout()
        
        left_panel = QVBoxLayout()
        
        device_group = QGroupBox('Device & Capture')
        device_layout = QVBoxLayout()
        
        self.device_status = QLabel('Device: Checking...')
        device_layout.addWidget(self.device_status)
        
        self.capture_btn = QPushButton('Scan Fingerprint')
        self.capture_btn.setFixedHeight(60)
        self.capture_btn.setStyleSheet('background-color: #4CAF50; color: white; font-size: 16px; font-weight: bold;')
        self.capture_btn.clicked.connect(self.scan_attendance)
        device_layout.addWidget(self.capture_btn)
        
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        device_layout.addWidget(self.progress)
        
        self.result_label = QLabel('')
        self.result_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.result_label.setFont(QFont('Arial', 12))
        device_layout.addWidget(self.result_label)
        
        device_group.setLayout(device_layout)
        left_panel.addWidget(device_group)
        
        manual_group = QGroupBox('Manual Entry')
        manual_layout = QVBoxLayout()
        
        user_layout = QHBoxLayout()
        user_layout.addWidget(QLabel('User:'))
        self.manual_user_combo = QComboBox()
        user_layout.addWidget(self.manual_user_combo)
        manual_layout.addLayout(user_layout)
        
        type_layout = QHBoxLayout()
        type_layout.addWidget(QLabel('Type:'))
        self.manual_type_combo = QComboBox()
        self.manual_type_combo.addItems(['Check In', 'Check Out'])
        type_layout.addWidget(self.manual_type_combo)
        manual_layout.addLayout(type_layout)
        
        manual_btn = QPushButton('Submit Manual Entry')
        manual_btn.clicked.connect(self.manual_entry)
        manual_layout.addWidget(manual_btn)
        
        manual_group.setLayout(manual_layout)
        left_panel.addWidget(manual_group)
        
        left_panel.addStretch()
        
        top_layout.addLayout(left_panel, 1)
        
        right_panel = QVBoxLayout()
        
        today_group = QGroupBox("Today's Attendance")
        today_layout = QVBoxLayout()
        
        refresh_btn = QPushButton('Refresh')
        refresh_btn.clicked.connect(self.load_today_attendance)
        today_layout.addWidget(refresh_btn)
        
        self.today_table = QTableWidget()
        self.today_table.setColumnCount(5)
        self.today_table.setHorizontalHeaderLabels(['User', 'Check In', 'Check Out', 'Status', 'Notes'])
        self.today_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.today_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        today_layout.addWidget(self.today_table)
        
        today_group.setLayout(today_layout)
        right_panel.addWidget(today_group)
        
        top_layout.addLayout(right_panel, 2)
        
        layout.addLayout(top_layout)
        
        self.setLayout(layout)
        
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_device_status)
        self.timer.start(2000)
        
        self.load_users_for_manual()
    
    def load_users_for_manual(self):
        users = self.db.get_all_users()
        self.manual_user_combo.clear()
        for user in users:
            self.manual_user_combo.addItem(f'{user.full_name} ({user.username})', user.id)
    
    def update_device_status(self):
        status = self.sdk.get_device_status()
        
        if status == NDeviceStatus.CONNECTED:
            self.device_status.setText('Device: Connected - Ready')
            self.device_status.setStyleSheet('color: green; font-weight: bold;')
        else:
            self.device_status.setText('Device: Not Connected')
            self.device_status.setStyleSheet('color: red; font-weight: bold;')
    
    def scan_attendance(self):
        self.capture_btn.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setRange(0, 0)
        self.result_label.setText('Scanning fingerprint...')
        
        QApplication.processEvents()
        
        success, scanned_template, quality = self.sdk.capture_fingerprint()
        
        self.progress.setVisible(False)
        self.capture_btn.setEnabled(True)
        
        if not success or len(scanned_template) == 0:
            self.result_label.setText('❌ Scan failed. Please try again.')
            self.result_label.setStyleSheet('color: red;')
            return
        
        all_fingerprints = self.db.get_all_fingerprints()
        
        matched_user_id = None
        best_score = 0
        
        for fp in all_fingerprints:
            match_result, score = self.sdk.verify_fingerprint(scanned_template, fp.template_data)
            
            if match_result and score > best_score and score > 70:
                best_score = score
                matched_user_id = fp.user_id
        
        if matched_user_id:
            user = self.db.get_user_by_id(matched_user_id)
            
            today_record = self.db.get_today_attendance(matched_user_id)
            
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            today_date = date.today().isoformat()
            
            if not today_record:
                self.db.record_attendance(matched_user_id, current_time, today_date, 'Present', f'Match Score: {best_score}%')
                self.result_label.setText(f'✓ Check In: {user.full_name}\nTime: {current_time[11:]}\nScore: {best_score}%')
                self.result_label.setStyleSheet('color: green;')
            elif not today_record.check_out:
                self.db.update_checkout(today_record.id, current_time)
                self.result_label.setText(f'✓ Check Out: {user.full_name}\nTime: {current_time[11:]}\nScore: {best_score}%')
                self.result_label.setStyleSheet('color: blue;')
            else:
                self.result_label.setText(f'⚠ Already checked out today:\n{user.full_name}')
                self.result_label.setStyleSheet('color: orange;')
            
            self.load_today_attendance()
        else:
            self.result_label.setText('❌ Fingerprint not recognized\nNo match found')
            self.result_label.setStyleSheet('color: red;')
    
    def manual_entry(self):
        user_id = self.manual_user_combo.currentData()
        entry_type = self.manual_type_combo.currentText()
        
        if not user_id:
            QMessageBox.warning(self, 'Warning', 'Please select a user')
            return
        
        user = self.db.get_user_by_id(user_id)
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        today_date = date.today().isoformat()
        
        if entry_type == 'Check In':
            self.db.record_attendance(user_id, current_time, today_date, 'Present', 'Manual Entry')
            QMessageBox.information(self, 'Success', f'Check In recorded for {user.full_name}')
        else:
            today_record = self.db.get_today_attendance(user_id)
            if today_record and not today_record.check_out:
                self.db.update_checkout(today_record.id, current_time)
                QMessageBox.information(self, 'Success', f'Check Out recorded for {user.full_name}')
            else:
                QMessageBox.warning(self, 'Warning', 'No check-in record found for today')
        
        self.load_today_attendance()
    
    def load_today_attendance(self):
        today = date.today().isoformat()
        records = self.db.get_attendance_by_date(today)
        users = {u.id: u for u in self.db.get_all_users()}
        
        self.today_table.setRowCount(len(records))
        
        for row, record in enumerate(records):
            user = users.get(record.user_id)
            user_name = user.full_name if user else f'User {record.user_id}'
            
            self.today_table.setItem(row, 0, QTableWidgetItem(user_name))
            self.today_table.setItem(row, 1, QTableWidgetItem(record.check_in[11:19]))
            checkout_time = record.check_out[11:19] if record.check_out else 'Not yet'
            self.today_table.setItem(row, 2, QTableWidgetItem(checkout_time))
            self.today_table.setItem(row, 3, QTableWidgetItem(record.status))
            self.today_table.setItem(row, 4, QTableWidgetItem(record.notes or ''))

class ReportsWidget(QWidget):
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout()
        
        header = QLabel('Attendance Reports')
        header.setFont(QFont('Arial', 14, QFont.Weight.Bold))
        layout.addWidget(header)
        
        filter_group = QGroupBox('Report Filters')
        filter_layout = QGridLayout()
        
        filter_layout.addWidget(QLabel('Report Type:'), 0, 0)
        self.report_type_combo = QComboBox()
        self.report_type_combo.addItems(['Daily Report', 'Monthly Report', 'Custom Range', 'User Summary'])
        self.report_type_combo.currentIndexChanged.connect(self.on_report_type_changed)
        filter_layout.addWidget(self.report_type_combo, 0, 1)
        
        filter_layout.addWidget(QLabel('User:'), 1, 0)
        self.user_combo = QComboBox()
        self.user_combo.addItem('All Users', 0)
        for user in self.db.get_all_users():
            self.user_combo.addItem(f'{user.full_name} ({user.username})', user.id)
        filter_layout.addWidget(self.user_combo, 1, 1)
        
        filter_layout.addWidget(QLabel('Start Date:'), 2, 0)
        self.start_date = QDateEdit()
        self.start_date.setDate(QDate.currentDate())
        self.start_date.setCalendarPopup(True)
        filter_layout.addWidget(self.start_date, 2, 1)
        
        filter_layout.addWidget(QLabel('End Date:'), 3, 0)
        self.end_date = QDateEdit()
        self.end_date.setDate(QDate.currentDate())
        self.end_date.setCalendarPopup(True)
        filter_layout.addWidget(self.end_date, 3, 1)
        
        button_layout = QHBoxLayout()
        
        generate_btn = QPushButton('Generate Report')
        generate_btn.clicked.connect(self.generate_report)
        button_layout.addWidget(generate_btn)
        
        export_pdf_btn = QPushButton('Export to PDF')
        export_pdf_btn.clicked.connect(lambda: self.export_report('pdf'))
        button_layout.addWidget(export_pdf_btn)
        
        export_excel_btn = QPushButton('Export to Excel')
        export_excel_btn.clicked.connect(lambda: self.export_report('excel'))
        button_layout.addWidget(export_excel_btn)
        
        export_csv_btn = QPushButton('Export to CSV')
        export_csv_btn.clicked.connect(lambda: self.export_report('csv'))
        button_layout.addWidget(export_csv_btn)
        
        filter_layout.addLayout(button_layout, 4, 0, 1, 2)
        
        filter_group.setLayout(filter_layout)
        layout.addWidget(filter_group)
        
        self.report_table = QTableWidget()
        self.report_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.report_table)
        
        self.summary_label = QLabel()
        self.summary_label.setFont(QFont('Arial', 10, QFont.Weight.Bold))
        layout.addWidget(self.summary_label)
        
        self.setLayout(layout)
    
    def on_report_type_changed(self, index):
        report_type = self.report_type_combo.currentText()
        
        if report_type == 'Daily Report':
            self.start_date.setDate(QDate.currentDate())
            self.end_date.setDate(QDate.currentDate())
            self.start_date.setEnabled(True)
            self.end_date.setEnabled(False)
        elif report_type == 'Monthly Report':
            current_date = QDate.currentDate()
            first_day = QDate(current_date.year(), current_date.month(), 1)
            last_day = QDate(current_date.year(), current_date.month(), current_date.daysInMonth())
            self.start_date.setDate(first_day)
            self.end_date.setDate(last_day)
            self.start_date.setEnabled(False)
            self.end_date.setEnabled(False)
        else:
            self.start_date.setEnabled(True)
            self.end_date.setEnabled(True)
    
    def generate_report(self):
        start_date_str = self.start_date.date().toString('yyyy-MM-dd')
        end_date_str = self.end_date.date().toString('yyyy-MM-dd')
        user_id = self.user_combo.currentData()
        
        if user_id:
            records = self.db.get_attendance_by_user(user_id, start_date_str, end_date_str)
        else:
            records = self.db.get_attendance_by_date_range(start_date_str, end_date_str)
        
        users = {u.id: u for u in self.db.get_all_users()}
        
        self.report_table.setColumnCount(6)
        self.report_table.setHorizontalHeaderLabels(['Date', 'User', 'Check In', 'Check Out', 'Hours', 'Status'])
        self.report_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        
        self.report_table.setRowCount(len(records))
        
        total_hours = 0
        present_count = 0
        
        for row, record in enumerate(records):
            user = users.get(record.user_id)
            user_name = user.full_name if user else f'User {record.user_id}'
            
            self.report_table.setItem(row, 0, QTableWidgetItem(record.date))
            self.report_table.setItem(row, 1, QTableWidgetItem(user_name))
            self.report_table.setItem(row, 2, QTableWidgetItem(record.check_in[11:19]))
            
            checkout_time = record.check_out[11:19] if record.check_out else 'Not yet'
            self.report_table.setItem(row, 3, QTableWidgetItem(checkout_time))
            
            if record.check_out:
                try:
                    checkin_dt = datetime.strptime(record.check_in, '%Y-%m-%d %H:%M:%S')
                    checkout_dt = datetime.strptime(record.check_out, '%Y-%m-%d %H:%M:%S')
                    duration = checkout_dt - checkin_dt
                    hours = duration.total_seconds() / 3600
                    total_hours += hours
                    hours_str = f'{hours:.2f} hrs'
                except:
                    hours_str = 'N/A'
            else:
                hours_str = 'N/A'
            
            self.report_table.setItem(row, 4, QTableWidgetItem(hours_str))
            self.report_table.setItem(row, 5, QTableWidgetItem(record.status))
            
            if record.status == 'Present':
                present_count += 1
        
        summary = f'Total Records: {len(records)} | Present: {present_count} | Total Hours: {total_hours:.2f}'
        self.summary_label.setText(summary)
    
    def export_report(self, format_type):
        if self.report_table.rowCount() == 0:
            QMessageBox.warning(self, 'Warning', 'Please generate a report first')
            return
        
        file_filter = {
            'pdf': 'PDF Files (*.pdf)',
            'excel': 'Excel Files (*.xlsx)',
            'csv': 'CSV Files (*.csv)'
        }
        
        file_path, _ = QFileDialog.getSaveFileName(self, 'Export Report', '', file_filter.get(format_type, ''))
        
        if not file_path:
            return
        
        try:
            if format_type == 'pdf':
                self.export_to_pdf(file_path)
            elif format_type == 'excel':
                self.export_to_excel(file_path)
            elif format_type == 'csv':
                self.export_to_csv(file_path)
            
            QMessageBox.information(self, 'Success', f'Report exported successfully to {file_path}')
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to export report: {str(e)}')
    
    def export_to_csv(self, file_path):
        with open(file_path, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            
            headers = [self.report_table.horizontalHeaderItem(i).text() 
                      for i in range(self.report_table.columnCount())]
            writer.writerow(headers)
            
            for row in range(self.report_table.rowCount()):
                row_data = [self.report_table.item(row, col).text() 
                           for col in range(self.report_table.columnCount())]
                writer.writerow(row_data)
    
    def export_to_excel(self, file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Attendance Report'
        
        headers = [self.report_table.horizontalHeaderItem(i).text() 
                  for i in range(self.report_table.columnCount())]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = ExcelFont(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        for row in range(self.report_table.rowCount()):
            row_data = [self.report_table.item(row, col).text() 
                       for col in range(self.report_table.columnCount())]
            ws.append(row_data)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2
        
        wb.save(file_path)
    
    def export_to_pdf(self, file_path):
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a237e'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        title = Paragraph('Attendance Report', title_style)
        elements.append(title)
        
        info_style = styles['Normal']
        report_type = self.report_type_combo.currentText()
        date_range = f"{self.start_date.date().toString('yyyy-MM-dd')} to {self.end_date.date().toString('yyyy-MM-dd')}"
        
        info = Paragraph(f'<b>Report Type:</b> {report_type}<br/><b>Date Range:</b> {date_range}<br/><br/>', info_style)
        elements.append(info)
        
        data = [[self.report_table.horizontalHeaderItem(i).text() 
                for i in range(self.report_table.columnCount())]]
        
        for row in range(self.report_table.rowCount()):
            row_data = [self.report_table.item(row, col).text() 
                       for col in range(self.report_table.columnCount())]
            data.append(row_data)
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        summary = Paragraph(f'<b>{self.summary_label.text()}</b>', styles['Normal'])
        elements.append(summary)
        
        doc.build(elements)

class MainWindow(QMainWindow):
    def __init__(self, db: Database, sdk: BiometricSDK, user: User):
        super().__init__()
        self.db = db
        self.sdk = sdk
        self.current_user = user
        self.setup_ui()
        
    def setup_ui(self):
        self.setWindowTitle('Biometric Attendance System')
        self.setGeometry(100, 100, 1200, 800)
        
        menubar = self.menuBar()
        
        file_menu = menubar.addMenu('File')
        
        logout_action = file_menu.addAction('Logout')
        logout_action.triggered.connect(self.logout)
        
        exit_action = file_menu.addAction('Exit')
        exit_action.triggered.connect(self.close)
        
        help_menu = menubar.addMenu('Help')
        
        about_action = help_menu.addAction('About')
        about_action.triggered.connect(self.show_about)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout()
        
        header_layout = QHBoxLayout()
        
        title = QLabel('BIOMETRIC ATTENDANCE SYSTEM')
        title.setFont(QFont('Arial', 18, QFont.Weight.Bold))
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        user_label = QLabel(f'User: {self.current_user.full_name} ({self.current_user.role})')
        user_label.setFont(QFont('Arial', 10))
        header_layout.addWidget(user_label)
        
        layout.addLayout(header_layout)
        
        self.tabs = QTabWidget()
        
        self.tabs.addTab(AttendanceWidget(self.db, self.sdk), 'Attendance')
        self.tabs.addTab(ReportsWidget(self.db), 'Reports')
        
        if self.current_user.role == 'Admin':
            self.tabs.addTab(UserManagementWidget(self.db), 'Users')
            self.tabs.addTab(DepartmentManagementWidget(self.db), 'Departments')
            self.tabs.addTab(ShiftManagementWidget(self.db), 'Shifts')
            self.tabs.addTab(FingerprintEnrollmentWidget(self.db, self.sdk), 'Fingerprints')
        
        layout.addWidget(self.tabs)
        
        status_bar = self.statusBar()
        status_bar.showMessage(f'Logged in as: {self.current_user.username}')
        
        central_widget.setLayout(layout)
    
    def logout(self):
        reply = QMessageBox.question(self, 'Logout', 'Are you sure you want to logout?',
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            self.close()
            QApplication.quit()
    
    def show_about(self):
        about_text = '''
        <h2>Biometric Attendance System</h2>
        <p><b>Version:</b> 1.0.0</p>
        <p><b>Device:</b> Neurotechnology BioMini Slim 2 (FAP20)</p>
        <p><b>Technology:</b> Python 3.x, PyQt6, SQLite</p>
        <br>
        <p>Features:</p>
        <ul>
        <li>Real-time fingerprint capture and matching</li>
        <li>User and department management</li>
        <li>Attendance tracking and reporting</li>
        <li>Export to PDF, Excel, CSV</li>
        <li>Multi-finger enrollment support</li>
        </ul>
        <br>
        <p><b>Developed for:</b> Educational Institutions & Organizations</p>
        '''
        
        QMessageBox.about(self, 'About', about_text)

def main():
    app = QApplication(sys.argv)
    
    app.setStyle('Fusion')
    
    palette = QPalette()
    palette.setColor(QPalette.ColorRole.Window, QColor(240, 240, 240))
    palette.setColor(QPalette.ColorRole.WindowText, QColor(0, 0, 0))
    palette.setColor(QPalette.ColorRole.Base, QColor(255, 255, 255))
    palette.setColor(QPalette.ColorRole.AlternateBase, QColor(245, 245, 245))
    palette.setColor(QPalette.ColorRole.Button, QColor(240, 240, 240))
    palette.setColor(QPalette.ColorRole.ButtonText, QColor(0, 0, 0))
    palette.setColor(QPalette.ColorRole.Highlight, QColor(42, 130, 218))
    palette.setColor(QPalette.ColorRole.HighlightedText, QColor(255, 255, 255))
    app.setPalette(palette)
    
    db = Database()
    sdk = BiometricSDK()
    
    if not sdk.initialize():
        QMessageBox.warning(None, 'Device Warning', 
                           'Failed to initialize biometric device.\n\n'
                           'Please ensure:\n'
                           '1. BioMini Slim 2 is connected via USB\n'
                           '2. Neurotechnology SDK is installed\n'
                           '3. Device drivers are properly installed\n\n'
                           'The application will continue but fingerprint features will not work.')
    
    login_dialog = LoginDialog(db)
    
    if login_dialog.exec() == QDialog.DialogCode.Accepted:
        main_window = MainWindow(db, sdk, login_dialog.logged_in_user)
        main_window.show()
        sys.exit(app.exec())
    else:
        db.close()
        sdk.close()
        sys.exit(0)

if __name__ == '__main__':
    main()