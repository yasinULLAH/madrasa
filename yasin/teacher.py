import sys
import os
import sqlite3
import hashlib
import base64
import json
import csv
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Any
import ctypes
from ctypes import c_int, c_uint, c_void_p, c_char_p, POINTER, Structure, byref
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QPushButton, QLabel, QLineEdit, QTableWidget, QTableWidgetItem,
                             QComboBox, QDateEdit, QTimeEdit, QTextEdit, QMessageBox, QDialog,
                             QFileDialog, QTabWidget, QSpinBox, QCheckBox, QFrame, QScrollArea,
                             QGridLayout, QStackedWidget, QHeaderView, QDialogButtonBox, QGroupBox)
from PyQt6.QtCore import Qt, QDate, QTime, QTimer, pyqtSignal, QThread
from PyQt6.QtGui import QPixmap, QImage, QIcon, QPalette, QColor, QFont
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class DPFPDD_DEV_INFO(Structure):
    _fields_ = [("size", c_uint), ("name", c_char_p)]

class BiometricDevice:
    def __init__(self):
        self.connected = False
        self.device_handle = None
        self.lib = None
        try:
            if sys.platform == 'win32':
                self.lib = ctypes.CDLL('dpfpdd.dll')
            else:
                self.lib = None
        except:
            self.lib = None
        
    def initialize(self):
        if not self.lib:
            return False
        try:
            init_func = self.lib.dpfpdd_init
            init_func.restype = c_int
            result = init_func()
            if result == 0:
                self.connected = True
                return True
        except:
            pass
        return False
    
    def open_device(self):
        if not self.lib or not self.connected:
            return False
        try:
            open_func = self.lib.dpfpdd_open
            open_func.argtypes = [c_char_p, POINTER(c_void_p)]
            open_func.restype = c_int
            handle = c_void_p()
            result = open_func(b"00000000", byref(handle))
            if result == 0:
                self.device_handle = handle
                return True
        except:
            pass
        return False
    
    def capture_fingerprint(self):
        if not self.device_handle:
            return None
        try:
            capture_func = self.lib.dpfpdd_capture
            capture_func.argtypes = [c_void_p, POINTER(c_void_p), POINTER(c_uint), c_uint]
            capture_func.restype = c_int
            data = c_void_p()
            size = c_uint()
            result = capture_func(self.device_handle, byref(data), byref(size), 5000)
            if result == 0 and size.value > 0:
                buffer = ctypes.string_at(data, size.value)
                return base64.b64encode(buffer).decode('utf-8')
        except:
            pass
        return None
    
    def match_fingerprints(self, template1, template2):
        if not self.lib:
            return False
        try:
            data1 = base64.b64decode(template1)
            data2 = base64.b64decode(template2)
            match_func = self.lib.dpfpdd_compare
            match_func.argtypes = [c_void_p, c_uint, c_void_p, c_uint, POINTER(c_uint)]
            match_func.restype = c_int
            score = c_uint()
            result = match_func(data1, len(data1), data2, len(data2), byref(score))
            return result == 0 and score.value > 50000
        except:
            pass
        return False
    
    def close_device(self):
        if self.device_handle and self.lib:
            try:
                close_func = self.lib.dpfpdd_close
                close_func.argtypes = [c_void_p]
                close_func.restype = c_int
                close_func(self.device_handle)
                self.device_handle = None
            except:
                pass
    
    def exit(self):
        self.close_device()
        if self.lib:
            try:
                exit_func = self.lib.dpfpdd_exit
                exit_func.restype = c_int
                exit_func()
                self.connected = False
            except:
                pass

class Database:
    def __init__(self, db_path="biometric_attendance.db"):
        self.db_path = db_path
        self.conn = None
        self.initialize()
    
    def initialize(self):
        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row
        cursor = self.conn.cursor()
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS departments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            code TEXT UNIQUE,
            status TEXT DEFAULT 'Active',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS shifts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            start_time TEXT NOT NULL,
            end_time TEXT NOT NULL,
            grace_time INTEGER DEFAULT 0,
            status TEXT DEFAULT 'Active',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS teachers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            emp_id TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            cnic TEXT UNIQUE,
            phone TEXT,
            email TEXT,
            department_id INTEGER,
            designation TEXT,
            shift_id INTEGER,
            photo TEXT,
            status TEXT DEFAULT 'Active',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (department_id) REFERENCES departments(id),
            FOREIGN KEY (shift_id) REFERENCES shifts(id)
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS fingerprints (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            teacher_id INTEGER NOT NULL,
            finger_name TEXT NOT NULL,
            template TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (teacher_id) REFERENCES teachers(id) ON DELETE CASCADE
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            teacher_id INTEGER NOT NULL,
            date TEXT NOT NULL,
            check_in TEXT,
            check_out TEXT,
            status TEXT,
            is_late INTEGER DEFAULT 0,
            is_early_leave INTEGER DEFAULT 0,
            manual_entry INTEGER DEFAULT 0,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (teacher_id) REFERENCES teachers(id),
            UNIQUE(teacher_id, date)
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS holidays (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            date TEXT NOT NULL,
            type TEXT DEFAULT 'Public',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS leaves (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            teacher_id INTEGER NOT NULL,
            leave_type TEXT NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            reason TEXT,
            status TEXT DEFAULT 'Pending',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (teacher_id) REFERENCES teachers(id)
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            action TEXT NOT NULL,
            details TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        
        cursor.execute("SELECT COUNT(*) FROM users WHERE role='SUPER_ADMIN'")
        if cursor.fetchone()[0] == 0:
            password_hash = hashlib.sha256("admin123".encode()).hexdigest()
            cursor.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                          ("admin", password_hash, "SUPER_ADMIN"))
        
        cursor.execute("SELECT COUNT(*) FROM departments")
        if cursor.fetchone()[0] == 0:
            cursor.execute("INSERT INTO departments (name, code) VALUES (?, ?)", ("Computer Science", "CS"))
            cursor.execute("INSERT INTO departments (name, code) VALUES (?, ?)", ("Mathematics", "MATH"))
            cursor.execute("INSERT INTO departments (name, code) VALUES (?, ?)", ("Physics", "PHY"))
            cursor.execute("INSERT INTO departments (name, code) VALUES (?, ?)", ("Chemistry", "CHEM"))
        
        cursor.execute("SELECT COUNT(*) FROM shifts")
        if cursor.fetchone()[0] == 0:
            cursor.execute("INSERT INTO shifts (name, start_time, end_time, grace_time) VALUES (?, ?, ?, ?)",
                          ("Morning", "08:00", "14:00", 15))
            cursor.execute("INSERT INTO shifts (name, start_time, end_time, grace_time) VALUES (?, ?, ?, ?)",
                          ("Evening", "14:00", "20:00", 15))
            cursor.execute("INSERT INTO shifts (name, start_time, end_time, grace_time) VALUES (?, ?, ?, ?)",
                          ("Full Day", "08:00", "16:00", 15))
            cursor.execute("INSERT INTO shifts (name, start_time, end_time, grace_time) VALUES (?, ?, ?, ?)",
                          ("Night", "20:00", "02:00", 15))
        
        cursor.execute("SELECT value FROM settings WHERE key='institute_name'")
        if not cursor.fetchone():
            cursor.execute("INSERT INTO settings (key, value) VALUES (?, ?)", ("institute_name", "Excellence Institute"))
            cursor.execute("INSERT INTO settings (key, value) VALUES (?, ?)", ("institute_address", "Main Street, City"))
            cursor.execute("INSERT INTO settings (key, value) VALUES (?, ?)", ("theme", "Light"))
        
        self.conn.commit()
    
    def execute(self, query, params=()):
        cursor = self.conn.cursor()
        cursor.execute(query, params)
        self.conn.commit()
        return cursor
    
    def fetchall(self, query, params=()):
        cursor = self.conn.cursor()
        cursor.execute(query, params)
        return cursor.fetchall()
    
    def fetchone(self, query, params=()):
        cursor = self.conn.cursor()
        cursor.execute(query, params)
        return cursor.fetchone()
    
    def backup(self, filepath):
        backup_conn = sqlite3.connect(filepath)
        self.conn.backup(backup_conn)
        backup_conn.close()
    
    def restore(self, filepath):
        self.conn.close()
        os.remove(self.db_path)
        restore_conn = sqlite3.connect(filepath)
        self.conn = sqlite3.connect(self.db_path)
        restore_conn.backup(self.conn)
        restore_conn.close()
        self.conn.row_factory = sqlite3.Row

class LoginDialog(QDialog):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.user_data = None
        self.setWindowTitle("Login")
        self.setFixedSize(400, 300)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(20)
        
        title = QLabel("Biometric Attendance System")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("font-size: 20px; font-weight: bold; margin: 20px;")
        layout.addWidget(title)
        
        form_layout = QVBoxLayout()
        
        username_label = QLabel("Username:")
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("Enter username")
        form_layout.addWidget(username_label)
        form_layout.addWidget(self.username_input)
        
        password_label = QLabel("Password:")
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_input.setPlaceholderText("Enter password")
        self.password_input.returnPressed.connect(self.login)
        form_layout.addWidget(password_label)
        form_layout.addWidget(self.password_input)
        
        layout.addLayout(form_layout)
        
        login_btn = QPushButton("Login")
        login_btn.setStyleSheet("padding: 10px; font-size: 14px;")
        login_btn.clicked.connect(self.login)
        layout.addWidget(login_btn)
        
        self.setLayout(layout)
    
    def login(self):
        username = self.username_input.text().strip()
        password = self.password_input.text()
        
        if not username or not password:
            QMessageBox.warning(self, "Error", "Please enter username and password")
            return
        
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        user = self.db.fetchone("SELECT * FROM users WHERE username=? AND password=?", 
                               (username, password_hash))
        
        if user:
            self.user_data = dict(user)
            self.accept()
        else:
            QMessageBox.warning(self, "Error", "Invalid username or password")

class FingerEnrollDialog(QDialog):
    def __init__(self, device, teacher_id, db):
        super().__init__()
        self.device = device
        self.teacher_id = teacher_id
        self.db = db
        self.template = None
        self.setWindowTitle("Fingerprint Enrollment")
        self.setFixedSize(400, 300)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        
        self.status_label = QLabel("Select finger and click Capture")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_label.setStyleSheet("font-size: 14px; padding: 10px;")
        layout.addWidget(self.status_label)
        
        self.finger_combo = QComboBox()
        self.finger_combo.addItems(["Right Thumb", "Right Index", "Right Middle", "Right Ring", "Right Pinky",
                                   "Left Thumb", "Left Index", "Left Middle", "Left Ring", "Left Pinky"])
        layout.addWidget(self.finger_combo)
        
        self.capture_btn = QPushButton("Capture Fingerprint")
        self.capture_btn.clicked.connect(self.capture)
        layout.addWidget(self.capture_btn)
        
        button_layout = QHBoxLayout()
        self.save_btn = QPushButton("Save")
        self.save_btn.setEnabled(False)
        self.save_btn.clicked.connect(self.save)
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def capture(self):
        self.status_label.setText("Place finger on scanner...")
        QApplication.processEvents()
        
        template = self.device.capture_fingerprint()
        if template:
            self.template = template
            self.status_label.setText("Fingerprint captured successfully!")
            self.status_label.setStyleSheet("color: green; font-size: 14px; padding: 10px;")
            self.save_btn.setEnabled(True)
        else:
            self.status_label.setText("Capture failed. Try again.")
            self.status_label.setStyleSheet("color: red; font-size: 14px; padding: 10px;")
    
    def save(self):
        finger_name = self.finger_combo.currentText()
        existing = self.db.fetchone("SELECT id FROM fingerprints WHERE teacher_id=? AND finger_name=?",
                                   (self.teacher_id, finger_name))
        if existing:
            reply = QMessageBox.question(self, "Confirm", "This finger already exists. Replace?",
                                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                self.db.execute("DELETE FROM fingerprints WHERE id=?", (existing['id'],))
            else:
                return
        
        self.db.execute("INSERT INTO fingerprints (teacher_id, finger_name, template) VALUES (?, ?, ?)",
                       (self.teacher_id, finger_name, self.template))
        QMessageBox.information(self, "Success", "Fingerprint saved successfully!")
        self.accept()

class TeacherFormDialog(QDialog):
    def __init__(self, db, device, teacher_id=None):
        super().__init__()
        self.db = db
        self.device = device
        self.teacher_id = teacher_id
        self.photo_path = None
        self.setWindowTitle("Add Teacher" if not teacher_id else "Edit Teacher")
        self.setMinimumSize(600, 700)
        self.setup_ui()
        if teacher_id:
            self.load_teacher_data()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        form_layout = QVBoxLayout()
        
        self.emp_id_input = QLineEdit()
        form_layout.addWidget(QLabel("Employee ID:"))
        form_layout.addWidget(self.emp_id_input)
        
        self.name_input = QLineEdit()
        form_layout.addWidget(QLabel("Name:"))
        form_layout.addWidget(self.name_input)
        
        self.cnic_input = QLineEdit()
        form_layout.addWidget(QLabel("CNIC:"))
        form_layout.addWidget(self.cnic_input)
        
        self.phone_input = QLineEdit()
        form_layout.addWidget(QLabel("Phone:"))
        form_layout.addWidget(self.phone_input)
        
        self.email_input = QLineEdit()
        form_layout.addWidget(QLabel("Email:"))
        form_layout.addWidget(self.email_input)
        
        self.dept_combo = QComboBox()
        departments = self.db.fetchall("SELECT id, name FROM departments WHERE status='Active'")
        for dept in departments:
            self.dept_combo.addItem(dept['name'], dept['id'])
        form_layout.addWidget(QLabel("Department:"))
        form_layout.addWidget(self.dept_combo)
        
        self.designation_input = QLineEdit()
        form_layout.addWidget(QLabel("Designation:"))
        form_layout.addWidget(self.designation_input)
        
        self.shift_combo = QComboBox()
        shifts = self.db.fetchall("SELECT id, name FROM shifts WHERE status='Active'")
        for shift in shifts:
            self.shift_combo.addItem(shift['name'], shift['id'])
        form_layout.addWidget(QLabel("Shift:"))
        form_layout.addWidget(self.shift_combo)
        
        self.status_combo = QComboBox()
        self.status_combo.addItems(["Active", "Inactive"])
        form_layout.addWidget(QLabel("Status:"))
        form_layout.addWidget(self.status_combo)
        
        photo_btn = QPushButton("Upload Photo")
        photo_btn.clicked.connect(self.upload_photo)
        form_layout.addWidget(photo_btn)
        
        scroll_widget.setLayout(form_layout)
        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)
        
        button_layout = QHBoxLayout()
        save_btn = QPushButton("Save")
        save_btn.clicked.connect(self.save)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def upload_photo(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Photo", "", "Images (*.png *.jpg *.jpeg)")
        if file_path:
            self.photo_path = file_path
            QMessageBox.information(self, "Success", "Photo selected")
    
    def load_teacher_data(self):
        teacher = self.db.fetchone("SELECT * FROM teachers WHERE id=?", (self.teacher_id,))
        if teacher:
            self.emp_id_input.setText(teacher['emp_id'])
            self.name_input.setText(teacher['name'])
            self.cnic_input.setText(teacher['cnic'] or "")
            self.phone_input.setText(teacher['phone'] or "")
            self.email_input.setText(teacher['email'] or "")
            
            for i in range(self.dept_combo.count()):
                if self.dept_combo.itemData(i) == teacher['department_id']:
                    self.dept_combo.setCurrentIndex(i)
                    break
            
            self.designation_input.setText(teacher['designation'] or "")
            
            for i in range(self.shift_combo.count()):
                if self.shift_combo.itemData(i) == teacher['shift_id']:
                    self.shift_combo.setCurrentIndex(i)
                    break
            
            self.status_combo.setCurrentText(teacher['status'])
            self.photo_path = teacher['photo']
    
    def save(self):
        emp_id = self.emp_id_input.text().strip()
        name = self.name_input.text().strip()
        
        if not emp_id or not name:
            QMessageBox.warning(self, "Error", "Employee ID and Name are required")
            return
        
        photo_data = None
        if self.photo_path and os.path.exists(self.photo_path):
            with open(self.photo_path, 'rb') as f:
                photo_data = base64.b64encode(f.read()).decode('utf-8')
        
        data = (
            emp_id,
            name,
            self.cnic_input.text().strip() or None,
            self.phone_input.text().strip() or None,
            self.email_input.text().strip() or None,
            self.dept_combo.currentData(),
            self.designation_input.text().strip() or None,
            self.shift_combo.currentData(),
            photo_data,
            self.status_combo.currentText()
        )
        
        try:
            if self.teacher_id:
                self.db.execute('''UPDATE teachers SET emp_id=?, name=?, cnic=?, phone=?, email=?, 
                                  department_id=?, designation=?, shift_id=?, photo=?, status=? WHERE id=?''',
                              data + (self.teacher_id,))
            else:
                cursor = self.db.execute('''INSERT INTO teachers (emp_id, name, cnic, phone, email, 
                                           department_id, designation, shift_id, photo, status) 
                                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', data)
                self.teacher_id = cursor.lastrowid
            
            QMessageBox.information(self, "Success", "Teacher saved successfully!")
            self.accept()
        except sqlite3.IntegrityError as e:
            QMessageBox.warning(self, "Error", f"Duplicate entry: {str(e)}")

class AttendanceWidget(QWidget):
    def __init__(self, db, device, user_data):
        super().__init__()
        self.db = db
        self.device = device
        self.user_data = user_data
        self.setup_ui()
        self.update_status()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        
        header = QLabel("Mark Attendance")
        header.setStyleSheet("font-size: 24px; font-weight: bold; margin: 10px;")
        layout.addWidget(header)
        
        status_layout = QHBoxLayout()
        self.device_status = QLabel("Device: Disconnected")
        self.device_status.setStyleSheet("color: red; font-size: 14px;")
        status_layout.addWidget(self.device_status)
        
        self.date_label = QLabel(f"Date: {date.today().strftime('%Y-%m-%d')}")
        self.date_label.setStyleSheet("font-size: 14px;")
        status_layout.addWidget(self.date_label)
        status_layout.addStretch()
        layout.addLayout(status_layout)
        
        self.info_label = QLabel("Place finger on scanner...")
        self.info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.info_label.setStyleSheet("font-size: 18px; padding: 20px; border: 2px dashed #ccc; margin: 20px;")
        layout.addWidget(self.info_label)
        
        btn_layout = QHBoxLayout()
        self.scan_btn = QPushButton("Scan Fingerprint")
        self.scan_btn.setStyleSheet("padding: 15px; font-size: 16px;")
        self.scan_btn.clicked.connect(self.scan_fingerprint)
        btn_layout.addWidget(self.scan_btn)
        
        self.manual_btn = QPushButton("Manual Entry")
        self.manual_btn.setStyleSheet("padding: 15px; font-size: 16px;")
        self.manual_btn.clicked.connect(self.manual_entry)
        btn_layout.addWidget(self.manual_btn)
        layout.addLayout(btn_layout)
        
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Name", "Employee ID", "Check-In", "Check-Out", "Status", "Notes"])
        self.table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.table)
        
        self.load_today_attendance()
        
        timer = QTimer(self)
        timer.timeout.connect(self.update_status)
        timer.start(5000)
        
        self.setLayout(layout)
    
    def update_status(self):
        if self.device.connected:
            self.device_status.setText("Device: Connected")
            self.device_status.setStyleSheet("color: green; font-size: 14px;")
        else:
            self.device_status.setText("Device: Disconnected")
            self.device_status.setStyleSheet("color: red; font-size: 14px;")
        
        self.date_label.setText(f"Date: {date.today().strftime('%Y-%m-%d')}")
    
    def scan_fingerprint(self):
        if not self.device.connected:
            QMessageBox.warning(self, "Error", "Biometric device not connected")
            return
        
        self.info_label.setText("Scanning... Place finger on scanner")
        QApplication.processEvents()
        
        captured_template = self.device.capture_fingerprint()
        if not captured_template:
            self.info_label.setText("Capture failed. Try again.")
            return
        
        teachers = self.db.fetchall("SELECT id, name, emp_id FROM teachers WHERE status='Active'")
        matched_teacher = None
        
        for teacher in teachers:
            fingerprints = self.db.fetchall("SELECT template FROM fingerprints WHERE teacher_id=?", (teacher['id'],))
            for fp in fingerprints:
                if self.device.match_fingerprints(captured_template, fp['template']):
                    matched_teacher = teacher
                    break
            if matched_teacher:
                break
        
        if matched_teacher:
            self.mark_attendance(matched_teacher['id'], matched_teacher['name'], matched_teacher['emp_id'])
        else:
            self.info_label.setText("Fingerprint not recognized")
            self.info_label.setStyleSheet("color: red; font-size: 18px; padding: 20px; border: 2px dashed #ccc; margin: 20px;")
    
    def mark_attendance(self, teacher_id, name, emp_id):
        today = date.today().strftime('%Y-%m-%d')
        current_time = datetime.now().strftime('%H:%M:%S')
        
        existing = self.db.fetchone("SELECT * FROM attendance WHERE teacher_id=? AND date=?", (teacher_id, today))
        
        teacher = self.db.fetchone("SELECT shift_id FROM teachers WHERE id=?", (teacher_id,))
        shift = self.db.fetchone("SELECT start_time, end_time, grace_time FROM shifts WHERE id=?", (teacher['shift_id'],))
        
        if existing:
            if existing['check_out']:
                self.info_label.setText(f"{name} already checked out today")
                return
            
            early_leave = self.is_early_leave(current_time, shift['end_time'], shift['grace_time'])
            self.db.execute("UPDATE attendance SET check_out=?, is_early_leave=? WHERE id=?",
                          (current_time, 1 if early_leave else 0, existing['id']))
            status = "Checked Out"
            if early_leave:
                status += " (Early)"
        else:
            is_late = self.is_late(current_time, shift['start_time'], shift['grace_time'])
            self.db.execute('''INSERT INTO attendance (teacher_id, date, check_in, status, is_late) 
                              VALUES (?, ?, ?, ?, ?)''',
                          (teacher_id, today, current_time, 'Present', 1 if is_late else 0))
            status = "Checked In"
            if is_late:
                status += " (Late)"
        
        self.info_label.setText(f"{name} - {status} at {current_time}")
        self.info_label.setStyleSheet("color: green; font-size: 18px; padding: 20px; border: 2px dashed #ccc; margin: 20px;")
        self.load_today_attendance()
    
    def is_late(self, check_in, start_time, grace_time):
        check_in_dt = datetime.strptime(check_in, '%H:%M:%S')
        start_dt = datetime.strptime(start_time, '%H:%M')
        grace_dt = start_dt + timedelta(minutes=grace_time)
        return check_in_dt.time() > grace_dt.time()
    
    def is_early_leave(self, check_out, end_time, grace_time):
        check_out_dt = datetime.strptime(check_out, '%H:%M:%S')
        end_dt = datetime.strptime(end_time, '%H:%M')
        grace_dt = end_dt - timedelta(minutes=grace_time)
        return check_out_dt.time() < grace_dt.time()
    
    def manual_entry(self):
        if self.user_data['role'] not in ['SUPER_ADMIN', 'ADMIN']:
            QMessageBox.warning(self, "Access Denied", "Only admins can make manual entries")
            return
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Manual Attendance Entry")
        dialog.setMinimumSize(400, 300)
        
        layout = QVBoxLayout()
        
        teacher_combo = QComboBox()
        teachers = self.db.fetchall("SELECT id, name, emp_id FROM teachers WHERE status='Active'")
        for t in teachers:
            teacher_combo.addItem(f"{t['name']} ({t['emp_id']})", t['id'])
        layout.addWidget(QLabel("Teacher:"))
        layout.addWidget(teacher_combo)
        
        date_edit = QDateEdit()
        date_edit.setDate(QDate.currentDate())
        layout.addWidget(QLabel("Date:"))
        layout.addWidget(date_edit)
        
        checkin_edit = QTimeEdit()
        checkin_edit.setTime(QTime.currentTime())
        layout.addWidget(QLabel("Check-In:"))
        layout.addWidget(checkin_edit)
        
        checkout_edit = QTimeEdit()
        checkout_edit.setTime(QTime.currentTime())
        layout.addWidget(QLabel("Check-Out:"))
        layout.addWidget(checkout_edit)
        
        notes_edit = QLineEdit()
        layout.addWidget(QLabel("Notes:"))
        layout.addWidget(notes_edit)
        
        def save_manual():
            teacher_id = teacher_combo.currentData()
            att_date = date_edit.date().toString('yyyy-MM-dd')
            check_in = checkin_edit.time().toString('HH:mm:ss')
            check_out = checkout_edit.time().toString('HH:mm:ss')
            notes = notes_edit.text()
            
            existing = self.db.fetchone("SELECT id FROM attendance WHERE teacher_id=? AND date=?",
                                       (teacher_id, att_date))
            
            if existing:
                self.db.execute('''UPDATE attendance SET check_in=?, check_out=?, manual_entry=1, notes=? 
                                  WHERE id=?''', (check_in, check_out, notes, existing['id']))
            else:
                self.db.execute('''INSERT INTO attendance (teacher_id, date, check_in, check_out, 
                                  status, manual_entry, notes) VALUES (?, ?, ?, ?, ?, ?, ?)''',
                              (teacher_id, att_date, check_in, check_out, 'Present', 1, notes))
            
            QMessageBox.information(dialog, "Success", "Manual entry saved")
            dialog.accept()
            self.load_today_attendance()
        
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("Save")
        save_btn.clicked.connect(save_manual)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def load_today_attendance(self):
        today = date.today().strftime('%Y-%m-%d')
        records = self.db.fetchall('''SELECT t.name, t.emp_id, a.check_in, a.check_out, a.status, a.notes,
                                     a.is_late, a.is_early_leave FROM attendance a
                                     JOIN teachers t ON a.teacher_id = t.id
                                     WHERE a.date = ? ORDER BY a.check_in DESC''', (today,))
        
        self.table.setRowCount(len(records))
        for i, rec in enumerate(records):
            self.table.setItem(i, 0, QTableWidgetItem(rec['name']))
            self.table.setItem(i, 1, QTableWidgetItem(rec['emp_id']))
            self.table.setItem(i, 2, QTableWidgetItem(rec['check_in'] or '-'))
            self.table.setItem(i, 3, QTableWidgetItem(rec['check_out'] or '-'))
            
            status = rec['status']
            if rec['is_late']:
                status += " (Late)"
            if rec['is_early_leave']:
                status += " (Early)"
            self.table.setItem(i, 4, QTableWidgetItem(status))
            self.table.setItem(i, 5, QTableWidgetItem(rec['notes'] or ''))

class TeachersWidget(QWidget):
    def __init__(self, db, device, user_data):
        super().__init__()
        self.db = db
        self.device = device
        self.user_data = user_data
        self.setup_ui()
        self.load_data()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        
        header_layout = QHBoxLayout()
        header = QLabel("Teachers Management")
        header.setStyleSheet("font-size: 24px; font-weight: bold;")
        header_layout.addWidget(header)
        header_layout.addStretch()
        
        add_btn = QPushButton("Add Teacher")
        add_btn.clicked.connect(self.add_teacher)
        header_layout.addWidget(add_btn)
        
        layout.addLayout(header_layout)
        
        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by name, employee ID, or CNIC")
        self.search_input.textChanged.connect(self.load_data)
        search_layout.addWidget(self.search_input)
        layout.addLayout(search_layout)
        
        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels(["ID", "Emp ID", "Name", "Department", "Designation", 
                                              "Phone", "Status", "Actions", "Fingerprints"])
        self.table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.table)
        
        self.setLayout(layout)
    
    def load_data(self):
        search = self.search_input.text().strip()
        query = '''SELECT t.*, d.name as dept_name FROM teachers t 
                  LEFT JOIN departments d ON t.department_id = d.id WHERE 1=1'''
        params = []
        
        if search:
            query += " AND (t.name LIKE ? OR t.emp_id LIKE ? OR t.cnic LIKE ?)"
            params = [f'%{search}%', f'%{search}%', f'%{search}%']
        
        query += " ORDER BY t.id DESC"
        teachers = self.db.fetchall(query, params)
        
        self.table.setRowCount(len(teachers))
        for i, teacher in enumerate(teachers):
            self.table.setItem(i, 0, QTableWidgetItem(str(teacher['id'])))
            self.table.setItem(i, 1, QTableWidgetItem(teacher['emp_id']))
            self.table.setItem(i, 2, QTableWidgetItem(teacher['name']))
            self.table.setItem(i, 3, QTableWidgetItem(teacher['dept_name'] or '-'))
            self.table.setItem(i, 4, QTableWidgetItem(teacher['designation'] or '-'))
            self.table.setItem(i, 5, QTableWidgetItem(teacher['phone'] or '-'))
            self.table.setItem(i, 6, QTableWidgetItem(teacher['status']))
            
            action_widget = QWidget()
            action_layout = QHBoxLayout()
            action_layout.setContentsMargins(0, 0, 0, 0)
            
            edit_btn = QPushButton("Edit")
            edit_btn.clicked.connect(lambda checked, tid=teacher['id']: self.edit_teacher(tid))
            action_layout.addWidget(edit_btn)
            
            delete_btn = QPushButton("Delete")
            delete_btn.clicked.connect(lambda checked, tid=teacher['id']: self.delete_teacher(tid))
            action_layout.addWidget(delete_btn)
            
            action_widget.setLayout(action_layout)
            self.table.setCellWidget(i, 7, action_widget)
            
            fp_widget = QWidget()
            fp_layout = QHBoxLayout()
            fp_layout.setContentsMargins(0, 0, 0, 0)
            
            enroll_btn = QPushButton("Enroll")
            enroll_btn.clicked.connect(lambda checked, tid=teacher['id']: self.enroll_finger(tid))
            fp_layout.addWidget(enroll_btn)
            
            view_btn = QPushButton("View")
            view_btn.clicked.connect(lambda checked, tid=teacher['id']: self.view_fingerprints(tid))
            fp_layout.addWidget(view_btn)
            
            fp_widget.setLayout(fp_layout)
            self.table.setCellWidget(i, 8, fp_widget)
    
    def add_teacher(self):
        dialog = TeacherFormDialog(self.db, self.device)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_data()
    
    def edit_teacher(self, teacher_id):
        dialog = TeacherFormDialog(self.db, self.device, teacher_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_data()
    
    def delete_teacher(self, teacher_id):
        reply = QMessageBox.question(self, "Confirm Delete", "Delete this teacher?",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.db.execute("DELETE FROM teachers WHERE id=?", (teacher_id,))
            QMessageBox.information(self, "Success", "Teacher deleted")
            self.load_data()
    
    def enroll_finger(self, teacher_id):
        if not self.device.connected:
            QMessageBox.warning(self, "Error", "Biometric device not connected")
            return
        
        dialog = FingerEnrollDialog(self.device, teacher_id, self.db)
        dialog.exec()
    
    def view_fingerprints(self, teacher_id):
        fingerprints = self.db.fetchall("SELECT * FROM fingerprints WHERE teacher_id=?", (teacher_id,))
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Enrolled Fingerprints")
        dialog.setMinimumSize(400, 300)
        
        layout = QVBoxLayout()
        
        table = QTableWidget()
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(["Finger", "Enrolled Date", "Action"])
        table.setRowCount(len(fingerprints))
        
        for i, fp in enumerate(fingerprints):
            table.setItem(i, 0, QTableWidgetItem(fp['finger_name']))
            table.setItem(i, 1, QTableWidgetItem(fp['created_at']))
            
            delete_btn = QPushButton("Delete")
            delete_btn.clicked.connect(lambda checked, fid=fp['id']: self.delete_fingerprint(fid, dialog))
            table.setCellWidget(i, 2, delete_btn)
        
        layout.addWidget(table)
        
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def delete_fingerprint(self, fp_id, parent_dialog):
        reply = QMessageBox.question(parent_dialog, "Confirm", "Delete this fingerprint?",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.db.execute("DELETE FROM fingerprints WHERE id=?", (fp_id,))
            QMessageBox.information(parent_dialog, "Success", "Fingerprint deleted")
            parent_dialog.accept()

class ReportsWidget(QWidget):
    def __init__(self, db, device, user_data):
        super().__init__()
        self.db = db
        self.device = device
        self.user_data = user_data
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        
        header = QLabel("Reports & Analytics")
        header.setStyleSheet("font-size: 24px; font-weight: bold; margin: 10px;")
        layout.addWidget(header)
        
        filters_group = QGroupBox("Filters")
        filters_layout = QGridLayout()
        
        filters_layout.addWidget(QLabel("Report Type:"), 0, 0)
        self.report_type_combo = QComboBox()
        self.report_type_combo.addItems(["Daily Report", "Monthly Report", "Yearly Report", 
                                        "Teacher Report", "Department Report", "Late Report",
                                        "Absent Report"])
        filters_layout.addWidget(self.report_type_combo, 0, 1)
        
        filters_layout.addWidget(QLabel("Start Date:"), 1, 0)
        self.start_date = QDateEdit()
        self.start_date.setDate(QDate.currentDate())
        filters_layout.addWidget(self.start_date, 1, 1)
        
        filters_layout.addWidget(QLabel("End Date:"), 2, 0)
        self.end_date = QDateEdit()
        self.end_date.setDate(QDate.currentDate())
        filters_layout.addWidget(self.end_date, 2, 1)
        
        filters_layout.addWidget(QLabel("Teacher:"), 3, 0)
        self.teacher_combo = QComboBox()
        self.teacher_combo.addItem("All Teachers", None)
        teachers = self.db.fetchall("SELECT id, name, emp_id FROM teachers WHERE status='Active'")
        for t in teachers:
            self.teacher_combo.addItem(f"{t['name']} ({t['emp_id']})", t['id'])
        filters_layout.addWidget(self.teacher_combo, 3, 1)
        
        filters_layout.addWidget(QLabel("Department:"), 4, 0)
        self.dept_combo = QComboBox()
        self.dept_combo.addItem("All Departments", None)
        depts = self.db.fetchall("SELECT id, name FROM departments WHERE status='Active'")
        for d in depts:
            self.dept_combo.addItem(d['name'], d['id'])
        filters_layout.addWidget(self.dept_combo, 4, 1)
        
        generate_btn = QPushButton("Generate Report")
        generate_btn.clicked.connect(self.generate_report)
        filters_layout.addWidget(generate_btn, 5, 0, 1, 2)
        
        filters_group.setLayout(filters_layout)
        layout.addWidget(filters_group)
        
        export_layout = QHBoxLayout()
        
        export_pdf_btn = QPushButton("Export PDF")
        export_pdf_btn.clicked.connect(lambda: self.export_report('pdf'))
        export_layout.addWidget(export_pdf_btn)
        
        export_excel_btn = QPushButton("Export Excel")
        export_excel_btn.clicked.connect(lambda: self.export_report('excel'))
        export_layout.addWidget(export_excel_btn)
        
        export_csv_btn = QPushButton("Export CSV")
        export_csv_btn.clicked.connect(lambda: self.export_report('csv'))
        export_layout.addWidget(export_csv_btn)
        
        layout.addLayout(export_layout)
        
        self.table = QTableWidget()
        layout.addWidget(self.table)
        
        self.setLayout(layout)
    
    def generate_report(self):
        report_type = self.report_type_combo.currentText()
        start = self.start_date.date().toString('yyyy-MM-dd')
        end = self.end_date.date().toString('yyyy-MM-dd')
        teacher_id = self.teacher_combo.currentData()
        dept_id = self.dept_combo.currentData()
        
        query = '''SELECT t.name, t.emp_id, d.name as dept, a.date, a.check_in, a.check_out, 
                  a.status, a.is_late, a.is_early_leave FROM attendance a
                  JOIN teachers t ON a.teacher_id = t.id
                  LEFT JOIN departments d ON t.department_id = d.id
                  WHERE a.date BETWEEN ? AND ?'''
        params = [start, end]
        
        if teacher_id:
            query += " AND t.id = ?"
            params.append(teacher_id)
        
        if dept_id:
            query += " AND t.department_id = ?"
            params.append(dept_id)
        
        if "Late" in report_type:
            query += " AND a.is_late = 1"
        elif "Absent" in report_type:
            query = '''SELECT t.name, t.emp_id, d.name as dept FROM teachers t
                      LEFT JOIN departments d ON t.department_id = d.id
                      WHERE t.status='Active' AND t.id NOT IN 
                      (SELECT teacher_id FROM attendance WHERE date BETWEEN ? AND ?)'''
            params = [start, end]
        
        query += " ORDER BY a.date DESC, t.name"
        
        records = self.db.fetchall(query, params)
        
        if "Absent" in report_type:
            self.table.setColumnCount(3)
            self.table.setHorizontalHeaderLabels(["Name", "Employee ID", "Department"])
            self.table.setRowCount(len(records))
            for i, rec in enumerate(records):
                self.table.setItem(i, 0, QTableWidgetItem(rec['name']))
                self.table.setItem(i, 1, QTableWidgetItem(rec['emp_id']))
                self.table.setItem(i, 2, QTableWidgetItem(rec['dept'] or '-'))
        else:
            self.table.setColumnCount(8)
            self.table.setHorizontalHeaderLabels(["Name", "Emp ID", "Department", "Date", 
                                                  "Check-In", "Check-Out", "Status", "Remarks"])
            self.table.setRowCount(len(records))
            for i, rec in enumerate(records):
                self.table.setItem(i, 0, QTableWidgetItem(rec['name']))
                self.table.setItem(i, 1, QTableWidgetItem(rec['emp_id']))
                self.table.setItem(i, 2, QTableWidgetItem(rec['dept'] or '-'))
                self.table.setItem(i, 3, QTableWidgetItem(rec['date']))
                self.table.setItem(i, 4, QTableWidgetItem(rec['check_in'] or '-'))
                self.table.setItem(i, 5, QTableWidgetItem(rec['check_out'] or '-'))
                self.table.setItem(i, 6, QTableWidgetItem(rec['status']))
                
                remarks = []
                if rec['is_late']:
                    remarks.append("Late")
                if rec['is_early_leave']:
                    remarks.append("Early Leave")
                self.table.setItem(i, 7, QTableWidgetItem(", ".join(remarks)))
        
        self.table.horizontalHeader().setStretchLastSection(True)
    
    def export_report(self, format_type):
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "Error", "No data to export. Generate report first.")
            return
        
        if format_type == 'pdf':
            if not REPORTLAB_AVAILABLE:
                QMessageBox.warning(self, "Error", "ReportLab not installed. Install: pip install reportlab")
                return
            self.export_pdf()
        elif format_type == 'excel':
            if not OPENPYXL_AVAILABLE:
                QMessageBox.warning(self, "Error", "openpyxl not installed. Install: pip install openpyxl")
                return
            self.export_excel()
        elif format_type == 'csv':
            self.export_csv()
    
    def export_pdf(self):
        filepath, _ = QFileDialog.getSaveFileName(self, "Save PDF", "", "PDF Files (*.pdf)")
        if not filepath:
            return
        
        doc = SimpleDocTemplate(filepath, pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], alignment=TA_CENTER)
        
        institute_name = self.db.fetchone("SELECT value FROM settings WHERE key='institute_name'")
        title = Paragraph(institute_name['value'] if institute_name else "Attendance Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        data = []
        headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
        data.append(headers)
        
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else '')
            data.append(row_data)
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        QMessageBox.information(self, "Success", f"PDF exported to {filepath}")
    
    def export_excel(self):
        filepath, _ = QFileDialog.getSaveFileName(self, "Save Excel", "", "Excel Files (*.xlsx)")
        if not filepath:
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                ws.cell(row=row+2, column=col+1, value=item.text() if item else '')
        
        wb.save(filepath)
        QMessageBox.information(self, "Success", f"Excel exported to {filepath}")
    
    def export_csv(self):
        filepath, _ = QFileDialog.getSaveFileName(self, "Save CSV", "", "CSV Files (*.csv)")
        if not filepath:
            return
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            headers = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
            writer.writerow(headers)
            
            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else '')
                writer.writerow(row_data)
        
        QMessageBox.information(self, "Success", f"CSV exported to {filepath}")

class DashboardWidget(QWidget):
    def __init__(self, db, device, user_data):
        super().__init__()
        self.db = db
        self.device = device
        self.user_data = user_data
        self.setup_ui()
        self.update_stats()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        
        header = QLabel("Dashboard")
        header.setStyleSheet("font-size: 24px; font-weight: bold; margin: 10px;")
        layout.addWidget(header)
        
        stats_layout = QGridLayout()
        
        self.total_label = self.create_stat_card("Total Teachers", "0", "#3498db")
        stats_layout.addWidget(self.total_label, 0, 0)
        
        self.present_label = self.create_stat_card("Present Today", "0", "#2ecc71")
        stats_layout.addWidget(self.present_label, 0, 1)
        
        self.absent_label = self.create_stat_card("Absent Today", "0", "#e74c3c")
        stats_layout.addWidget(self.absent_label, 0, 2)
        
        self.late_label = self.create_stat_card("Late Today", "0", "#f39c12")
        stats_layout.addWidget(self.late_label, 1, 0)
        
        self.device_label = self.create_stat_card("Device Status", "Disconnected", "#95a5a6")
        stats_layout.addWidget(self.device_label, 1, 1)
        
        layout.addLayout(stats_layout)
        
        refresh_btn = QPushButton("Refresh Stats")
        refresh_btn.clicked.connect(self.update_stats)
        layout.addWidget(refresh_btn)
        
        self.setLayout(layout)
        
        timer = QTimer(self)
        timer.timeout.connect(self.update_stats)
        timer.start(30000)
    
    def create_stat_card(self, title, value, color):
        card = QFrame()
        card.setFrameShape(QFrame.Shape.StyledPanel)
        card.setStyleSheet(f"background-color: {color}; color: white; padding: 20px; border-radius: 10px;")
        
        layout = QVBoxLayout()
        
        title_label = QLabel(title)
        title_label.setStyleSheet("font-size: 14px; font-weight: bold;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        value_label = QLabel(value)
        value_label.setStyleSheet("font-size: 32px; font-weight: bold;")
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        value_label.setObjectName("value")
        
        layout.addWidget(title_label)
        layout.addWidget(value_label)
        
        card.setLayout(layout)
        return card
    
    def update_stats(self):
        total = self.db.fetchone("SELECT COUNT(*) as count FROM teachers WHERE status='Active'")
        self.update_card_value(self.total_label, str(total['count']))
        
        today = date.today().strftime('%Y-%m-%d')
        present = self.db.fetchone("SELECT COUNT(*) as count FROM attendance WHERE date=? AND status='Present'", (today,))
        self.update_card_value(self.present_label, str(present['count']))
        
        absent_count = total['count'] - present['count']
        self.update_card_value(self.absent_label, str(absent_count))
        
        late = self.db.fetchone("SELECT COUNT(*) as count FROM attendance WHERE date=? AND is_late=1", (today,))
        self.update_card_value(self.late_label, str(late['count']))
        
        device_status = "Connected" if self.device.connected else "Disconnected"
        self.update_card_value(self.device_label, device_status)
    
    def update_card_value(self, card, value):
        value_label = card.findChild(QLabel, "value")
        if value_label:
            value_label.setText(value)

class SettingsWidget(QWidget):
    def __init__(self, db, device, user_data, main_window):
        super().__init__()
        self.db = db
        self.device = device
        self.user_data = user_data
        self.main_window = main_window
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        
        header = QLabel("Settings")
        header.setStyleSheet("font-size: 24px; font-weight: bold; margin: 10px;")
        layout.addWidget(header)
        
        tabs = QTabWidget()
        
        tabs.addTab(self.create_institute_tab(), "Institute")
        tabs.addTab(self.create_appearance_tab(), "Appearance")
        tabs.addTab(self.create_backup_tab(), "Backup & Restore")
        tabs.addTab(self.create_password_tab(), "Change Password")
        
        layout.addWidget(tabs)
        self.setLayout(layout)
    
    def create_institute_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        name_label = QLabel("Institute Name:")
        self.name_input = QLineEdit()
        name_value = self.db.fetchone("SELECT value FROM settings WHERE key='institute_name'")
        if name_value:
            self.name_input.setText(name_value['value'])
        layout.addWidget(name_label)
        layout.addWidget(self.name_input)
        
        address_label = QLabel("Address:")
        self.address_input = QTextEdit()
        self.address_input.setMaximumHeight(100)
        address_value = self.db.fetchone("SELECT value FROM settings WHERE key='institute_address'")
        if address_value:
            self.address_input.setText(address_value['value'])
        layout.addWidget(address_label)
        layout.addWidget(self.address_input)
        
        save_btn = QPushButton("Save Institute Info")
        save_btn.clicked.connect(self.save_institute_info)
        layout.addWidget(save_btn)
        
        layout.addStretch()
        widget.setLayout(layout)
        return widget
    
    def create_appearance_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        theme_label = QLabel("Theme:")
        self.theme_combo = QComboBox()
        self.theme_combo.addItems(["Light", "Dark"])
        
        theme_value = self.db.fetchone("SELECT value FROM settings WHERE key='theme'")
        if theme_value:
            self.theme_combo.setCurrentText(theme_value['value'])
        
        self.theme_combo.currentTextChanged.connect(self.change_theme)
        
        layout.addWidget(theme_label)
        layout.addWidget(self.theme_combo)
        layout.addStretch()
        
        widget.setLayout(layout)
        return widget
    
    def create_backup_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        backup_btn = QPushButton("Backup Database")
        backup_btn.clicked.connect(self.backup_database)
        layout.addWidget(backup_btn)
        
        restore_btn = QPushButton("Restore Database")
        restore_btn.clicked.connect(self.restore_database)
        layout.addWidget(restore_btn)
        
        layout.addStretch()
        widget.setLayout(layout)
        return widget
    
    def create_password_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        current_label = QLabel("Current Password:")
        self.current_password = QLineEdit()
        self.current_password.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(current_label)
        layout.addWidget(self.current_password)
        
        new_label = QLabel("New Password:")
        self.new_password = QLineEdit()
        self.new_password.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(new_label)
        layout.addWidget(self.new_password)
        
        confirm_label = QLabel("Confirm Password:")
        self.confirm_password = QLineEdit()
        self.confirm_password.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(confirm_label)
        layout.addWidget(self.confirm_password)
        
        change_btn = QPushButton("Change Password")
        change_btn.clicked.connect(self.change_password)
        layout.addWidget(change_btn)
        
        layout.addStretch()
        widget.setLayout(layout)
        return widget
    
    def save_institute_info(self):
        name = self.name_input.text().strip()
        address = self.address_input.toPlainText().strip()
        
        self.db.execute("UPDATE settings SET value=? WHERE key='institute_name'", (name,))
        self.db.execute("UPDATE settings SET value=? WHERE key='institute_address'", (address,))
        
        QMessageBox.information(self, "Success", "Institute information saved")
    
    def change_theme(self, theme):
        self.db.execute("UPDATE settings SET value=? WHERE key='theme'", (theme,))
        
        if theme == "Dark":
            dark_palette = QPalette()
            dark_palette.setColor(QPalette.ColorRole.Window, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.ColorRole.WindowText, Qt.GlobalColor.white)
            dark_palette.setColor(QPalette.ColorRole.Base, QColor(35, 35, 35))
            dark_palette.setColor(QPalette.ColorRole.AlternateBase, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.ColorRole.ToolTipBase, Qt.GlobalColor.white)
            dark_palette.setColor(QPalette.ColorRole.ToolTipText, Qt.GlobalColor.white)
            dark_palette.setColor(QPalette.ColorRole.Text, Qt.GlobalColor.white)
            dark_palette.setColor(QPalette.ColorRole.Button, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.ColorRole.ButtonText, Qt.GlobalColor.white)
            dark_palette.setColor(QPalette.ColorRole.Link, QColor(42, 130, 218))
            dark_palette.setColor(QPalette.ColorRole.Highlight, QColor(42, 130, 218))
            dark_palette.setColor(QPalette.ColorRole.HighlightedText, Qt.GlobalColor.black)
            QApplication.instance().setPalette(dark_palette)
        else:
            QApplication.instance().setPalette(QApplication.style().standardPalette())
    
    def backup_database(self):
        filepath, _ = QFileDialog.getSaveFileName(self, "Backup Database", 
                                                  f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db",
                                                  "Database Files (*.db)")
        if filepath:
            self.db.backup(filepath)
            QMessageBox.information(self, "Success", f"Database backed up to {filepath}")
    
    def restore_database(self):
        filepath, _ = QFileDialog.getOpenFileName(self, "Restore Database", "", "Database Files (*.db)")
        if filepath:
            reply = QMessageBox.question(self, "Confirm", 
                                        "This will replace all current data. Continue?",
                                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                self.db.restore(filepath)
                QMessageBox.information(self, "Success", "Database restored. Please restart the application.")
    
    def change_password(self):
        current = self.current_password.text()
        new = self.new_password.text()
        confirm = self.confirm_password.text()
        
        if not current or not new or not confirm:
            QMessageBox.warning(self, "Error", "All fields are required")
            return
        
        if new != confirm:
            QMessageBox.warning(self, "Error", "New passwords do not match")
            return
        
        current_hash = hashlib.sha256(current.encode()).hexdigest()
        user = self.db.fetchone("SELECT * FROM users WHERE id=? AND password=?",
                               (self.user_data['id'], current_hash))
        
        if not user:
            QMessageBox.warning(self, "Error", "Current password is incorrect")
            return
        
        new_hash = hashlib.sha256(new.encode()).hexdigest()
        self.db.execute("UPDATE users SET password=? WHERE id=?", (new_hash, self.user_data['id']))
        
        QMessageBox.information(self, "Success", "Password changed successfully")
        self.current_password.clear()
        self.new_password.clear()
        self.confirm_password.clear()

class MainWindow(QMainWindow):
    def __init__(self, db, device, user_data):
        super().__init__()
        self.db = db
        self.device = device
        self.user_data = user_data
        self.setWindowTitle("Biometric Attendance Management System")
        self.setMinimumSize(1200, 800)
        self.setup_ui()
        self.apply_theme()
    
    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QHBoxLayout()
        
        sidebar = QWidget()
        sidebar.setFixedWidth(200)
        sidebar.setStyleSheet("background-color: #2c3e50; color: white;")
        sidebar_layout = QVBoxLayout()
        
        logo_label = QLabel("BAMS")
        logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        logo_label.setStyleSheet("font-size: 24px; font-weight: bold; padding: 20px;")
        sidebar_layout.addWidget(logo_label)
        
        user_label = QLabel(f"User: {self.user_data['username']}\nRole: {self.user_data['role']}")
        user_label.setStyleSheet("padding: 10px; font-size: 12px;")
        user_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sidebar_layout.addWidget(user_label)
        
        self.nav_buttons = []
        
        dashboard_btn = self.create_nav_button("Dashboard")
        dashboard_btn.clicked.connect(lambda: self.switch_page(0))
        sidebar_layout.addWidget(dashboard_btn)
        
        attendance_btn = self.create_nav_button("Attendance")
        attendance_btn.clicked.connect(lambda: self.switch_page(1))
        sidebar_layout.addWidget(attendance_btn)
        
        teachers_btn = self.create_nav_button("Teachers")
        teachers_btn.clicked.connect(lambda: self.switch_page(2))
        sidebar_layout.addWidget(teachers_btn)
        
        reports_btn = self.create_nav_button("Reports")
        reports_btn.clicked.connect(lambda: self.switch_page(3))
        sidebar_layout.addWidget(reports_btn)
        
        settings_btn = self.create_nav_button("Settings")
        settings_btn.clicked.connect(lambda: self.switch_page(4))
        sidebar_layout.addWidget(settings_btn)
        
        sidebar_layout.addStretch()
        
        logout_btn = QPushButton("Logout")
        logout_btn.setStyleSheet("background-color: #e74c3c; color: white; padding: 10px; border: none;")
        logout_btn.clicked.connect(self.logout)
        sidebar_layout.addWidget(logout_btn)
        
        sidebar.setLayout(sidebar_layout)
        main_layout.addWidget(sidebar)
        
        self.content_stack = QStackedWidget()
        
        self.content_stack.addWidget(DashboardWidget(self.db, self.device, self.user_data))
        self.content_stack.addWidget(AttendanceWidget(self.db, self.device, self.user_data))
        self.content_stack.addWidget(TeachersWidget(self.db, self.device, self.user_data))
        self.content_stack.addWidget(ReportsWidget(self.db, self.device, self.user_data))
        self.content_stack.addWidget(SettingsWidget(self.db, self.device, self.user_data, self))
        
        main_layout.addWidget(self.content_stack)
        
        central_widget.setLayout(main_layout)
    
    def create_nav_button(self, text):
        btn = QPushButton(text)
        btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: white;
                text-align: left;
                padding: 15px;
                border: none;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #34495e;
            }
        """)
        self.nav_buttons.append(btn)
        return btn
    
    def switch_page(self, index):
        self.content_stack.setCurrentIndex(index)
    
    def apply_theme(self):
        theme = self.db.fetchone("SELECT value FROM settings WHERE key='theme'")
        if theme and theme['value'] == "Dark":
            dark_palette = QPalette()
            dark_palette.setColor(QPalette.ColorRole.Window, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.ColorRole.WindowText, Qt.GlobalColor.white)
            dark_palette.setColor(QPalette.ColorRole.Base, QColor(35, 35, 35))
            dark_palette.setColor(QPalette.ColorRole.AlternateBase, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.ColorRole.ToolTipBase, Qt.GlobalColor.white)
            dark_palette.setColor(QPalette.ColorRole.ToolTipText, Qt.GlobalColor.white)
            dark_palette.setColor(QPalette.ColorRole.Text, Qt.GlobalColor.white)
            dark_palette.setColor(QPalette.ColorRole.Button, QColor(53, 53, 53))
            dark_palette.setColor(QPalette.ColorRole.ButtonText, Qt.GlobalColor.white)
            dark_palette.setColor(QPalette.ColorRole.Link, QColor(42, 130, 218))
            dark_palette.setColor(QPalette.ColorRole.Highlight, QColor(42, 130, 218))
            dark_palette.setColor(QPalette.ColorRole.HighlightedText, Qt.GlobalColor.black)
            QApplication.instance().setPalette(dark_palette)
    
    def logout(self):
        reply = QMessageBox.question(self, "Logout", "Are you sure you want to logout?",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.close()
            login_dialog = LoginDialog(self.db)
            if login_dialog.exec() == QDialog.DialogCode.Accepted:
                new_window = MainWindow(self.db, self.device, login_dialog.user_data)
                new_window.show()

def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    db = Database()
    device = BiometricDevice()
    
    device.initialize()
    device.open_device()
    
    login_dialog = LoginDialog(db)
    if login_dialog.exec() == QDialog.DialogCode.Accepted:
        window = MainWindow(db, device, login_dialog.user_data)
        window.show()
        exit_code = app.exec()
        
        device.exit()
        sys.exit(exit_code)
    else:
        sys.exit(0)

if __name__ == '__main__':
    main()