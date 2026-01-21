import sys
import os
import sqlite3
import json
import hashlib
import base64
from datetime import datetime, date, time, timedelta
from pathlib import Path
import ctypes
import platform
from typing import Optional, List, Dict, Any, Tuple
from dataclasses import dataclass
from enum import Enum

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QPushButton, QLabel, QLineEdit, QTableWidget, QTableWidgetItem,
    QComboBox, QDateEdit, QTimeEdit, QTextEdit, QMessageBox, QDialog,
    QFormLayout, QGroupBox, QCheckBox, QSpinBox, QTabWidget, QFileDialog,
    QStackedWidget, QScrollArea, QFrame, QHeaderView, QProgressBar,
    QDialogButtonBox, QRadioButton, QButtonGroup, QSplitter, QCalendarWidget,
    QListWidget, QListWidgetItem, QGridLayout, QDoubleSpinBox
)
from PyQt6.QtCore import Qt, QDate, QTime, QTimer, pyqtSignal, QThread, QSize
from PyQt6.QtGui import QIcon, QPixmap, QImage, QPalette, QColor, QFont, QPainter

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    REPORTLAB_AVAILABLE = True
except:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except:
    PANDAS_AVAILABLE = False


class NBioBSPWrapper:
    def __init__(self):
        self.initialized = False
        self.device_connected = False
        self.handle = None
        self.dll = None
        self._init_sdk()
    
    def _init_sdk(self):
        try:
            if platform.system() == 'Windows':
                sdk_paths = [
                    r"C:\Program Files\Nitgen\NBioBSP\Bin\NBioBSP.dll",
                    r"C:\Program Files (x86)\Nitgen\NBioBSP\Bin\NBioBSP.dll",
                    r"C:\Program Files\Common Files\Nitgen\NBioBSP\NBioBSP.dll",
                    "NBioBSP.dll"
                ]
                
                for path in sdk_paths:
                    if os.path.exists(path):
                        self.dll = ctypes.WinDLL(path)
                        break
                
                if not self.dll:
                    self.dll = ctypes.WinDLL("NBioBSP.dll")
                
                self.dll.NBioAPI_Init.argtypes = [ctypes.POINTER(ctypes.c_void_p)]
                self.dll.NBioAPI_Init.restype = ctypes.c_uint32
                
                self.dll.NBioAPI_Terminate.argtypes = [ctypes.c_void_p]
                self.dll.NBioAPI_Terminate.restype = ctypes.c_uint32
                
                self.dll.NBioAPI_OpenDevice.argtypes = [ctypes.c_void_p, ctypes.c_uint32]
                self.dll.NBioAPI_OpenDevice.restype = ctypes.c_uint32
                
                self.dll.NBioAPI_CloseDevice.argtypes = [ctypes.c_void_p, ctypes.c_uint32]
                self.dll.NBioAPI_CloseDevice.restype = ctypes.c_uint32
                
                self.dll.NBioAPI_Capture.argtypes = [ctypes.c_void_p, ctypes.c_int32, ctypes.POINTER(ctypes.c_void_p), ctypes.c_int32, ctypes.c_void_p]
                self.dll.NBioAPI_Capture.restype = ctypes.c_uint32
                
                self.dll.NBioAPI_VerifyMatch.argtypes = [ctypes.c_void_p, ctypes.c_void_p, ctypes.c_void_p, ctypes.POINTER(ctypes.c_bool), ctypes.c_void_p]
                self.dll.NBioAPI_VerifyMatch.restype = ctypes.c_uint32
                
                self.dll.NBioAPI_GetFIRFromHandle.argtypes = [ctypes.c_void_p, ctypes.c_void_p, ctypes.POINTER(ctypes.c_void_p)]
                self.dll.NBioAPI_GetFIRFromHandle.restype = ctypes.c_uint32
                
                self.dll.NBioAPI_FreeFIRHandle.argtypes = [ctypes.c_void_p, ctypes.c_void_p]
                self.dll.NBioAPI_FreeFIRHandle.restype = ctypes.c_uint32
                
                handle = ctypes.c_void_p()
                ret = self.dll.NBioAPI_Init(ctypes.byref(handle))
                
                if ret == 0:
                    self.handle = handle
                    device_ret = self.dll.NBioAPI_OpenDevice(self.handle, 0)
                    if device_ret == 0:
                        self.device_connected = True
                    self.initialized = True
                    return True
                    
        except Exception as e:
            pass
        
        return False
    
    def capture_fingerprint(self, timeout=10000):
        if not self.initialized or not self.handle:
            return None
            
        try:
            fir_handle = ctypes.c_void_p()
            ret = self.dll.NBioAPI_Capture(
                self.handle,
                -1,
                ctypes.byref(fir_handle),
                timeout,
                None
            )
            
            if ret == 0 and fir_handle:
                fir_data = ctypes.c_void_p()
                ret = self.dll.NBioAPI_GetFIRFromHandle(self.handle, fir_handle, ctypes.byref(fir_data))
                
                if ret == 0 and fir_data:
                    size = ctypes.c_uint32()
                    ctypes.memmove(ctypes.byref(size), fir_data.value, 4)
                    
                    buffer = ctypes.create_string_buffer(size.value)
                    ctypes.memmove(buffer, fir_data.value, size.value)
                    
                    self.dll.NBioAPI_FreeFIRHandle(self.handle, fir_handle)
                    
                    return base64.b64encode(buffer.raw).decode('utf-8')
                    
            return None
        except:
            return None
    
    def verify_fingerprint(self, captured_template, stored_template):
        if not self.initialized or not self.handle:
            return False
            
        try:
            captured_data = base64.b64decode(captured_template)
            stored_data = base64.b64decode(stored_template)
            
            captured_fir = ctypes.create_string_buffer(captured_data)
            stored_fir = ctypes.create_string_buffer(stored_data)
            
            match_result = ctypes.c_bool()
            ret = self.dll.NBioAPI_VerifyMatch(
                self.handle,
                ctypes.cast(captured_fir, ctypes.c_void_p),
                ctypes.cast(stored_fir, ctypes.c_void_p),
                ctypes.byref(match_result),
                None
            )
            
            if ret == 0:
                return match_result.value
                
        except:
            pass
        
        return False
    
    def is_device_connected(self):
        return self.device_connected
    
    def cleanup(self):
        if self.initialized and self.handle:
            try:
                if self.device_connected:
                    self.dll.NBioAPI_CloseDevice(self.handle, 0)
                self.dll.NBioAPI_Terminate(self.handle)
            except:
                pass
        self.initialized = False
        self.device_connected = False


class UserRole(Enum):
    SUPER_ADMIN = "Super Admin"
    ADMIN = "Admin"
    OPERATOR = "Operator"
    TEACHER = "Teacher"
    STUDENT = "Student"
    STAFF = "Staff"


class AttendanceStatus(Enum):
    PRESENT = "Present"
    ABSENT = "Absent"
    LATE = "Late"
    HALF_DAY = "Half Day"
    LEAVE = "Leave"
    HOLIDAY = "Holiday"


class LeaveStatus(Enum):
    PENDING = "Pending"
    APPROVED = "Approved"
    REJECTED = "Rejected"


class Database:
    def __init__(self, db_path='biometric_attendance7.db'):
        self.db_path = db_path
        self.conn = None
        self.init_database()
    
    def connect(self):
        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row
        return self.conn
    
    def init_database(self):
        conn = self.connect()
        cursor = conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                role TEXT NOT NULL,
                full_name TEXT NOT NULL,
                email TEXT,
                phone TEXT,
                created_at TEXT NOT NULL,
                is_active INTEGER DEFAULT 1
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS institute_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                institute_name TEXT NOT NULL,
                address TEXT,
                phone TEXT,
                email TEXT,
                logo BLOB,
                established_year INTEGER,
                registration_no TEXT
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS departments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                head_id INTEGER,
                description TEXT,
                created_at TEXT NOT NULL,
                is_active INTEGER DEFAULT 1
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS classes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                description TEXT,
                created_at TEXT NOT NULL,
                is_active INTEGER DEFAULT 1
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS sections (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                class_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                capacity INTEGER,
                created_at TEXT NOT NULL,
                is_active INTEGER DEFAULT 1,
                FOREIGN KEY (class_id) REFERENCES classes(id),
                UNIQUE(class_id, name)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS subjects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                code TEXT UNIQUE,
                description TEXT,
                created_at TEXT NOT NULL,
                is_active INTEGER DEFAULT 1
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                is_current INTEGER DEFAULT 0,
                created_at TEXT NOT NULL
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS shifts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL,
                grace_time INTEGER DEFAULT 0,
                created_at TEXT NOT NULL,
                is_active INTEGER DEFAULT 1
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS teachers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                teacher_id TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                cnic TEXT UNIQUE NOT NULL,
                phone TEXT,
                email TEXT,
                address TEXT,
                gender TEXT,
                date_of_birth TEXT,
                qualification TEXT,
                department_id INTEGER,
                designation TEXT,
                shift_id INTEGER,
                joining_date TEXT,
                photo BLOB,
                status INTEGER DEFAULT 1,
                created_at TEXT NOT NULL,
                FOREIGN KEY (department_id) REFERENCES departments(id),
                FOREIGN KEY (shift_id) REFERENCES shifts(id)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS students (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                roll_no TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                father_name TEXT,
                cnic_bform TEXT UNIQUE,
                phone TEXT,
                address TEXT,
                gender TEXT,
                date_of_birth TEXT,
                class_id INTEGER,
                section_id INTEGER,
                session_id INTEGER,
                admission_date TEXT,
                photo BLOB,
                status INTEGER DEFAULT 1,
                created_at TEXT NOT NULL,
                FOREIGN KEY (class_id) REFERENCES classes(id),
                FOREIGN KEY (section_id) REFERENCES sections(id),
                FOREIGN KEY (session_id) REFERENCES sessions(id)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS staff (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_id TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                cnic TEXT UNIQUE NOT NULL,
                phone TEXT,
                email TEXT,
                address TEXT,
                gender TEXT,
                date_of_birth TEXT,
                role TEXT NOT NULL,
                shift_id INTEGER,
                salary_type TEXT,
                joining_date TEXT,
                photo BLOB,
                status INTEGER DEFAULT 1,
                created_at TEXT NOT NULL,
                FOREIGN KEY (shift_id) REFERENCES shifts(id)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS guardians (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                cnic TEXT NOT NULL,
                phone TEXT NOT NULL,
                relation TEXT NOT NULL,
                email TEXT,
                address TEXT,
                occupation TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (student_id) REFERENCES students(id)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS biometric_data (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                person_type TEXT NOT NULL,
                person_id INTEGER NOT NULL,
                finger_index INTEGER NOT NULL,
                template_data TEXT NOT NULL,
                quality_score INTEGER,
                enrolled_at TEXT NOT NULL,
                UNIQUE(person_type, person_id, finger_index)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS attendance_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                person_type TEXT NOT NULL,
                person_id INTEGER NOT NULL,
                log_date TEXT NOT NULL,
                check_in_time TEXT,
                check_out_time TEXT,
                status TEXT NOT NULL,
                is_late INTEGER DEFAULT 0,
                is_early_leave INTEGER DEFAULT 0,
                punch_type TEXT,
                remarks TEXT,
                created_at TEXT NOT NULL
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS holidays (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                date TEXT NOT NULL,
                type TEXT NOT NULL,
                description TEXT,
                created_at TEXT NOT NULL
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS leaves (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                person_type TEXT NOT NULL,
                person_id INTEGER NOT NULL,
                leave_type TEXT NOT NULL,
                from_date TEXT NOT NULL,
                to_date TEXT NOT NULL,
                reason TEXT NOT NULL,
                status TEXT NOT NULL,
                approved_by INTEGER,
                approved_at TEXT,
                remarks TEXT,
                created_at TEXT NOT NULL
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS attendance_policies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                grace_time INTEGER DEFAULT 0,
                late_fine_amount REAL DEFAULT 0,
                half_day_hours REAL DEFAULT 4,
                full_day_hours REAL DEFAULT 8,
                created_at TEXT NOT NULL
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS system_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                action TEXT NOT NULL,
                entity_type TEXT,
                entity_id INTEGER,
                details TEXT,
                created_at TEXT NOT NULL
            )
        ''')
        
        cursor.execute("SELECT COUNT(*) FROM users")
        if cursor.fetchone()[0] == 0:
            password_hash = hashlib.sha256('admin123'.encode()).hexdigest()
            cursor.execute('''
                INSERT INTO users (username, password, role, full_name, email, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', ('admin', password_hash, UserRole.SUPER_ADMIN.value, 'System Administrator', 
                  'admin@system.com', datetime.now().isoformat()))
        
        cursor.execute("SELECT COUNT(*) FROM institute_settings")
        if cursor.fetchone()[0] == 0:
            cursor.execute('''
                INSERT INTO institute_settings (institute_name, address, phone, email, established_year)
                VALUES (?, ?, ?, ?, ?)
            ''', ('Educational Institute', '123 Main Street', '+92-300-1234567', 
                  'info@institute.edu', 2020))
        
        cursor.execute("SELECT COUNT(*) FROM departments")
        if cursor.fetchone()[0] == 0:
            depts = [
                ('Computer Science', 'Department of Computer Science'),
                ('Mathematics', 'Department of Mathematics'),
                ('Physics', 'Department of Physics'),
                ('English', 'Department of English')
            ]
            for name, desc in depts:
                cursor.execute('''
                    INSERT INTO departments (name, description, created_at)
                    VALUES (?, ?, ?)
                ''', (name, desc, datetime.now().isoformat()))
        
        cursor.execute("SELECT COUNT(*) FROM classes")
        if cursor.fetchone()[0] == 0:
            classes = ['Class 1', 'Class 2', 'Class 3', 'Class 4']
            for cls in classes:
                cursor.execute('''
                    INSERT INTO classes (name, description, created_at)
                    VALUES (?, ?, ?)
                ''', (cls, f'Standard {cls}', datetime.now().isoformat()))
        
        cursor.execute("SELECT COUNT(*) FROM shifts")
        if cursor.fetchone()[0] == 0:
            shifts = [
                ('Morning Shift', '08:00', '14:00', 15),
                ('Evening Shift', '14:00', '20:00', 15),
                ('Day Shift', '09:00', '17:00', 15),
                ('Night Shift', '20:00', '04:00', 15)
            ]
            for name, start, end, grace in shifts:
                cursor.execute('''
                    INSERT INTO shifts (name, start_time, end_time, grace_time, created_at)
                    VALUES (?, ?, ?, ?, ?)
                ''', (name, start, end, grace, datetime.now().isoformat()))
        
        cursor.execute("SELECT COUNT(*) FROM sessions")
        if cursor.fetchone()[0] == 0:
            current_year = datetime.now().year
            sessions = [
                (f'{current_year-1}-{current_year}', f'{current_year-1}-01-01', f'{current_year}-12-31', 0),
                (f'{current_year}-{current_year+1}', f'{current_year}-01-01', f'{current_year+1}-12-31', 1),
                (f'{current_year+1}-{current_year+2}', f'{current_year+1}-01-01', f'{current_year+2}-12-31', 0),
                (f'{current_year+2}-{current_year+3}', f'{current_year+2}-01-01', f'{current_year+3}-12-31', 0)
            ]
            for name, start, end, current in sessions:
                cursor.execute('''
                    INSERT INTO sessions (name, start_date, end_date, is_current, created_at)
                    VALUES (?, ?, ?, ?, ?)
                ''', (name, start, end, current, datetime.now().isoformat()))
        
        cursor.execute("SELECT COUNT(*) FROM teachers")
        if cursor.fetchone()[0] == 0:
            cursor.execute("SELECT id FROM departments LIMIT 1")
            dept_id = cursor.fetchone()[0]
            cursor.execute("SELECT id FROM shifts WHERE name = 'Morning Shift'")
            shift_id = cursor.fetchone()[0]
            
            teachers = [
                ('TCH001', 'Ahmed Ali', '12345-6789012-3', '+92-300-1111111', 'ahmed@school.com', 'House 10, Street 5', 'Male', '1985-05-15', 'MSc Computer Science', dept_id, 'Senior Teacher', shift_id, '2020-01-15'),
                ('TCH002', 'Fatima Khan', '23456-7890123-4', '+92-301-2222222', 'fatima@school.com', 'House 20, Street 6', 'Female', '1988-08-20', 'MSc Mathematics', dept_id, 'Teacher', shift_id, '2021-03-10'),
                ('TCH003', 'Hassan Raza', '34567-8901234-5', '+92-302-3333333', 'hassan@school.com', 'House 30, Street 7', 'Male', '1990-12-25', 'MSc Physics', dept_id, 'Teacher', shift_id, '2022-06-01'),
                ('TCH004', 'Ayesha Malik', '45678-9012345-6', '+92-303-4444444', 'ayesha@school.com', 'House 40, Street 8', 'Female', '1992-03-18', 'MA English', dept_id, 'Junior Teacher', shift_id, '2023-09-15')
            ]
            
            for teacher in teachers:
                cursor.execute('''
                    INSERT INTO teachers (teacher_id, name, cnic, phone, email, address, gender, date_of_birth, 
                                        qualification, department_id, designation, shift_id, joining_date, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (*teacher, datetime.now().isoformat()))
        
        cursor.execute("SELECT COUNT(*) FROM students")
        if cursor.fetchone()[0] == 0:
            cursor.execute("SELECT id FROM classes LIMIT 1")
            class_id = cursor.fetchone()[0]
            cursor.execute("SELECT id FROM sections WHERE class_id = ? LIMIT 1", (class_id,))
            section_result = cursor.fetchone()
            if not section_result:
                cursor.execute('''
                    INSERT INTO sections (class_id, name, capacity, created_at)
                    VALUES (?, ?, ?, ?)
                ''', (class_id, 'Section A', 40, datetime.now().isoformat()))
                section_id = cursor.lastrowid
            else:
                section_id = section_result[0]
            
            cursor.execute("SELECT id FROM sessions WHERE is_current = 1")
            session_id = cursor.fetchone()[0]
            
            students = [
                ('STD001', 'Ali Ahmed', 'Muhammad Ahmed', '12345-1234567-1', '+92-304-5555555', 'House 50, Street 9', 'Male', '2010-01-10', class_id, section_id, session_id, '2024-01-15'),
                ('STD002', 'Sara Khan', 'Kamran Khan', '23456-2345678-2', '+92-305-6666666', 'House 60, Street 10', 'Female', '2010-03-20', class_id, section_id, session_id, '2024-01-15'),
                ('STD003', 'Usman Malik', 'Khalid Malik', '34567-3456789-3', '+92-306-7777777', 'House 70, Street 11', 'Male', '2010-06-15', class_id, section_id, session_id, '2024-01-15'),
                ('STD004', 'Zainab Ali', 'Ali Raza', '45678-4567890-4', '+92-307-8888888', 'House 80, Street 12', 'Female', '2010-09-25', class_id, section_id, session_id, '2024-01-15')
            ]
            
            for student in students:
                cursor.execute('''
                    INSERT INTO students (roll_no, name, father_name, cnic_bform, phone, address, gender, 
                                        date_of_birth, class_id, section_id, session_id, admission_date, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (*student, datetime.now().isoformat()))
        
        cursor.execute("SELECT COUNT(*) FROM staff")
        if cursor.fetchone()[0] == 0:
            cursor.execute("SELECT id FROM shifts WHERE name = 'Day Shift'")
            shift_id = cursor.fetchone()[0]
            
            staff_members = [
                ('STF001', 'Akbar Hussain', '56789-0123456-7', '+92-308-9999999', 'akbar@school.com', 'House 90, Street 13', 'Male', '1980-07-10', 'Office Assistant', shift_id, 'Monthly', '2019-05-01'),
                ('STF002', 'Nazia Bibi', '67890-1234567-8', '+92-309-1010101', 'nazia@school.com', 'House 100, Street 14', 'Female', '1983-11-22', 'Librarian', shift_id, 'Monthly', '2020-08-15'),
                ('STF003', 'Imran Khan', '78901-2345678-9', '+92-310-1111222', 'imran@school.com', 'House 110, Street 15', 'Male', '1987-04-18', 'Lab Assistant', shift_id, 'Monthly', '2021-02-10'),
                ('STF004', 'Rabia Ahmed', '89012-3456789-0', '+92-311-2223333', 'rabia@school.com', 'House 120, Street 16', 'Female', '1991-09-30', 'Cleaner', shift_id, 'Daily', '2022-11-20')
            ]
            
            for staff in staff_members:
                cursor.execute('''
                    INSERT INTO staff (staff_id, name, cnic, phone, email, address, gender, date_of_birth, 
                                     role, shift_id, salary_type, joining_date, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (*staff, datetime.now().isoformat()))
        
        conn.commit()
        conn.close()
    
    def execute_query(self, query, params=None):
        conn = self.connect()
        cursor = conn.cursor()
        if params:
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        conn.commit()
        last_id = cursor.lastrowid
        conn.close()
        return last_id
    
    def fetch_all(self, query, params=None):
        conn = self.connect()
        cursor = conn.cursor()
        if params:
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        rows = cursor.fetchall()
        conn.close()
        return [dict(row) for row in rows]
    
    def fetch_one(self, query, params=None):
        conn = self.connect()
        cursor = conn.cursor()
        if params:
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        row = cursor.fetchone()
        conn.close()
        return dict(row) if row else None
    
    def backup_database(self, backup_path):
        import shutil
        try:
            shutil.copy2(self.db_path, backup_path)
            return True
        except:
            return False
    
    def restore_database(self, backup_path):
        import shutil
        try:
            if os.path.exists(backup_path):
                shutil.copy2(backup_path, self.db_path)
                self.init_database()
                return True
        except:
            pass
        return False


class BiometricEnrollmentDialog(QDialog):
    def __init__(self, parent, biometric_wrapper, person_type, person_id, person_name):
        super().__init__(parent)
        self.biometric = biometric_wrapper
        self.person_type = person_type
        self.person_id = person_id
        self.person_name = person_name
        self.db = parent.db
        self.captured_templates = {}
        
        self.setWindowTitle(f"Biometric Enrollment - {person_name}")
        self.setModal(True)
        self.resize(600, 500)
        
        layout = QVBoxLayout()
        
        info_label = QLabel(f"Enrolling fingerprints for: {person_name}")
        info_label.setStyleSheet("font-size: 14px; font-weight: bold; padding: 10px;")
        layout.addWidget(info_label)
        
        status_label = QLabel("Device Status: " + ("Connected" if self.biometric.is_device_connected() else "Disconnected"))
        status_label.setStyleSheet("padding: 5px;")
        layout.addWidget(status_label)
        
        fingers_group = QGroupBox("Select Fingers to Enroll")
        fingers_layout = QGridLayout()
        
        self.finger_checkboxes = {}
        fingers = [
            ("Right Thumb", 1), ("Right Index", 2), ("Right Middle", 3), ("Right Ring", 4), ("Right Little", 5),
            ("Left Thumb", 6), ("Left Index", 7), ("Left Middle", 8), ("Left Ring", 9), ("Left Little", 10)
        ]
        
        for idx, (name, index) in enumerate(fingers):
            cb = QCheckBox(name)
            cb.setProperty('finger_index', index)
            self.finger_checkboxes[index] = cb
            fingers_layout.addWidget(cb, idx // 2, idx % 2)
        
        fingers_group.setLayout(fingers_layout)
        layout.addWidget(fingers_group)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        self.status_text = QTextEdit()
        self.status_text.setReadOnly(True)
        self.status_text.setMaximumHeight(150)
        layout.addWidget(self.status_text)
        
        buttons_layout = QHBoxLayout()
        
        self.start_btn = QPushButton("Start Enrollment")
        self.start_btn.clicked.connect(self.start_enrollment)
        buttons_layout.addWidget(self.start_btn)
        
        self.save_btn = QPushButton("Save")
        self.save_btn.clicked.connect(self.save_enrollment)
        self.save_btn.setEnabled(False)
        buttons_layout.addWidget(self.save_btn)
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        buttons_layout.addWidget(cancel_btn)
        
        layout.addLayout(buttons_layout)
        self.setLayout(layout)
    
    def start_enrollment(self):
        selected_fingers = [idx for idx, cb in self.finger_checkboxes.items() if cb.isChecked()]
        
        if not selected_fingers:
            QMessageBox.warning(self, "Warning", "Please select at least one finger to enroll.")
            return
        
        if not self.biometric.is_device_connected():
            QMessageBox.warning(self, "Device Error", "Biometric device is not connected.")
            return
        
        self.start_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setMaximum(len(selected_fingers))
        self.progress_bar.setValue(0)
        self.status_text.clear()
        
        for idx, finger_idx in enumerate(selected_fingers):
            finger_name = None
            for name, index in [("Right Thumb", 1), ("Right Index", 2), ("Right Middle", 3), ("Right Ring", 4), ("Right Little", 5),
                               ("Left Thumb", 6), ("Left Index", 7), ("Left Middle", 8), ("Left Ring", 9), ("Left Little", 10)]:
                if index == finger_idx:
                    finger_name = name
                    break
            
            self.status_text.append(f"\nCapturing {finger_name}... Please place your finger on the scanner.")
            QApplication.processEvents()
            
            template = self.biometric.capture_fingerprint()
            
            if template:
                self.captured_templates[finger_idx] = template
                self.status_text.append(f"✓ {finger_name} captured successfully (Quality: Good)")
            else:
                self.status_text.append(f"✗ {finger_name} capture failed. Please try again.")
            
            self.progress_bar.setValue(idx + 1)
            QApplication.processEvents()
        
        self.start_btn.setEnabled(True)
        self.save_btn.setEnabled(len(self.captured_templates) > 0)
        
        if len(self.captured_templates) > 0:
            self.status_text.append(f"\n✓ Enrollment completed. {len(self.captured_templates)} fingerprint(s) captured.")
        else:
            self.status_text.append("\n✗ No fingerprints captured. Please try again.")
    
    def save_enrollment(self):
        if not self.captured_templates:
            return
        
        try:
            for finger_idx, template in self.captured_templates.items():
                existing = self.db.fetch_one(
                    "SELECT id FROM biometric_data WHERE person_type = ? AND person_id = ? AND finger_index = ?",
                    (self.person_type, self.person_id, finger_idx)
                )
                
                if existing:
                    self.db.execute_query(
                        "UPDATE biometric_data SET template_data = ?, quality_score = ?, enrolled_at = ? WHERE id = ?",
                        (template, 85, datetime.now().isoformat(), existing['id'])
                    )
                else:
                    self.db.execute_query(
                        "INSERT INTO biometric_data (person_type, person_id, finger_index, template_data, quality_score, enrolled_at) VALUES (?, ?, ?, ?, ?, ?)",
                        (self.person_type, self.person_id, finger_idx, template, 85, datetime.now().isoformat())
                    )
            
            QMessageBox.information(self, "Success", "Fingerprints enrolled successfully!")
            self.accept()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save fingerprints: {str(e)}")


class BiometricVerificationDialog(QDialog):
    verification_success = pyqtSignal(str, int, str)
    
    def __init__(self, parent, biometric_wrapper, db):
        super().__init__(parent)
        self.biometric = biometric_wrapper
        self.db = db
        
        self.setWindowTitle("Biometric Verification")
        self.setModal(True)
        self.resize(400, 300)
        
        layout = QVBoxLayout()
        
        status_label = QLabel("Device Status: " + ("Connected" if self.biometric.is_device_connected() else "Disconnected"))
        status_label.setStyleSheet("font-size: 12px; padding: 10px;")
        layout.addWidget(status_label)
        
        instruction_label = QLabel("Please place your finger on the scanner...")
        instruction_label.setStyleSheet("font-size: 14px; font-weight: bold; padding: 20px; color: blue;")
        instruction_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(instruction_label)
        
        self.result_label = QLabel("")
        self.result_label.setStyleSheet("font-size: 16px; padding: 20px;")
        self.result_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.result_label)
        
        buttons_layout = QHBoxLayout()
        
        retry_btn = QPushButton("Retry")
        retry_btn.clicked.connect(self.start_verification)
        buttons_layout.addWidget(retry_btn)
        
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.reject)
        buttons_layout.addWidget(close_btn)
        
        layout.addLayout(buttons_layout)
        self.setLayout(layout)
        
        QTimer.singleShot(500, self.start_verification)
    
    def start_verification(self):
        self.result_label.setText("Capturing...")
        QApplication.processEvents()
        
        if not self.biometric.is_device_connected():
            self.result_label.setText("Device not connected!")
            self.result_label.setStyleSheet("font-size: 16px; padding: 20px; color: red;")
            return
        
        captured = self.biometric.capture_fingerprint()
        
        if not captured:
            self.result_label.setText("Capture failed. Please retry.")
            self.result_label.setStyleSheet("font-size: 16px; padding: 20px; color: orange;")
            return
        
        all_templates = self.db.fetch_all("SELECT * FROM biometric_data")
        
        for template_data in all_templates:
            if self.biometric.verify_fingerprint(captured, template_data['template_data']):
                person_type = template_data['person_type']
                person_id = template_data['person_id']
                
                person_name = "Unknown"
                if person_type == 'teacher':
                    person = self.db.fetch_one("SELECT name FROM teachers WHERE id = ?", (person_id,))
                    if person:
                        person_name = person['name']
                elif person_type == 'student':
                    person = self.db.fetch_one("SELECT name FROM students WHERE id = ?", (person_id,))
                    if person:
                        person_name = person['name']
                elif person_type == 'staff':
                    person = self.db.fetch_one("SELECT name FROM staff WHERE id = ?", (person_id,))
                    if person:
                        person_name = person['name']
                
                self.result_label.setText(f"✓ Verified: {person_name}")
                self.result_label.setStyleSheet("font-size: 16px; padding: 20px; color: green; font-weight: bold;")
                
                self.verification_success.emit(person_type, person_id, person_name)
                
                QTimer.singleShot(2000, self.accept)
                return
        
        self.result_label.setText("✗ Fingerprint not recognized")
        self.result_label.setStyleSheet("font-size: 16px; padding: 20px; color: red;")


class LoginWindow(QDialog):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.current_user = None
        
        self.setWindowTitle("Biometric Attendance System - Login")
        self.setFixedSize(400, 300)
        
        layout = QVBoxLayout()
        
        title = QLabel("Biometric Attendance System")
        title.setStyleSheet("font-size: 20px; font-weight: bold; padding: 20px;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)
        
        form_layout = QFormLayout()
        
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("Enter username")
        form_layout.addRow("Username:", self.username_input)
        
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_input.setPlaceholderText("Enter password")
        self.password_input.returnPressed.connect(self.login)
        form_layout.addRow("Password:", self.password_input)
        
        layout.addLayout(form_layout)
        
        login_btn = QPushButton("Login")
        login_btn.clicked.connect(self.login)
        login_btn.setStyleSheet("padding: 10px; font-size: 14px;")
        layout.addWidget(login_btn)
        
        self.setLayout(layout)
    
    def login(self):
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()
        
        if not username or not password:
            QMessageBox.warning(self, "Warning", "Please enter both username and password.")
            return
        
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        user = self.db.fetch_one(
            "SELECT * FROM users WHERE username = ? AND password = ? AND is_active = 1",
            (username, password_hash)
        )
        
        if user:
            self.current_user = user
            self.accept()
        else:
            QMessageBox.warning(self, "Login Failed", "Invalid username or password.")


class MainWindow(QMainWindow):
    def __init__(self, db, biometric, current_user):
        super().__init__()
        self.db = db
        self.biometric = biometric
        self.current_user = current_user
        
        self.setWindowTitle("Biometric Attendance Management System")
        self.setGeometry(100, 100, 1400, 900)
        
        self.dark_mode = False
        self.apply_theme()
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QHBoxLayout()
        
        self.sidebar = self.create_sidebar()
        main_layout.addWidget(self.sidebar)
        
        self.content_stack = QStackedWidget()
        main_layout.addWidget(self.content_stack, 1)
        
        self.dashboard_widget = self.create_dashboard()
        self.content_stack.addWidget(self.dashboard_widget)
        
        self.teachers_widget = self.create_teachers_management()
        self.content_stack.addWidget(self.teachers_widget)
        
        self.students_widget = self.create_students_management()
        self.content_stack.addWidget(self.students_widget)
        
        self.staff_widget = self.create_staff_management()
        self.content_stack.addWidget(self.staff_widget)
        
        self.attendance_widget = self.create_attendance_management()
        self.content_stack.addWidget(self.attendance_widget)
        
        self.reports_widget = self.create_reports()
        self.content_stack.addWidget(self.reports_widget)
        
        self.settings_widget = self.create_settings()
        self.content_stack.addWidget(self.settings_widget)
        
        central_widget.setLayout(main_layout)
        
        self.auto_refresh_timer = QTimer()
        self.auto_refresh_timer.timeout.connect(self.refresh_dashboard)
        self.auto_refresh_timer.start(60000)
    
    def apply_theme(self):
        if self.dark_mode:
            self.setStyleSheet("""
                QMainWindow, QWidget { background-color: #2b2b2b; color: #ffffff; }
                QPushButton { background-color: #3c3c3c; color: #ffffff; border: 1px solid #555555; padding: 8px; border-radius: 4px; }
                QPushButton:hover { background-color: #4c4c4c; }
                QLineEdit, QTextEdit, QComboBox, QSpinBox, QDateEdit, QTimeEdit { background-color: #3c3c3c; color: #ffffff; border: 1px solid #555555; padding: 5px; border-radius: 3px; }
                QTableWidget { background-color: #3c3c3c; color: #ffffff; gridline-color: #555555; }
                QHeaderView::section { background-color: #4c4c4c; color: #ffffff; padding: 5px; border: 1px solid #555555; }
                QGroupBox { border: 2px solid #555555; border-radius: 5px; margin-top: 10px; font-weight: bold; }
                QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; }
            """)
        else:
            self.setStyleSheet("""
                QMainWindow, QWidget { background-color: #f5f5f5; color: #333333; }
                QPushButton { background-color: #0078d4; color: white; border: none; padding: 8px; border-radius: 4px; font-weight: bold; }
                QPushButton:hover { background-color: #005a9e; }
                QLineEdit, QTextEdit, QComboBox, QSpinBox, QDateEdit, QTimeEdit { border: 1px solid #cccccc; padding: 5px; border-radius: 3px; }
                QTableWidget { background-color: white; gridline-color: #e0e0e0; }
                QHeaderView::section { background-color: #0078d4; color: white; padding: 5px; font-weight: bold; }
                QGroupBox { border: 2px solid #0078d4; border-radius: 5px; margin-top: 10px; font-weight: bold; }
                QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; }
            """)
    
    def create_sidebar(self):
        sidebar = QWidget()
        sidebar.setFixedWidth(200)
        layout = QVBoxLayout()
        
        logo_label = QLabel("BAMS")
        logo_label.setStyleSheet("font-size: 24px; font-weight: bold; padding: 20px; color: #0078d4;")
        logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(logo_label)
        
        user_label = QLabel(f"Welcome,\n{self.current_user['full_name']}")
        user_label.setStyleSheet("padding: 10px; font-size: 12px;")
        user_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(user_label)
        
        role_label = QLabel(f"Role: {self.current_user['role']}")
        role_label.setStyleSheet("padding: 5px; font-size: 11px; color: #666;")
        role_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(role_label)
        
        menu_buttons = [
            ("Dashboard", 0),
            ("Teachers", 1),
            ("Students", 2),
            ("Staff", 3),
            ("Attendance", 4),
            ("Reports", 5),
            ("Settings", 6)
        ]
        
        for text, index in menu_buttons:
            btn = QPushButton(text)
            btn.clicked.connect(lambda checked, idx=index: self.switch_page(idx))
            btn.setStyleSheet("text-align: left; padding: 12px;")
            layout.addWidget(btn)
        
        layout.addStretch()
        
        theme_btn = QPushButton("Toggle Theme")
        theme_btn.clicked.connect(self.toggle_theme)
        layout.addWidget(theme_btn)
        
        logout_btn = QPushButton("Logout")
        logout_btn.clicked.connect(self.logout)
        layout.addWidget(logout_btn)
        
        sidebar.setLayout(layout)
        return sidebar
    
    def switch_page(self, index):
        self.content_stack.setCurrentIndex(index)
        if index == 0:
            self.refresh_dashboard()
    
    def toggle_theme(self):
        self.dark_mode = not self.dark_mode
        self.apply_theme()
    
    def logout(self):
        reply = QMessageBox.question(self, 'Logout', 'Are you sure you want to logout?',
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.close()
    
    def create_dashboard(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        title = QLabel("Dashboard")
        title.setStyleSheet("font-size: 24px; font-weight: bold; padding: 10px;")
        layout.addWidget(title)
        
        stats_layout = QHBoxLayout()
        
        self.total_teachers_label = QLabel("Total Teachers\n0")
        self.total_teachers_label.setStyleSheet("background-color: #4CAF50; color: white; padding: 20px; border-radius: 8px; font-size: 18px; font-weight: bold;")
        self.total_teachers_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        stats_layout.addWidget(self.total_teachers_label)
        
        self.total_students_label = QLabel("Total Students\n0")
        self.total_students_label.setStyleSheet("background-color: #2196F3; color: white; padding: 20px; border-radius: 8px; font-size: 18px; font-weight: bold;")
        self.total_students_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        stats_layout.addWidget(self.total_students_label)
        
        self.total_staff_label = QLabel("Total Staff\n0")
        self.total_staff_label.setStyleSheet("background-color: #FF9800; color: white; padding: 20px; border-radius: 8px; font-size: 18px; font-weight: bold;")
        self.total_staff_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        stats_layout.addWidget(self.total_staff_label)
        
        self.present_today_label = QLabel("Present Today\n0")
        self.present_today_label.setStyleSheet("background-color: #9C27B0; color: white; padding: 20px; border-radius: 8px; font-size: 18px; font-weight: bold;")
        self.present_today_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        stats_layout.addWidget(self.present_today_label)
        
        layout.addLayout(stats_layout)
        
        device_status_label = QLabel("Device Status: " + ("✓ Connected" if self.biometric.is_device_connected() else "✗ Disconnected"))
        device_status_label.setStyleSheet("padding: 10px; font-size: 14px; font-weight: bold; color: " + ("#4CAF50" if self.biometric.is_device_connected() else "#f44336"))
        layout.addWidget(device_status_label)
        
        quick_actions_group = QGroupBox("Quick Actions")
        quick_actions_layout = QHBoxLayout()
        
        mark_attendance_btn = QPushButton("Mark Attendance (Biometric)")
        mark_attendance_btn.clicked.connect(self.quick_mark_attendance)
        quick_actions_layout.addWidget(mark_attendance_btn)
        
        view_today_btn = QPushButton("View Today's Attendance")
        view_today_btn.clicked.connect(self.view_today_attendance)
        quick_actions_layout.addWidget(view_today_btn)
        
        quick_actions_group.setLayout(quick_actions_layout)
        layout.addWidget(quick_actions_group)
        
        recent_group = QGroupBox("Recent Attendance")
        recent_layout = QVBoxLayout()
        
        self.recent_attendance_table = QTableWidget()
        self.recent_attendance_table.setColumnCount(5)
        self.recent_attendance_table.setHorizontalHeaderLabels(["Time", "Name", "Type", "Status", "Remarks"])
        self.recent_attendance_table.horizontalHeader().setStretchLastSection(True)
        recent_layout.addWidget(self.recent_attendance_table)
        
        recent_group.setLayout(recent_layout)
        layout.addWidget(recent_group)
        
        widget.setLayout(layout)
        return widget
    
    def refresh_dashboard(self):
        teachers = self.db.fetch_all("SELECT COUNT(*) as count FROM teachers WHERE status = 1")
        students = self.db.fetch_all("SELECT COUNT(*) as count FROM students WHERE status = 1")
        staff = self.db.fetch_all("SELECT COUNT(*) as count FROM staff WHERE status = 1")
        
        self.total_teachers_label.setText(f"Total Teachers\n{teachers[0]['count']}")
        self.total_students_label.setText(f"Total Students\n{students[0]['count']}")
        self.total_staff_label.setText(f"Total Staff\n{staff[0]['count']}")
        
        today = date.today().isoformat()
        present_count = self.db.fetch_all(
            "SELECT COUNT(DISTINCT person_type || '-' || person_id) as count FROM attendance_logs WHERE log_date = ? AND status = 'Present'",
            (today,)
        )
        self.present_today_label.setText(f"Present Today\n{present_count[0]['count']}")
        
        recent_logs = self.db.fetch_all(
            "SELECT * FROM attendance_logs ORDER BY created_at DESC LIMIT 10"
        )
        
        self.recent_attendance_table.setRowCount(len(recent_logs))
        for idx, log in enumerate(recent_logs):
            person_name = "Unknown"
            if log['person_type'] == 'teacher':
                person = self.db.fetch_one("SELECT name FROM teachers WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
            elif log['person_type'] == 'student':
                person = self.db.fetch_one("SELECT name FROM students WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
            elif log['person_type'] == 'staff':
                person = self.db.fetch_one("SELECT name FROM staff WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
            
            self.recent_attendance_table.setItem(idx, 0, QTableWidgetItem(log['created_at'][:19]))
            self.recent_attendance_table.setItem(idx, 1, QTableWidgetItem(person_name))
            self.recent_attendance_table.setItem(idx, 2, QTableWidgetItem(log['person_type'].title()))
            self.recent_attendance_table.setItem(idx, 3, QTableWidgetItem(log['status']))
            self.recent_attendance_table.setItem(idx, 4, QTableWidgetItem(log['remarks'] or ''))
    
    def quick_mark_attendance(self):
        dialog = BiometricVerificationDialog(self, self.biometric, self.db)
        dialog.verification_success.connect(self.process_attendance_marking)
        dialog.exec()
    
    def process_attendance_marking(self, person_type, person_id, person_name):
        today = date.today().isoformat()
        now = datetime.now()
        current_time = now.time().isoformat()[:8]
        
        existing_log = self.db.fetch_one(
            "SELECT * FROM attendance_logs WHERE person_type = ? AND person_id = ? AND log_date = ?",
            (person_type, person_id, today)
        )
        
        if existing_log:
            if existing_log['check_out_time']:
                QMessageBox.information(self, "Already Marked", f"{person_name} has already checked out today.")
                return
            
            self.db.execute_query(
                "UPDATE attendance_logs SET check_out_time = ?, remarks = ? WHERE id = ?",
                (current_time, "Check-out via biometric", existing_log['id'])
            )
            QMessageBox.information(self, "Success", f"Check-out marked for {person_name} at {current_time}")
        else:
            shift_id = None
            expected_time = None
            
            if person_type == 'teacher':
                person = self.db.fetch_one("SELECT shift_id FROM teachers WHERE id = ?", (person_id,))
                if person:
                    shift_id = person['shift_id']
            elif person_type == 'staff':
                person = self.db.fetch_one("SELECT shift_id FROM staff WHERE id = ?", (person_id,))
                if person:
                    shift_id = person['shift_id']
            
            is_late = 0
            if shift_id:
                shift = self.db.fetch_one("SELECT start_time, grace_time FROM shifts WHERE id = ?", (shift_id,))
                if shift:
                    expected_time = shift['start_time']
                    grace_time = shift['grace_time']
                    
                    shift_start = datetime.strptime(expected_time, '%H:%M').time()
                    grace_end = (datetime.combine(date.today(), shift_start) + timedelta(minutes=grace_time)).time()
                    current_time_obj = datetime.strptime(current_time, '%H:%M:%S').time()
                    
                    if current_time_obj > grace_end:
                        is_late = 1
            
            self.db.execute_query(
                "INSERT INTO attendance_logs (person_type, person_id, log_date, check_in_time, status, is_late, punch_type, remarks, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (person_type, person_id, today, current_time, 'Present', is_late, 'biometric', "Check-in via biometric", now.isoformat())
            )
            
            status_text = "on time" if not is_late else "LATE"
            QMessageBox.information(self, "Success", f"Check-in marked for {person_name} at {current_time} ({status_text})")
        
        self.refresh_dashboard()
    
    def view_today_attendance(self):
        today = date.today().isoformat()
        logs = self.db.fetch_all(
            "SELECT * FROM attendance_logs WHERE log_date = ? ORDER BY created_at DESC",
            (today,)
        )
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Today's Attendance")
        dialog.resize(800, 600)
        
        layout = QVBoxLayout()
        
        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(["Name", "Type", "Check-In", "Check-Out", "Status", "Remarks"])
        table.horizontalHeader().setStretchLastSection(True)
        
        table.setRowCount(len(logs))
        for idx, log in enumerate(logs):
            person_name = "Unknown"
            if log['person_type'] == 'teacher':
                person = self.db.fetch_one("SELECT name FROM teachers WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
            elif log['person_type'] == 'student':
                person = self.db.fetch_one("SELECT name FROM students WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
            elif log['person_type'] == 'staff':
                person = self.db.fetch_one("SELECT name FROM staff WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
            
            table.setItem(idx, 0, QTableWidgetItem(person_name))
            table.setItem(idx, 1, QTableWidgetItem(log['person_type'].title()))
            table.setItem(idx, 2, QTableWidgetItem(log['check_in_time'] or ''))
            table.setItem(idx, 3, QTableWidgetItem(log['check_out_time'] or ''))
            table.setItem(idx, 4, QTableWidgetItem(log['status']))
            table.setItem(idx, 5, QTableWidgetItem(log['remarks'] or ''))
        
        layout.addWidget(table)
        
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def create_teachers_management(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        title = QLabel("Teachers Management")
        title.setStyleSheet("font-size: 24px; font-weight: bold; padding: 10px;")
        layout.addWidget(title)
        
        toolbar_layout = QHBoxLayout()
        
        add_btn = QPushButton("Add Teacher")
        add_btn.clicked.connect(self.add_teacher)
        toolbar_layout.addWidget(add_btn)
        
        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(self.load_teachers)
        toolbar_layout.addWidget(refresh_btn)
        
        toolbar_layout.addStretch()
        layout.addLayout(toolbar_layout)
        
        self.teachers_table = QTableWidget()
        self.teachers_table.setColumnCount(9)
        self.teachers_table.setHorizontalHeaderLabels(["ID", "Teacher ID", "Name", "CNIC", "Phone", "Department", "Designation", "Status", "Actions"])
        self.teachers_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.teachers_table)
        
        widget.setLayout(layout)
        self.load_teachers()
        return widget
    
    def load_teachers(self):
        teachers = self.db.fetch_all('''
            SELECT t.*, d.name as dept_name 
            FROM teachers t 
            LEFT JOIN departments d ON t.department_id = d.id 
            ORDER BY t.created_at DESC
        ''')
        
        self.teachers_table.setRowCount(len(teachers))
        for idx, teacher in enumerate(teachers):
            self.teachers_table.setItem(idx, 0, QTableWidgetItem(str(teacher['id'])))
            self.teachers_table.setItem(idx, 1, QTableWidgetItem(teacher['teacher_id']))
            self.teachers_table.setItem(idx, 2, QTableWidgetItem(teacher['name']))
            self.teachers_table.setItem(idx, 3, QTableWidgetItem(teacher['cnic']))
            self.teachers_table.setItem(idx, 4, QTableWidgetItem(teacher['phone'] or ''))
            self.teachers_table.setItem(idx, 5, QTableWidgetItem(teacher['dept_name'] or ''))
            self.teachers_table.setItem(idx, 6, QTableWidgetItem(teacher['designation'] or ''))
            self.teachers_table.setItem(idx, 7, QTableWidgetItem("Active" if teacher['status'] else "Inactive"))
            
            actions_widget = QWidget()
            actions_layout = QHBoxLayout()
            actions_layout.setContentsMargins(0, 0, 0, 0)
            
            edit_btn = QPushButton("Edit")
            edit_btn.clicked.connect(lambda checked, t=teacher: self.edit_teacher(t))
            actions_layout.addWidget(edit_btn)
            
            biometric_btn = QPushButton("Biometric")
            biometric_btn.clicked.connect(lambda checked, t=teacher: self.enroll_teacher_biometric(t))
            actions_layout.addWidget(biometric_btn)
            
            delete_btn = QPushButton("Delete")
            delete_btn.clicked.connect(lambda checked, t=teacher: self.delete_teacher(t))
            actions_layout.addWidget(delete_btn)
            
            actions_widget.setLayout(actions_layout)
            self.teachers_table.setCellWidget(idx, 8, actions_widget)
    
    def add_teacher(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Teacher")
        dialog.resize(600, 700)
        
        layout = QVBoxLayout()
        form = QFormLayout()
        
        teacher_id_input = QLineEdit()
        teacher_id_input.setPlaceholderText("TCH001")
        form.addRow("Teacher ID:", teacher_id_input)
        
        name_input = QLineEdit()
        form.addRow("Name:", name_input)
        
        cnic_input = QLineEdit()
        cnic_input.setPlaceholderText("12345-1234567-1")
        form.addRow("CNIC:", cnic_input)
        
        phone_input = QLineEdit()
        form.addRow("Phone:", phone_input)
        
        email_input = QLineEdit()
        form.addRow("Email:", email_input)
        
        address_input = QTextEdit()
        address_input.setMaximumHeight(60)
        form.addRow("Address:", address_input)
        
        gender_combo = QComboBox()
        gender_combo.addItems(["Male", "Female", "Other"])
        form.addRow("Gender:", gender_combo)
        
        dob_input = QDateEdit()
        dob_input.setCalendarPopup(True)
        dob_input.setDate(QDate.currentDate().addYears(-25))
        form.addRow("Date of Birth:", dob_input)
        
        qualification_input = QLineEdit()
        form.addRow("Qualification:", qualification_input)
        
        departments = self.db.fetch_all("SELECT id, name FROM departments WHERE is_active = 1")
        dept_combo = QComboBox()
        for dept in departments:
            dept_combo.addItem(dept['name'], dept['id'])
        form.addRow("Department:", dept_combo)
        
        designation_input = QLineEdit()
        form.addRow("Designation:", designation_input)
        
        shifts = self.db.fetch_all("SELECT id, name FROM shifts WHERE is_active = 1")
        shift_combo = QComboBox()
        for shift in shifts:
            shift_combo.addItem(shift['name'], shift['id'])
        form.addRow("Shift:", shift_combo)
        
        joining_date = QDateEdit()
        joining_date.setCalendarPopup(True)
        joining_date.setDate(QDate.currentDate())
        form.addRow("Joining Date:", joining_date)
        
        layout.addLayout(form)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(lambda: self.save_teacher(dialog, {
            'teacher_id': teacher_id_input.text(),
            'name': name_input.text(),
            'cnic': cnic_input.text(),
            'phone': phone_input.text(),
            'email': email_input.text(),
            'address': address_input.toPlainText(),
            'gender': gender_combo.currentText(),
            'dob': dob_input.date().toString('yyyy-MM-dd'),
            'qualification': qualification_input.text(),
            'department_id': dept_combo.currentData(),
            'designation': designation_input.text(),
            'shift_id': shift_combo.currentData(),
            'joining_date': joining_date.date().toString('yyyy-MM-dd')
        }))
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def save_teacher(self, dialog, data):
        if not data['teacher_id'] or not data['name'] or not data['cnic']:
            QMessageBox.warning(dialog, "Validation Error", "Teacher ID, Name, and CNIC are required.")
            return
        
        try:
            self.db.execute_query('''
                INSERT INTO teachers (teacher_id, name, cnic, phone, email, address, gender, date_of_birth, 
                                    qualification, department_id, designation, shift_id, joining_date, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (data['teacher_id'], data['name'], data['cnic'], data['phone'], data['email'], 
                  data['address'], data['gender'], data['dob'], data['qualification'], 
                  data['department_id'], data['designation'], data['shift_id'], 
                  data['joining_date'], datetime.now().isoformat()))
            
            QMessageBox.information(dialog, "Success", "Teacher added successfully!")
            dialog.accept()
            self.load_teachers()
        except Exception as e:
            QMessageBox.critical(dialog, "Error", f"Failed to add teacher: {str(e)}")
    
    def edit_teacher(self, teacher):
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Teacher")
        dialog.resize(600, 700)
        
        layout = QVBoxLayout()
        form = QFormLayout()
        
        teacher_id_input = QLineEdit(teacher['teacher_id'])
        teacher_id_input.setReadOnly(True)
        form.addRow("Teacher ID:", teacher_id_input)
        
        name_input = QLineEdit(teacher['name'])
        form.addRow("Name:", name_input)
        
        cnic_input = QLineEdit(teacher['cnic'])
        form.addRow("CNIC:", cnic_input)
        
        phone_input = QLineEdit(teacher['phone'] or '')
        form.addRow("Phone:", phone_input)
        
        email_input = QLineEdit(teacher['email'] or '')
        form.addRow("Email:", email_input)
        
        address_input = QTextEdit(teacher['address'] or '')
        address_input.setMaximumHeight(60)
        form.addRow("Address:", address_input)
        
        gender_combo = QComboBox()
        gender_combo.addItems(["Male", "Female", "Other"])
        if teacher['gender']:
            gender_combo.setCurrentText(teacher['gender'])
        form.addRow("Gender:", gender_combo)
        
        dob_input = QDateEdit()
        dob_input.setCalendarPopup(True)
        if teacher['date_of_birth']:
            dob_input.setDate(QDate.fromString(teacher['date_of_birth'], 'yyyy-MM-dd'))
        form.addRow("Date of Birth:", dob_input)
        
        qualification_input = QLineEdit(teacher['qualification'] or '')
        form.addRow("Qualification:", qualification_input)
        
        departments = self.db.fetch_all("SELECT id, name FROM departments WHERE is_active = 1")
        dept_combo = QComboBox()
        for dept in departments:
            dept_combo.addItem(dept['name'], dept['id'])
        if teacher['department_id']:
            dept_combo.setCurrentIndex(dept_combo.findData(teacher['department_id']))
        form.addRow("Department:", dept_combo)
        
        designation_input = QLineEdit(teacher['designation'] or '')
        form.addRow("Designation:", designation_input)
        
        shifts = self.db.fetch_all("SELECT id, name FROM shifts WHERE is_active = 1")
        shift_combo = QComboBox()
        for shift in shifts:
            shift_combo.addItem(shift['name'], shift['id'])
        if teacher['shift_id']:
            shift_combo.setCurrentIndex(shift_combo.findData(teacher['shift_id']))
        form.addRow("Shift:", shift_combo)
        
        joining_date = QDateEdit()
        joining_date.setCalendarPopup(True)
        if teacher['joining_date']:
            joining_date.setDate(QDate.fromString(teacher['joining_date'], 'yyyy-MM-dd'))
        form.addRow("Joining Date:", joining_date)
        
        status_check = QCheckBox("Active")
        status_check.setChecked(teacher['status'] == 1)
        form.addRow("Status:", status_check)
        
        layout.addLayout(form)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(lambda: self.update_teacher(dialog, teacher['id'], {
            'name': name_input.text(),
            'cnic': cnic_input.text(),
            'phone': phone_input.text(),
            'email': email_input.text(),
            'address': address_input.toPlainText(),
            'gender': gender_combo.currentText(),
            'dob': dob_input.date().toString('yyyy-MM-dd'),
            'qualification': qualification_input.text(),
            'department_id': dept_combo.currentData(),
            'designation': designation_input.text(),
            'shift_id': shift_combo.currentData(),
            'joining_date': joining_date.date().toString('yyyy-MM-dd'),
            'status': 1 if status_check.isChecked() else 0
        }))
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def update_teacher(self, dialog, teacher_id, data):
        if not data['name'] or not data['cnic']:
            QMessageBox.warning(dialog, "Validation Error", "Name and CNIC are required.")
            return
        
        try:
            self.db.execute_query('''
                UPDATE teachers SET name = ?, cnic = ?, phone = ?, email = ?, address = ?, gender = ?, 
                                  date_of_birth = ?, qualification = ?, department_id = ?, designation = ?, 
                                  shift_id = ?, joining_date = ?, status = ? WHERE id = ?
            ''', (data['name'], data['cnic'], data['phone'], data['email'], data['address'], 
                  data['gender'], data['dob'], data['qualification'], data['department_id'], 
                  data['designation'], data['shift_id'], data['joining_date'], data['status'], teacher_id))
            
            QMessageBox.information(dialog, "Success", "Teacher updated successfully!")
            dialog.accept()
            self.load_teachers()
        except Exception as e:
            QMessageBox.critical(dialog, "Error", f"Failed to update teacher: {str(e)}")
    
    def enroll_teacher_biometric(self, teacher):
        dialog = BiometricEnrollmentDialog(self, self.biometric, 'teacher', teacher['id'], teacher['name'])
        dialog.exec()
    
    def delete_teacher(self, teacher):
        reply = QMessageBox.question(self, 'Delete Teacher', 
                                    f"Are you sure you want to delete {teacher['name']}?",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.db.execute_query("DELETE FROM teachers WHERE id = ?", (teacher['id'],))
                self.db.execute_query("DELETE FROM biometric_data WHERE person_type = 'teacher' AND person_id = ?", (teacher['id'],))
                QMessageBox.information(self, "Success", "Teacher deleted successfully!")
                self.load_teachers()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete teacher: {str(e)}")
    
    def create_students_management(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        title = QLabel("Students Management")
        title.setStyleSheet("font-size: 24px; font-weight: bold; padding: 10px;")
        layout.addWidget(title)
        
        toolbar_layout = QHBoxLayout()
        
        add_btn = QPushButton("Add Student")
        add_btn.clicked.connect(self.add_student)
        toolbar_layout.addWidget(add_btn)
        
        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(self.load_students)
        toolbar_layout.addWidget(refresh_btn)
        
        toolbar_layout.addStretch()
        layout.addLayout(toolbar_layout)
        
        self.students_table = QTableWidget()
        self.students_table.setColumnCount(9)
        self.students_table.setHorizontalHeaderLabels(["ID", "Roll No", "Name", "Father Name", "Class", "Section", "Phone", "Status", "Actions"])
        self.students_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.students_table)
        
        widget.setLayout(layout)
        self.load_students()
        return widget
    
    def load_students(self):
        students = self.db.fetch_all('''
            SELECT s.*, c.name as class_name, sec.name as section_name 
            FROM students s 
            LEFT JOIN classes c ON s.class_id = c.id 
            LEFT JOIN sections sec ON s.section_id = sec.id 
            ORDER BY s.created_at DESC
        ''')
        
        self.students_table.setRowCount(len(students))
        for idx, student in enumerate(students):
            self.students_table.setItem(idx, 0, QTableWidgetItem(str(student['id'])))
            self.students_table.setItem(idx, 1, QTableWidgetItem(student['roll_no']))
            self.students_table.setItem(idx, 2, QTableWidgetItem(student['name']))
            self.students_table.setItem(idx, 3, QTableWidgetItem(student['father_name'] or ''))
            self.students_table.setItem(idx, 4, QTableWidgetItem(student['class_name'] or ''))
            self.students_table.setItem(idx, 5, QTableWidgetItem(student['section_name'] or ''))
            self.students_table.setItem(idx, 6, QTableWidgetItem(student['phone'] or ''))
            self.students_table.setItem(idx, 7, QTableWidgetItem("Active" if student['status'] else "Inactive"))
            
            actions_widget = QWidget()
            actions_layout = QHBoxLayout()
            actions_layout.setContentsMargins(0, 0, 0, 0)
            
            edit_btn = QPushButton("Edit")
            edit_btn.clicked.connect(lambda checked, s=student: self.edit_student(s))
            actions_layout.addWidget(edit_btn)
            
            biometric_btn = QPushButton("Biometric")
            biometric_btn.clicked.connect(lambda checked, s=student: self.enroll_student_biometric(s))
            actions_layout.addWidget(biometric_btn)
            
            delete_btn = QPushButton("Delete")
            delete_btn.clicked.connect(lambda checked, s=student: self.delete_student(s))
            actions_layout.addWidget(delete_btn)
            
            actions_widget.setLayout(actions_layout)
            self.students_table.setCellWidget(idx, 8, actions_widget)
    
    def add_student(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Student")
        dialog.resize(600, 700)
        
        layout = QVBoxLayout()
        form = QFormLayout()
        
        roll_no_input = QLineEdit()
        roll_no_input.setPlaceholderText("STD001")
        form.addRow("Roll No:", roll_no_input)
        
        name_input = QLineEdit()
        form.addRow("Name:", name_input)
        
        father_name_input = QLineEdit()
        form.addRow("Father Name:", father_name_input)
        
        cnic_input = QLineEdit()
        cnic_input.setPlaceholderText("12345-1234567-1")
        form.addRow("CNIC/B-Form:", cnic_input)
        
        phone_input = QLineEdit()
        form.addRow("Phone:", phone_input)
        
        address_input = QTextEdit()
        address_input.setMaximumHeight(60)
        form.addRow("Address:", address_input)
        
        gender_combo = QComboBox()
        gender_combo.addItems(["Male", "Female"])
        form.addRow("Gender:", gender_combo)
        
        dob_input = QDateEdit()
        dob_input.setCalendarPopup(True)
        dob_input.setDate(QDate.currentDate().addYears(-10))
        form.addRow("Date of Birth:", dob_input)
        
        classes = self.db.fetch_all("SELECT id, name FROM classes WHERE is_active = 1")
        class_combo = QComboBox()
        for cls in classes:
            class_combo.addItem(cls['name'], cls['id'])
        form.addRow("Class:", class_combo)
        
        section_combo = QComboBox()
        
        def load_sections():
            section_combo.clear()
            class_id = class_combo.currentData()
            if class_id:
                sections = self.db.fetch_all("SELECT id, name FROM sections WHERE class_id = ? AND is_active = 1", (class_id,))
                for sec in sections:
                    section_combo.addItem(sec['name'], sec['id'])
        
        class_combo.currentIndexChanged.connect(load_sections)
        load_sections()
        form.addRow("Section:", section_combo)
        
        sessions = self.db.fetch_all("SELECT id, name FROM sessions ORDER BY is_current DESC")
        session_combo = QComboBox()
        for sess in sessions:
            session_combo.addItem(sess['name'], sess['id'])
        form.addRow("Session:", session_combo)
        
        admission_date = QDateEdit()
        admission_date.setCalendarPopup(True)
        admission_date.setDate(QDate.currentDate())
        form.addRow("Admission Date:", admission_date)
        
        layout.addLayout(form)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(lambda: self.save_student(dialog, {
            'roll_no': roll_no_input.text(),
            'name': name_input.text(),
            'father_name': father_name_input.text(),
            'cnic_bform': cnic_input.text(),
            'phone': phone_input.text(),
            'address': address_input.toPlainText(),
            'gender': gender_combo.currentText(),
            'dob': dob_input.date().toString('yyyy-MM-dd'),
            'class_id': class_combo.currentData(),
            'section_id': section_combo.currentData(),
            'session_id': session_combo.currentData(),
            'admission_date': admission_date.date().toString('yyyy-MM-dd')
        }))
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def save_student(self, dialog, data):
        if not data['roll_no'] or not data['name']:
            QMessageBox.warning(dialog, "Validation Error", "Roll No and Name are required.")
            return
        
        try:
            self.db.execute_query('''
                INSERT INTO students (roll_no, name, father_name, cnic_bform, phone, address, gender, 
                                    date_of_birth, class_id, section_id, session_id, admission_date, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (data['roll_no'], data['name'], data['father_name'], data['cnic_bform'], 
                  data['phone'], data['address'], data['gender'], data['dob'], 
                  data['class_id'], data['section_id'], data['session_id'], 
                  data['admission_date'], datetime.now().isoformat()))
            
            QMessageBox.information(dialog, "Success", "Student added successfully!")
            dialog.accept()
            self.load_students()
        except Exception as e:
            QMessageBox.critical(dialog, "Error", f"Failed to add student: {str(e)}")
    
    def edit_student(self, student):
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Student")
        dialog.resize(600, 700)
        
        layout = QVBoxLayout()
        form = QFormLayout()
        
        roll_no_input = QLineEdit(student['roll_no'])
        roll_no_input.setReadOnly(True)
        form.addRow("Roll No:", roll_no_input)
        
        name_input = QLineEdit(student['name'])
        form.addRow("Name:", name_input)
        
        father_name_input = QLineEdit(student['father_name'] or '')
        form.addRow("Father Name:", father_name_input)
        
        cnic_input = QLineEdit(student['cnic_bform'] or '')
        form.addRow("CNIC/B-Form:", cnic_input)
        
        phone_input = QLineEdit(student['phone'] or '')
        form.addRow("Phone:", phone_input)
        
        address_input = QTextEdit(student['address'] or '')
        address_input.setMaximumHeight(60)
        form.addRow("Address:", address_input)
        
        gender_combo = QComboBox()
        gender_combo.addItems(["Male", "Female"])
        if student['gender']:
            gender_combo.setCurrentText(student['gender'])
        form.addRow("Gender:", gender_combo)
        
        dob_input = QDateEdit()
        dob_input.setCalendarPopup(True)
        if student['date_of_birth']:
            dob_input.setDate(QDate.fromString(student['date_of_birth'], 'yyyy-MM-dd'))
        form.addRow("Date of Birth:", dob_input)
        
        classes = self.db.fetch_all("SELECT id, name FROM classes WHERE is_active = 1")
        class_combo = QComboBox()
        for cls in classes:
            class_combo.addItem(cls['name'], cls['id'])
        if student['class_id']:
            class_combo.setCurrentIndex(class_combo.findData(student['class_id']))
        form.addRow("Class:", class_combo)
        
        section_combo = QComboBox()
        
        def load_sections():
            section_combo.clear()
            class_id = class_combo.currentData()
            if class_id:
                sections = self.db.fetch_all("SELECT id, name FROM sections WHERE class_id = ? AND is_active = 1", (class_id,))
                for sec in sections:
                    section_combo.addItem(sec['name'], sec['id'])
                if student['section_id']:
                    section_combo.setCurrentIndex(section_combo.findData(student['section_id']))
        
        class_combo.currentIndexChanged.connect(load_sections)
        load_sections()
        form.addRow("Section:", section_combo)
        
        sessions = self.db.fetch_all("SELECT id, name FROM sessions ORDER BY is_current DESC")
        session_combo = QComboBox()
        for sess in sessions:
            session_combo.addItem(sess['name'], sess['id'])
        if student['session_id']:
            session_combo.setCurrentIndex(session_combo.findData(student['session_id']))
        form.addRow("Session:", session_combo)
        
        admission_date = QDateEdit()
        admission_date.setCalendarPopup(True)
        if student['admission_date']:
            admission_date.setDate(QDate.fromString(student['admission_date'], 'yyyy-MM-dd'))
        form.addRow("Admission Date:", admission_date)
        
        status_check = QCheckBox("Active")
        status_check.setChecked(student['status'] == 1)
        form.addRow("Status:", status_check)
        
        layout.addLayout(form)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(lambda: self.update_student(dialog, student['id'], {
            'name': name_input.text(),
            'father_name': father_name_input.text(),
            'cnic_bform': cnic_input.text(),
            'phone': phone_input.text(),
            'address': address_input.toPlainText(),
            'gender': gender_combo.currentText(),
            'dob': dob_input.date().toString('yyyy-MM-dd'),
            'class_id': class_combo.currentData(),
            'section_id': section_combo.currentData(),
            'session_id': session_combo.currentData(),
            'admission_date': admission_date.date().toString('yyyy-MM-dd'),
            'status': 1 if status_check.isChecked() else 0
        }))
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def update_student(self, dialog, student_id, data):
        if not data['name']:
            QMessageBox.warning(dialog, "Validation Error", "Name is required.")
            return
        
        try:
            self.db.execute_query('''
                UPDATE students SET name = ?, father_name = ?, cnic_bform = ?, phone = ?, address = ?, 
                                  gender = ?, date_of_birth = ?, class_id = ?, section_id = ?, 
                                  session_id = ?, admission_date = ?, status = ? WHERE id = ?
            ''', (data['name'], data['father_name'], data['cnic_bform'], data['phone'], 
                  data['address'], data['gender'], data['dob'], data['class_id'], 
                  data['section_id'], data['session_id'], data['admission_date'], 
                  data['status'], student_id))
            
            QMessageBox.information(dialog, "Success", "Student updated successfully!")
            dialog.accept()
            self.load_students()
        except Exception as e:
            QMessageBox.critical(dialog, "Error", f"Failed to update student: {str(e)}")
    
    def enroll_student_biometric(self, student):
        dialog = BiometricEnrollmentDialog(self, self.biometric, 'student', student['id'], student['name'])
        dialog.exec()
    
    def delete_student(self, student):
        reply = QMessageBox.question(self, 'Delete Student', 
                                    f"Are you sure you want to delete {student['name']}?",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.db.execute_query("DELETE FROM students WHERE id = ?", (student['id'],))
                self.db.execute_query("DELETE FROM biometric_data WHERE person_type = 'student' AND person_id = ?", (student['id'],))
                QMessageBox.information(self, "Success", "Student deleted successfully!")
                self.load_students()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete student: {str(e)}")
    
    def create_staff_management(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        title = QLabel("Staff Management")
        title.setStyleSheet("font-size: 24px; font-weight: bold; padding: 10px;")
        layout.addWidget(title)
        
        toolbar_layout = QHBoxLayout()
        
        add_btn = QPushButton("Add Staff")
        add_btn.clicked.connect(self.add_staff)
        toolbar_layout.addWidget(add_btn)
        
        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(self.load_staff)
        toolbar_layout.addWidget(refresh_btn)
        
        toolbar_layout.addStretch()
        layout.addLayout(toolbar_layout)
        
        self.staff_table = QTableWidget()
        self.staff_table.setColumnCount(8)
        self.staff_table.setHorizontalHeaderLabels(["ID", "Staff ID", "Name", "CNIC", "Phone", "Role", "Status", "Actions"])
        self.staff_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.staff_table)
        
        widget.setLayout(layout)
        self.load_staff()
        return widget
    
    def load_staff(self):
        staff_members = self.db.fetch_all("SELECT * FROM staff ORDER BY created_at DESC")
        
        self.staff_table.setRowCount(len(staff_members))
        for idx, staff in enumerate(staff_members):
            self.staff_table.setItem(idx, 0, QTableWidgetItem(str(staff['id'])))
            self.staff_table.setItem(idx, 1, QTableWidgetItem(staff['staff_id']))
            self.staff_table.setItem(idx, 2, QTableWidgetItem(staff['name']))
            self.staff_table.setItem(idx, 3, QTableWidgetItem(staff['cnic']))
            self.staff_table.setItem(idx, 4, QTableWidgetItem(staff['phone'] or ''))
            self.staff_table.setItem(idx, 5, QTableWidgetItem(staff['role']))
            self.staff_table.setItem(idx, 6, QTableWidgetItem("Active" if staff['status'] else "Inactive"))
            
            actions_widget = QWidget()
            actions_layout = QHBoxLayout()
            actions_layout.setContentsMargins(0, 0, 0, 0)
            
            edit_btn = QPushButton("Edit")
            edit_btn.clicked.connect(lambda checked, s=staff: self.edit_staff(s))
            actions_layout.addWidget(edit_btn)
            
            biometric_btn = QPushButton("Biometric")
            biometric_btn.clicked.connect(lambda checked, s=staff: self.enroll_staff_biometric(s))
            actions_layout.addWidget(biometric_btn)
            
            delete_btn = QPushButton("Delete")
            delete_btn.clicked.connect(lambda checked, s=staff: self.delete_staff(s))
            actions_layout.addWidget(delete_btn)
            
            actions_widget.setLayout(actions_layout)
            self.staff_table.setCellWidget(idx, 7, actions_widget)
    
    def add_staff(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Staff")
        dialog.resize(600, 600)
        
        layout = QVBoxLayout()
        form = QFormLayout()
        
        staff_id_input = QLineEdit()
        staff_id_input.setPlaceholderText("STF001")
        form.addRow("Staff ID:", staff_id_input)
        
        name_input = QLineEdit()
        form.addRow("Name:", name_input)
        
        cnic_input = QLineEdit()
        cnic_input.setPlaceholderText("12345-1234567-1")
        form.addRow("CNIC:", cnic_input)
        
        phone_input = QLineEdit()
        form.addRow("Phone:", phone_input)
        
        email_input = QLineEdit()
        form.addRow("Email:", email_input)
        
        address_input = QTextEdit()
        address_input.setMaximumHeight(60)
        form.addRow("Address:", address_input)
        
        gender_combo = QComboBox()
        gender_combo.addItems(["Male", "Female", "Other"])
        form.addRow("Gender:", gender_combo)
        
        dob_input = QDateEdit()
        dob_input.setCalendarPopup(True)
        dob_input.setDate(QDate.currentDate().addYears(-25))
        form.addRow("Date of Birth:", dob_input)
        
        role_input = QLineEdit()
        form.addRow("Role:", role_input)
        
        shifts = self.db.fetch_all("SELECT id, name FROM shifts WHERE is_active = 1")
        shift_combo = QComboBox()
        for shift in shifts:
            shift_combo.addItem(shift['name'], shift['id'])
        form.addRow("Shift:", shift_combo)
        
        salary_combo = QComboBox()
        salary_combo.addItems(["Monthly", "Daily", "Hourly"])
        form.addRow("Salary Type:", salary_combo)
        
        joining_date = QDateEdit()
        joining_date.setCalendarPopup(True)
        joining_date.setDate(QDate.currentDate())
        form.addRow("Joining Date:", joining_date)
        
        layout.addLayout(form)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(lambda: self.save_staff(dialog, {
            'staff_id': staff_id_input.text(),
            'name': name_input.text(),
            'cnic': cnic_input.text(),
            'phone': phone_input.text(),
            'email': email_input.text(),
            'address': address_input.toPlainText(),
            'gender': gender_combo.currentText(),
            'dob': dob_input.date().toString('yyyy-MM-dd'),
            'role': role_input.text(),
            'shift_id': shift_combo.currentData(),
            'salary_type': salary_combo.currentText(),
            'joining_date': joining_date.date().toString('yyyy-MM-dd')
        }))
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def save_staff(self, dialog, data):
        if not data['staff_id'] or not data['name'] or not data['cnic']:
            QMessageBox.warning(dialog, "Validation Error", "Staff ID, Name, and CNIC are required.")
            return
        
        try:
            self.db.execute_query('''
                INSERT INTO staff (staff_id, name, cnic, phone, email, address, gender, date_of_birth, 
                                 role, shift_id, salary_type, joining_date, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (data['staff_id'], data['name'], data['cnic'], data['phone'], data['email'], 
                  data['address'], data['gender'], data['dob'], data['role'], 
                  data['shift_id'], data['salary_type'], data['joining_date'], 
                  datetime.now().isoformat()))
            
            QMessageBox.information(dialog, "Success", "Staff added successfully!")
            dialog.accept()
            self.load_staff()
        except Exception as e:
            QMessageBox.critical(dialog, "Error", f"Failed to add staff: {str(e)}")
    
    def edit_staff(self, staff):
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Staff")
        dialog.resize(600, 600)
        
        layout = QVBoxLayout()
        form = QFormLayout()
        
        staff_id_input = QLineEdit(staff['staff_id'])
        staff_id_input.setReadOnly(True)
        form.addRow("Staff ID:", staff_id_input)
        
        name_input = QLineEdit(staff['name'])
        form.addRow("Name:", name_input)
        
        cnic_input = QLineEdit(staff['cnic'])
        form.addRow("CNIC:", cnic_input)
        
        phone_input = QLineEdit(staff['phone'] or '')
        form.addRow("Phone:", phone_input)
        
        email_input = QLineEdit(staff['email'] or '')
        form.addRow("Email:", email_input)
        
        address_input = QTextEdit(staff['address'] or '')
        address_input.setMaximumHeight(60)
        form.addRow("Address:", address_input)
        
        gender_combo = QComboBox()
        gender_combo.addItems(["Male", "Female", "Other"])
        if staff['gender']:
            gender_combo.setCurrentText(staff['gender'])
        form.addRow("Gender:", gender_combo)
        
        dob_input = QDateEdit()
        dob_input.setCalendarPopup(True)
        if staff['date_of_birth']:
            dob_input.setDate(QDate.fromString(staff['date_of_birth'], 'yyyy-MM-dd'))
        form.addRow("Date of Birth:", dob_input)
        
        role_input = QLineEdit(staff['role'])
        form.addRow("Role:", role_input)
        
        shifts = self.db.fetch_all("SELECT id, name FROM shifts WHERE is_active = 1")
        shift_combo = QComboBox()
        for shift in shifts:
            shift_combo.addItem(shift['name'], shift['id'])
        if staff['shift_id']:
            shift_combo.setCurrentIndex(shift_combo.findData(staff['shift_id']))
        form.addRow("Shift:", shift_combo)
        
        salary_combo = QComboBox()
        salary_combo.addItems(["Monthly", "Daily", "Hourly"])
        if staff['salary_type']:
            salary_combo.setCurrentText(staff['salary_type'])
        form.addRow("Salary Type:", salary_combo)
        
        joining_date = QDateEdit()
        joining_date.setCalendarPopup(True)
        if staff['joining_date']:
            joining_date.setDate(QDate.fromString(staff['joining_date'], 'yyyy-MM-dd'))
        form.addRow("Joining Date:", joining_date)
        
        status_check = QCheckBox("Active")
        status_check.setChecked(staff['status'] == 1)
        form.addRow("Status:", status_check)
        
        layout.addLayout(form)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(lambda: self.update_staff(dialog, staff['id'], {
            'name': name_input.text(),
            'cnic': cnic_input.text(),
            'phone': phone_input.text(),
            'email': email_input.text(),
            'address': address_input.toPlainText(),
            'gender': gender_combo.currentText(),
            'dob': dob_input.date().toString('yyyy-MM-dd'),
            'role': role_input.text(),
            'shift_id': shift_combo.currentData(),
            'salary_type': salary_combo.currentText(),
            'joining_date': joining_date.date().toString('yyyy-MM-dd'),
            'status': 1 if status_check.isChecked() else 0
        }))
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def update_staff(self, dialog, staff_id, data):
        if not data['name'] or not data['cnic']:
            QMessageBox.warning(dialog, "Validation Error", "Name and CNIC are required.")
            return
        
        try:
            self.db.execute_query('''
                UPDATE staff SET name = ?, cnic = ?, phone = ?, email = ?, address = ?, gender = ?, 
                               date_of_birth = ?, role = ?, shift_id = ?, salary_type = ?, 
                               joining_date = ?, status = ? WHERE id = ?
            ''', (data['name'], data['cnic'], data['phone'], data['email'], data['address'], 
                  data['gender'], data['dob'], data['role'], data['shift_id'], 
                  data['salary_type'], data['joining_date'], data['status'], staff_id))
            
            QMessageBox.information(dialog, "Success", "Staff updated successfully!")
            dialog.accept()
            self.load_staff()
        except Exception as e:
            QMessageBox.critical(dialog, "Error", f"Failed to update staff: {str(e)}")
    
    def enroll_staff_biometric(self, staff):
        dialog = BiometricEnrollmentDialog(self, self.biometric, 'staff', staff['id'], staff['name'])
        dialog.exec()
    
    def delete_staff(self, staff):
        reply = QMessageBox.question(self, 'Delete Staff', 
                                    f"Are you sure you want to delete {staff['name']}?",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.db.execute_query("DELETE FROM staff WHERE id = ?", (staff['id'],))
                self.db.execute_query("DELETE FROM biometric_data WHERE person_type = 'staff' AND person_id = ?", (staff['id'],))
                QMessageBox.information(self, "Success", "Staff deleted successfully!")
                self.load_staff()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete staff: {str(e)}")
    
    def create_attendance_management(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        title = QLabel("Attendance Management")
        title.setStyleSheet("font-size: 24px; font-weight: bold; padding: 10px;")
        layout.addWidget(title)
        
        tabs = QTabWidget()
        
        mark_tab = self.create_mark_attendance_tab()
        tabs.addTab(mark_tab, "Mark Attendance")
        
        view_tab = self.create_view_attendance_tab()
        tabs.addTab(view_tab, "View Attendance")
        
        manual_tab = self.create_manual_attendance_tab()
        tabs.addTab(manual_tab, "Manual Entry")
        
        layout.addWidget(tabs)
        widget.setLayout(layout)
        return widget
    
    def create_mark_attendance_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        info_label = QLabel("Use biometric scanner to mark attendance")
        info_label.setStyleSheet("font-size: 16px; padding: 20px;")
        layout.addWidget(info_label)
        
        scan_btn = QPushButton("Scan Fingerprint")
        scan_btn.setStyleSheet("padding: 20px; font-size: 18px;")
        scan_btn.clicked.connect(self.quick_mark_attendance)
        layout.addWidget(scan_btn)
        
        layout.addStretch()
        widget.setLayout(layout)
        return widget
    
    def create_view_attendance_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        filter_layout = QHBoxLayout()
        
        filter_layout.addWidget(QLabel("Date:"))
        self.attendance_date_filter = QDateEdit()
        self.attendance_date_filter.setCalendarPopup(True)
        self.attendance_date_filter.setDate(QDate.currentDate())
        filter_layout.addWidget(self.attendance_date_filter)
        
        filter_layout.addWidget(QLabel("Type:"))
        self.attendance_type_filter = QComboBox()
        self.attendance_type_filter.addItems(["All", "Teacher", "Student", "Staff"])
        filter_layout.addWidget(self.attendance_type_filter)
        
        filter_btn = QPushButton("Filter")
        filter_btn.clicked.connect(self.filter_attendance)
        filter_layout.addWidget(filter_btn)
        
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        
        self.attendance_view_table = QTableWidget()
        self.attendance_view_table.setColumnCount(7)
        self.attendance_view_table.setHorizontalHeaderLabels(["Name", "Type", "Check-In", "Check-Out", "Status", "Late", "Remarks"])
        self.attendance_view_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.attendance_view_table)
        
        widget.setLayout(layout)
        return widget
    
    def filter_attendance(self):
        selected_date = self.attendance_date_filter.date().toString('yyyy-MM-dd')
        selected_type = self.attendance_type_filter.currentText().lower()
        
        if selected_type == "all":
            logs = self.db.fetch_all(
                "SELECT * FROM attendance_logs WHERE log_date = ? ORDER BY created_at DESC",
                (selected_date,)
            )
        else:
            logs = self.db.fetch_all(
                "SELECT * FROM attendance_logs WHERE log_date = ? AND person_type = ? ORDER BY created_at DESC",
                (selected_date, selected_type)
            )
        
        self.attendance_view_table.setRowCount(len(logs))
        for idx, log in enumerate(logs):
            person_name = "Unknown"
            if log['person_type'] == 'teacher':
                person = self.db.fetch_one("SELECT name FROM teachers WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
            elif log['person_type'] == 'student':
                person = self.db.fetch_one("SELECT name FROM students WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
            elif log['person_type'] == 'staff':
                person = self.db.fetch_one("SELECT name FROM staff WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
            
            self.attendance_view_table.setItem(idx, 0, QTableWidgetItem(person_name))
            self.attendance_view_table.setItem(idx, 1, QTableWidgetItem(log['person_type'].title()))
            self.attendance_view_table.setItem(idx, 2, QTableWidgetItem(log['check_in_time'] or ''))
            self.attendance_view_table.setItem(idx, 3, QTableWidgetItem(log['check_out_time'] or ''))
            self.attendance_view_table.setItem(idx, 4, QTableWidgetItem(log['status']))
            self.attendance_view_table.setItem(idx, 5, QTableWidgetItem("Yes" if log['is_late'] else "No"))
            self.attendance_view_table.setItem(idx, 6, QTableWidgetItem(log['remarks'] or ''))
    
    def create_manual_attendance_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        form = QFormLayout()
        
        type_combo = QComboBox()
        type_combo.addItems(["Teacher", "Student", "Staff"])
        form.addRow("Person Type:", type_combo)
        
        person_combo = QComboBox()
        
        def load_persons():
            person_combo.clear()
            person_type = type_combo.currentText().lower()
            
            if person_type == 'teacher':
                persons = self.db.fetch_all("SELECT id, name, teacher_id FROM teachers WHERE status = 1")
                for p in persons:
                    person_combo.addItem(f"{p['name']} ({p['teacher_id']})", p['id'])
            elif person_type == 'student':
                persons = self.db.fetch_all("SELECT id, name, roll_no FROM students WHERE status = 1")
                for p in persons:
                    person_combo.addItem(f"{p['name']} ({p['roll_no']})", p['id'])
            elif person_type == 'staff':
                persons = self.db.fetch_all("SELECT id, name, staff_id FROM staff WHERE status = 1")
                for p in persons:
                    person_combo.addItem(f"{p['name']} ({p['staff_id']})", p['id'])
        
        type_combo.currentIndexChanged.connect(load_persons)
        load_persons()
        form.addRow("Person:", person_combo)
        
        date_input = QDateEdit()
        date_input.setCalendarPopup(True)
        date_input.setDate(QDate.currentDate())
        form.addRow("Date:", date_input)
        
        checkin_input = QTimeEdit()
        checkin_input.setTime(QTime.currentTime())
        form.addRow("Check-In Time:", checkin_input)
        
        checkout_input = QTimeEdit()
        checkout_input.setTime(QTime.currentTime())
        form.addRow("Check-Out Time:", checkout_input)
        
        status_combo = QComboBox()
        status_combo.addItems(["Present", "Absent", "Late", "Half Day", "Leave"])
        form.addRow("Status:", status_combo)
        
        remarks_input = QTextEdit()
        remarks_input.setMaximumHeight(60)
        form.addRow("Remarks:", remarks_input)
        
        layout.addLayout(form)
        
        save_btn = QPushButton("Save Manual Attendance")
        save_btn.clicked.connect(lambda: self.save_manual_attendance({
            'person_type': type_combo.currentText().lower(),
            'person_id': person_combo.currentData(),
            'date': date_input.date().toString('yyyy-MM-dd'),
            'checkin': checkin_input.time().toString('HH:mm:ss'),
            'checkout': checkout_input.time().toString('HH:mm:ss'),
            'status': status_combo.currentText(),
            'remarks': remarks_input.toPlainText()
        }))
        layout.addWidget(save_btn)
        
        layout.addStretch()
        widget.setLayout(layout)
        return widget
    
    def save_manual_attendance(self, data):
        try:
            existing = self.db.fetch_one(
                "SELECT id FROM attendance_logs WHERE person_type = ? AND person_id = ? AND log_date = ?",
                (data['person_type'], data['person_id'], data['date'])
            )
            
            if existing:
                self.db.execute_query(
                    "UPDATE attendance_logs SET check_in_time = ?, check_out_time = ?, status = ?, punch_type = 'manual', remarks = ? WHERE id = ?",
                    (data['checkin'], data['checkout'], data['status'], data['remarks'], existing['id'])
                )
            else:
                self.db.execute_query(
                    "INSERT INTO attendance_logs (person_type, person_id, log_date, check_in_time, check_out_time, status, punch_type, remarks, created_at) VALUES (?, ?, ?, ?, ?, ?, 'manual', ?, ?)",
                    (data['person_type'], data['person_id'], data['date'], data['checkin'], data['checkout'], data['status'], data['remarks'], datetime.now().isoformat())
                )
            
            QMessageBox.information(self, "Success", "Manual attendance saved successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save attendance: {str(e)}")
    
    def create_reports(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        title = QLabel("Reports")
        title.setStyleSheet("font-size: 24px; font-weight: bold; padding: 10px;")
        layout.addWidget(title)
        
        reports_group = QGroupBox("Generate Reports")
        reports_layout = QVBoxLayout()
        
        filter_layout = QFormLayout()
        
        report_type_combo = QComboBox()
        report_type_combo.addItems([
            "Daily Attendance Report",
            "Monthly Attendance Report",
            "Individual Person Report",
            "Class-wise Student Report",
            "Department-wise Staff Report",
            "Late  Arrivals Report",
            "Absent Report",
            "Leave Report"
        ])
        filter_layout.addRow("Report Type:", report_type_combo)
        
        self.report_from_date = QDateEdit()
        self.report_from_date.setCalendarPopup(True)
        self.report_from_date.setDate(QDate.currentDate().addMonths(-1))
        filter_layout.addRow("From Date:", self.report_from_date)
        
        self.report_to_date = QDateEdit()
        self.report_to_date.setCalendarPopup(True)
        self.report_to_date.setDate(QDate.currentDate())
        filter_layout.addRow("To Date:", self.report_to_date)
        
        self.report_person_type = QComboBox()
        self.report_person_type.addItems(["All", "Teacher", "Student", "Staff"])
        filter_layout.addRow("Person Type:", self.report_person_type)
        
        reports_layout.addLayout(filter_layout)
        
        buttons_layout = QHBoxLayout()
        
        generate_btn = QPushButton("Generate Report")
        generate_btn.clicked.connect(lambda: self.generate_report(report_type_combo.currentText()))
        buttons_layout.addWidget(generate_btn)
        
        if REPORTLAB_AVAILABLE:
            export_pdf_btn = QPushButton("Export to PDF")
            export_pdf_btn.clicked.connect(lambda: self.export_report_pdf(report_type_combo.currentText()))
            buttons_layout.addWidget(export_pdf_btn)
        
        if OPENPYXL_AVAILABLE:
            export_excel_btn = QPushButton("Export to Excel")
            export_excel_btn.clicked.connect(lambda: self.export_report_excel(report_type_combo.currentText()))
            buttons_layout.addWidget(export_excel_btn)
        
        reports_layout.addLayout(buttons_layout)
        reports_group.setLayout(reports_layout)
        layout.addWidget(reports_group)
        
        self.reports_table = QTableWidget()
        self.reports_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.reports_table)
        
        widget.setLayout(layout)
        return widget
    
    def generate_report(self, report_type):
        from_date = self.report_from_date.date().toString('yyyy-MM-dd')
        to_date = self.report_to_date.date().toString('yyyy-MM-dd')
        person_type = self.report_person_type.currentText().lower()
        
        if report_type == "Daily Attendance Report":
            self.generate_daily_report(to_date, person_type)
        elif report_type == "Monthly Attendance Report":
            self.generate_monthly_report(from_date, to_date, person_type)
        elif report_type == "Late Arrivals Report":
            self.generate_late_report(from_date, to_date, person_type)
        elif report_type == "Absent Report":
            self.generate_absent_report(from_date, to_date, person_type)
    
    def generate_daily_report(self, report_date, person_type):
        if person_type == "all":
            logs = self.db.fetch_all(
                "SELECT * FROM attendance_logs WHERE log_date = ? ORDER BY person_type, person_id",
                (report_date,)
            )
        else:
            logs = self.db.fetch_all(
                "SELECT * FROM attendance_logs WHERE log_date = ? AND person_type = ? ORDER BY person_id",
                (report_date, person_type)
            )
        
        self.reports_table.setColumnCount(8)
        self.reports_table.setHorizontalHeaderLabels(["Name", "Type", "ID", "Check-In", "Check-Out", "Status", "Late", "Remarks"])
        self.reports_table.setRowCount(len(logs))
        
        for idx, log in enumerate(logs):
            person_name = "Unknown"
            person_id_text = ""
            
            if log['person_type'] == 'teacher':
                person = self.db.fetch_one("SELECT name, teacher_id FROM teachers WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
                    person_id_text = person['teacher_id']
            elif log['person_type'] == 'student':
                person = self.db.fetch_one("SELECT name, roll_no FROM students WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
                    person_id_text = person['roll_no']
            elif log['person_type'] == 'staff':
                person = self.db.fetch_one("SELECT name, staff_id FROM staff WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
                    person_id_text = person['staff_id']
            
            self.reports_table.setItem(idx, 0, QTableWidgetItem(person_name))
            self.reports_table.setItem(idx, 1, QTableWidgetItem(log['person_type'].title()))
            self.reports_table.setItem(idx, 2, QTableWidgetItem(person_id_text))
            self.reports_table.setItem(idx, 3, QTableWidgetItem(log['check_in_time'] or ''))
            self.reports_table.setItem(idx, 4, QTableWidgetItem(log['check_out_time'] or ''))
            self.reports_table.setItem(idx, 5, QTableWidgetItem(log['status']))
            self.reports_table.setItem(idx, 6, QTableWidgetItem("Yes" if log['is_late'] else "No"))
            self.reports_table.setItem(idx, 7, QTableWidgetItem(log['remarks'] or ''))
    
    def generate_monthly_report(self, from_date, to_date, person_type):
        if person_type == "all":
            logs = self.db.fetch_all(
                "SELECT person_type, person_id, COUNT(*) as days_present, SUM(is_late) as late_days FROM attendance_logs WHERE log_date BETWEEN ? AND ? AND status = 'Present' GROUP BY person_type, person_id",
                (from_date, to_date)
            )
        else:
            logs = self.db.fetch_all(
                "SELECT person_type, person_id, COUNT(*) as days_present, SUM(is_late) as late_days FROM attendance_logs WHERE log_date BETWEEN ? AND ? AND person_type = ? AND status = 'Present' GROUP BY person_type, person_id",
                (from_date, to_date, person_type)
            )
        
        self.reports_table.setColumnCount(6)
        self.reports_table.setHorizontalHeaderLabels(["Name", "Type", "ID", "Days Present", "Late Days", "Attendance %"])
        self.reports_table.setRowCount(len(logs))
        
        total_days = (datetime.strptime(to_date, '%Y-%m-%d') - datetime.strptime(from_date, '%Y-%m-%d')).days + 1
        
        for idx, log in enumerate(logs):
            person_name = "Unknown"
            person_id_text = ""
            
            if log['person_type'] == 'teacher':
                person = self.db.fetch_one("SELECT name, teacher_id FROM teachers WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
                    person_id_text = person['teacher_id']
            elif log['person_type'] == 'student':
                person = self.db.fetch_one("SELECT name, roll_no FROM students WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
                    person_id_text = person['roll_no']
            elif log['person_type'] == 'staff':
                person = self.db.fetch_one("SELECT name, staff_id FROM staff WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
                    person_id_text = person['staff_id']
            
            attendance_percent = (log['days_present'] / total_days) * 100 if total_days > 0 else 0
            
            self.reports_table.setItem(idx, 0, QTableWidgetItem(person_name))
            self.reports_table.setItem(idx, 1, QTableWidgetItem(log['person_type'].title()))
            self.reports_table.setItem(idx, 2, QTableWidgetItem(person_id_text))
            self.reports_table.setItem(idx, 3, QTableWidgetItem(str(log['days_present'])))
            self.reports_table.setItem(idx, 4, QTableWidgetItem(str(log['late_days'])))
            self.reports_table.setItem(idx, 5, QTableWidgetItem(f"{attendance_percent:.2f}%"))
    
    def generate_late_report(self, from_date, to_date, person_type):
        if person_type == "all":
            logs = self.db.fetch_all(
                "SELECT * FROM attendance_logs WHERE log_date BETWEEN ? AND ? AND is_late = 1 ORDER BY log_date DESC",
                (from_date, to_date)
            )
        else:
            logs = self.db.fetch_all(
                "SELECT * FROM attendance_logs WHERE log_date BETWEEN ? AND ? AND person_type = ? AND is_late = 1 ORDER BY log_date DESC",
                (from_date, to_date, person_type)
            )
        
        self.reports_table.setColumnCount(6)
        self.reports_table.setHorizontalHeaderLabels(["Date", "Name", "Type", "ID", "Check-In Time", "Remarks"])
        self.reports_table.setRowCount(len(logs))
        
        for idx, log in enumerate(logs):
            person_name = "Unknown"
            person_id_text = ""
            
            if log['person_type'] == 'teacher':
                person = self.db.fetch_one("SELECT name, teacher_id FROM teachers WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
                    person_id_text = person['teacher_id']
            elif log['person_type'] == 'student':
                person = self.db.fetch_one("SELECT name, roll_no FROM students WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
                    person_id_text = person['roll_no']
            elif log['person_type'] == 'staff':
                person = self.db.fetch_one("SELECT name, staff_id FROM staff WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
                    person_id_text = person['staff_id']
            
            self.reports_table.setItem(idx, 0, QTableWidgetItem(log['log_date']))
            self.reports_table.setItem(idx, 1, QTableWidgetItem(person_name))
            self.reports_table.setItem(idx, 2, QTableWidgetItem(log['person_type'].title()))
            self.reports_table.setItem(idx, 3, QTableWidgetItem(person_id_text))
            self.reports_table.setItem(idx, 4, QTableWidgetItem(log['check_in_time'] or ''))
            self.reports_table.setItem(idx, 5, QTableWidgetItem(log['remarks'] or ''))
    
    def generate_absent_report(self, from_date, to_date, person_type):
        if person_type == "all":
            logs = self.db.fetch_all(
                "SELECT * FROM attendance_logs WHERE log_date BETWEEN ? AND ? AND status = 'Absent' ORDER BY log_date DESC",
                (from_date, to_date)
            )
        else:
            logs = self.db.fetch_all(
                "SELECT * FROM attendance_logs WHERE log_date BETWEEN ? AND ? AND person_type = ? AND status = 'Absent' ORDER BY log_date DESC",
                (from_date, to_date, person_type)
            )
        
        self.reports_table.setColumnCount(5)
        self.reports_table.setHorizontalHeaderLabels(["Date", "Name", "Type", "ID", "Remarks"])
        self.reports_table.setRowCount(len(logs))
        
        for idx, log in enumerate(logs):
            person_name = "Unknown"
            person_id_text = ""
            
            if log['person_type'] == 'teacher':
                person = self.db.fetch_one("SELECT name, teacher_id FROM teachers WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
                    person_id_text = person['teacher_id']
            elif log['person_type'] == 'student':
                person = self.db.fetch_one("SELECT name, roll_no FROM students WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
                    person_id_text = person['roll_no']
            elif log['person_type'] == 'staff':
                person = self.db.fetch_one("SELECT name, staff_id FROM staff WHERE id = ?", (log['person_id'],))
                if person:
                    person_name = person['name']
                    person_id_text = person['staff_id']
            
            self.reports_table.setItem(idx, 0, QTableWidgetItem(log['log_date']))
            self.reports_table.setItem(idx, 1, QTableWidgetItem(person_name))
            self.reports_table.setItem(idx, 2, QTableWidgetItem(log['person_type'].title()))
            self.reports_table.setItem(idx, 3, QTableWidgetItem(person_id_text))
            self.reports_table.setItem(idx, 4, QTableWidgetItem(log['remarks'] or ''))
    
    def export_report_pdf(self, report_type):
        if not REPORTLAB_AVAILABLE:
            QMessageBox.warning(self, "Library Missing", "ReportLab is not installed.")
            return
        
        filename, _ = QFileDialog.getSaveFileName(self, "Save PDF Report", "", "PDF Files (*.pdf)")
        if not filename:
            return
        
        try:
            doc = SimpleDocTemplate(filename, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#0078d4'), spaceAfter=30, alignment=TA_CENTER)
            elements.append(Paragraph(report_type, title_style))
            elements.append(Spacer(1, 0.2*inch))
            
            table_data = []
            headers = []
            for col in range(self.reports_table.columnCount()):
                headers.append(self.reports_table.horizontalHeaderItem(col).text())
            table_data.append(headers)
            
            for row in range(self.reports_table.rowCount()):
                row_data = []
                for col in range(self.reports_table.columnCount()):
                    item = self.reports_table.item(row, col)
                    row_data.append(item.text() if item else '')
                table_data.append(row_data)
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0078d4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            QMessageBox.information(self, "Success", f"Report exported to {filename}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export PDF: {str(e)}")
    
    def export_report_excel(self, report_type):
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "Library Missing", "openpyxl is not installed.")
            return
        
        filename, _ = QFileDialog.getSaveFileName(self, "Save Excel Report", "", "Excel Files (*.xlsx)")
        if not filename:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = report_type[:31]
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col in range(self.reports_table.columnCount()):
                cell = ws.cell(row=1, column=col+1)
                cell.value = self.reports_table.horizontalHeaderItem(col).text()
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            for row in range(self.reports_table.rowCount()):
                for col in range(self.reports_table.columnCount()):
                    item = self.reports_table.item(row, col)
                    ws.cell(row=row+2, column=col+1, value=item.text() if item else '')
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(filename)
            QMessageBox.information(self, "Success", f"Report exported to {filename}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export Excel: {str(e)}")
    
    def create_settings(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        title = QLabel("Settings")
        title.setStyleSheet("font-size: 24px; font-weight: bold; padding: 10px;")
        layout.addWidget(title)
        
        tabs = QTabWidget()
        
        institute_tab = self.create_institute_settings()
        tabs.addTab(institute_tab, "Institute")
        
        shifts_tab = self.create_shifts_settings()
        tabs.addTab(shifts_tab, "Shifts")
        
        holidays_tab = self.create_holidays_settings()
        tabs.addTab(holidays_tab, "Holidays")
        
        backup_tab = self.create_backup_settings()
        tabs.addTab(backup_tab, "Backup & Restore")
        
        layout.addWidget(tabs)
        widget.setLayout(layout)
        return widget
    
    def create_institute_settings(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        form = QFormLayout()
        
        institute_data = self.db.fetch_one("SELECT * FROM institute_settings LIMIT 1")
        
        name_input = QLineEdit(institute_data['institute_name'] if institute_data else '')
        form.addRow("Institute Name:", name_input)
        
        address_input = QTextEdit(institute_data['address'] if institute_data and institute_data['address'] else '')
        address_input.setMaximumHeight(60)
        form.addRow("Address:", address_input)
        
        phone_input = QLineEdit(institute_data['phone'] if institute_data and institute_data['phone'] else '')
        form.addRow("Phone:", phone_input)
        
        email_input = QLineEdit(institute_data['email'] if institute_data and institute_data['email'] else '')
        form.addRow("Email:", email_input)
        
        year_input = QSpinBox()
        year_input.setRange(1900, 2100)
        year_input.setValue(institute_data['established_year'] if institute_data and institute_data['established_year'] else 2020)
        form.addRow("Established Year:", year_input)
        
        reg_input = QLineEdit(institute_data['registration_no'] if institute_data and institute_data['registration_no'] else '')
        form.addRow("Registration No:", reg_input)
        
        layout.addLayout(form)
        
        save_btn = QPushButton("Save Institute Settings")
        save_btn.clicked.connect(lambda: self.save_institute_settings({
            'name': name_input.text(),
            'address': address_input.toPlainText(),
            'phone': phone_input.text(),
            'email': email_input.text(),
            'year': year_input.value(),
            'reg_no': reg_input.text()
        }))
        layout.addWidget(save_btn)
        
        layout.addStretch()
        widget.setLayout(layout)
        return widget
    
    def save_institute_settings(self, data):
        try:
            existing = self.db.fetch_one("SELECT id FROM institute_settings LIMIT 1")
            
            if existing:
                self.db.execute_query(
                    "UPDATE institute_settings SET institute_name = ?, address = ?, phone = ?, email = ?, established_year = ?, registration_no = ? WHERE id = ?",
                    (data['name'], data['address'], data['phone'], data['email'], data['year'], data['reg_no'], existing['id'])
                )
            else:
                self.db.execute_query(
                    "INSERT INTO institute_settings (institute_name, address, phone, email, established_year, registration_no) VALUES (?, ?, ?, ?, ?, ?)",
                    (data['name'], data['address'], data['phone'], data['email'], data['year'], data['reg_no'])
                )
            
            QMessageBox.information(self, "Success", "Institute settings saved successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save settings: {str(e)}")
    
    def create_shifts_settings(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        toolbar_layout = QHBoxLayout()
        
        add_shift_btn = QPushButton("Add Shift")
        add_shift_btn.clicked.connect(self.add_shift)
        toolbar_layout.addWidget(add_shift_btn)
        
        refresh_shifts_btn = QPushButton("Refresh")
        refresh_shifts_btn.clicked.connect(self.load_shifts)
        toolbar_layout.addWidget(refresh_shifts_btn)
        
        toolbar_layout.addStretch()
        layout.addLayout(toolbar_layout)
        
        self.shifts_table = QTableWidget()
        self.shifts_table.setColumnCount(6)
        self.shifts_table.setHorizontalHeaderLabels(["ID", "Name", "Start Time", "End Time", "Grace Time (min)", "Actions"])
        self.shifts_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.shifts_table)
        
        widget.setLayout(layout)
        self.load_shifts()
        return widget
    
    def load_shifts(self):
        shifts = self.db.fetch_all("SELECT * FROM shifts ORDER BY created_at DESC")
        
        self.shifts_table.setRowCount(len(shifts))
        for idx, shift in enumerate(shifts):
            self.shifts_table.setItem(idx, 0, QTableWidgetItem(str(shift['id'])))
            self.shifts_table.setItem(idx, 1, QTableWidgetItem(shift['name']))
            self.shifts_table.setItem(idx, 2, QTableWidgetItem(shift['start_time']))
            self.shifts_table.setItem(idx, 3, QTableWidgetItem(shift['end_time']))
            self.shifts_table.setItem(idx, 4, QTableWidgetItem(str(shift['grace_time'])))
            
            actions_widget = QWidget()
            actions_layout = QHBoxLayout()
            actions_layout.setContentsMargins(0, 0, 0, 0)
            
            edit_btn = QPushButton("Edit")
            edit_btn.clicked.connect(lambda checked, s=shift: self.edit_shift(s))
            actions_layout.addWidget(edit_btn)
            
            delete_btn = QPushButton("Delete")
            delete_btn.clicked.connect(lambda checked, s=shift: self.delete_shift(s))
            actions_layout.addWidget(delete_btn)
            
            actions_widget.setLayout(actions_layout)
            self.shifts_table.setCellWidget(idx, 5, actions_widget)
    
    def add_shift(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Shift")
        dialog.resize(400, 300)
        
        layout = QVBoxLayout()
        form = QFormLayout()
        
        name_input = QLineEdit()
        form.addRow("Shift Name:", name_input)
        
        start_time = QTimeEdit()
        start_time.setTime(QTime(8, 0))
        form.addRow("Start Time:", start_time)
        
        end_time = QTimeEdit()
        end_time.setTime(QTime(17, 0))
        form.addRow("End Time:", end_time)
        
        grace_input = QSpinBox()
        grace_input.setRange(0, 120)
        grace_input.setValue(15)
        form.addRow("Grace Time (minutes):", grace_input)
        
        layout.addLayout(form)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(lambda: self.save_shift(dialog, {
            'name': name_input.text(),
            'start_time': start_time.time().toString('HH:mm'),
            'end_time': end_time.time().toString('HH:mm'),
            'grace_time': grace_input.value()
        }))
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def save_shift(self, dialog, data):
        if not data['name']:
            QMessageBox.warning(dialog, "Validation Error", "Shift name is required.")
            return
        
        try:
            self.db.execute_query(
                "INSERT INTO shifts (name, start_time, end_time, grace_time, created_at) VALUES (?, ?, ?, ?, ?)",
                (data['name'], data['start_time'], data['end_time'], data['grace_time'], datetime.now().isoformat())
            )
            QMessageBox.information(dialog, "Success", "Shift added successfully!")
            dialog.accept()
            self.load_shifts()
        except Exception as e:
            QMessageBox.critical(dialog, "Error", f"Failed to add shift: {str(e)}")
    
    def edit_shift(self, shift):
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Shift")
        dialog.resize(400, 300)
        
        layout = QVBoxLayout()
        form = QFormLayout()
        
        name_input = QLineEdit(shift['name'])
        form.addRow("Shift Name:", name_input)
        
        start_time = QTimeEdit()
        start_time.setTime(QTime.fromString(shift['start_time'], 'HH:mm'))
        form.addRow("Start Time:", start_time)
        
        end_time = QTimeEdit()
        end_time.setTime(QTime.fromString(shift['end_time'], 'HH:mm'))
        form.addRow("End Time:", end_time)
        
        grace_input = QSpinBox()
        grace_input.setRange(0, 120)
        grace_input.setValue(shift['grace_time'])
        form.addRow("Grace Time (minutes):", grace_input)
        
        layout.addLayout(form)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(lambda: self.update_shift(dialog, shift['id'], {
            'name': name_input.text(),
            'start_time': start_time.time().toString('HH:mm'),
            'end_time': end_time.time().toString('HH:mm'),
            'grace_time': grace_input.value()
        }))
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def update_shift(self, dialog, shift_id, data):
        if not data['name']:
            QMessageBox.warning(dialog, "Validation Error", "Shift name is required.")
            return
        
        try:
            self.db.execute_query(
                "UPDATE shifts SET name = ?, start_time = ?, end_time = ?, grace_time = ? WHERE id = ?",
                (data['name'], data['start_time'], data['end_time'], data['grace_time'], shift_id)
            )
            QMessageBox.information(dialog, "Success", "Shift updated successfully!")
            dialog.accept()
            self.load_shifts()
        except Exception as e:
            QMessageBox.critical(dialog, "Error", f"Failed to update shift: {str(e)}")
    
    def delete_shift(self, shift):
        reply = QMessageBox.question(self, 'Delete Shift', 
                                    f"Are you sure you want to delete {shift['name']}?",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.db.execute_query("DELETE FROM shifts WHERE id = ?", (shift['id'],))
                QMessageBox.information(self, "Success", "Shift deleted successfully!")
                self.load_shifts()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete shift: {str(e)}")
    
    def create_holidays_settings(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        toolbar_layout = QHBoxLayout()
        
        add_holiday_btn = QPushButton("Add Holiday")
        add_holiday_btn.clicked.connect(self.add_holiday)
        toolbar_layout.addWidget(add_holiday_btn)
        
        refresh_holidays_btn = QPushButton("Refresh")
        refresh_holidays_btn.clicked.connect(self.load_holidays)
        toolbar_layout.addWidget(refresh_holidays_btn)
        
        toolbar_layout.addStretch()
        layout.addLayout(toolbar_layout)
        
        self.holidays_table = QTableWidget()
        self.holidays_table.setColumnCount(5)
        self.holidays_table.setHorizontalHeaderLabels(["ID", "Name", "Date", "Type", "Actions"])
        self.holidays_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.holidays_table)
        
        widget.setLayout(layout)
        self.load_holidays()
        return widget
    
    def load_holidays(self):
        holidays = self.db.fetch_all("SELECT * FROM holidays ORDER BY date DESC")
        
        self.holidays_table.setRowCount(len(holidays))
        for idx, holiday in enumerate(holidays):
            self.holidays_table.setItem(idx, 0, QTableWidgetItem(str(holiday['id'])))
            self.holidays_table.setItem(idx, 1, QTableWidgetItem(holiday['name']))
            self.holidays_table.setItem(idx, 2, QTableWidgetItem(holiday['date']))
            self.holidays_table.setItem(idx, 3, QTableWidgetItem(holiday['type']))
            
            actions_widget = QWidget()
            actions_layout = QHBoxLayout()
            actions_layout.setContentsMargins(0, 0, 0, 0)
            
            delete_btn = QPushButton("Delete")
            delete_btn.clicked.connect(lambda checked, h=holiday: self.delete_holiday(h))
            actions_layout.addWidget(delete_btn)
            
            actions_widget.setLayout(actions_layout)
            self.holidays_table.setCellWidget(idx, 4, actions_widget)
    
    def add_holiday(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Holiday")
        dialog.resize(400, 250)
        
        layout = QVBoxLayout()
        form = QFormLayout()
        
        name_input = QLineEdit()
        form.addRow("Holiday Name:", name_input)
        
        date_input = QDateEdit()
        date_input.setCalendarPopup(True)
        date_input.setDate(QDate.currentDate())
        form.addRow("Date:", date_input)
        
        type_combo = QComboBox()
        type_combo.addItems(["National", "Religious", "Custom"])
        form.addRow("Type:", type_combo)
        
        desc_input = QTextEdit()
        desc_input.setMaximumHeight(60)
        form.addRow("Description:", desc_input)
        
        layout.addLayout(form)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(lambda: self.save_holiday(dialog, {
            'name': name_input.text(),
            'date': date_input.date().toString('yyyy-MM-dd'),
            'type': type_combo.currentText(),
            'description': desc_input.toPlainText()
        }))
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def save_holiday(self, dialog, data):
        if not data['name']:
            QMessageBox.warning(dialog, "Validation Error", "Holiday name is required.")
            return
        
        try:
            self.db.execute_query(
                "INSERT INTO holidays (name, date, type, description, created_at) VALUES (?, ?, ?, ?, ?)",
                (data['name'], data['date'], data['type'], data['description'], datetime.now().isoformat())
            )
            QMessageBox.information(dialog, "Success", "Holiday added successfully!")
            dialog.accept()
            self.load_holidays()
        except Exception as e:
            QMessageBox.critical(dialog, "Error", f"Failed to add holiday: {str(e)}")
    
    def delete_holiday(self, holiday):
        reply = QMessageBox.question(self, 'Delete Holiday', 
                                    f"Are you sure you want to delete {holiday['name']}?",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.db.execute_query("DELETE FROM holidays WHERE id = ?", (holiday['id'],))
                QMessageBox.information(self, "Success", "Holiday deleted successfully!")
                self.load_holidays()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete holiday: {str(e)}")
    
    def create_backup_settings(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        backup_group = QGroupBox("Database Backup")
        backup_layout = QVBoxLayout()
        
        backup_btn = QPushButton("Create Backup")
        backup_btn.clicked.connect(self.create_backup)
        backup_layout.addWidget(backup_btn)
        
        backup_group.setLayout(backup_layout)
        layout.addWidget(backup_group)
        
        restore_group = QGroupBox("Database Restore")
        restore_layout = QVBoxLayout()
        
        restore_btn = QPushButton("Restore from Backup")
        restore_btn.clicked.connect(self.restore_backup)
        restore_layout.addWidget(restore_btn)
        
        restore_group.setLayout(restore_layout)
        layout.addWidget(restore_group)
        
        layout.addStretch()
        widget.setLayout(layout)
        return widget
    
    def create_backup(self):
        filename, _ = QFileDialog.getSaveFileName(
            self, 
            "Save Backup", 
            f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db", 
            "Database Files (*.db)"
        )
        
        if filename:
            if self.db.backup_database(filename):
                QMessageBox.information(self, "Success", f"Backup created successfully at {filename}")
            else:
                QMessageBox.critical(self, "Error", "Failed to create backup")
    
    def restore_backup(self):
        reply = QMessageBox.question(
            self, 
            'Confirm Restore', 
            'This will replace all current data. Are you sure?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            filename, _ = QFileDialog.getOpenFileName(
                self, 
                "Select Backup File", 
                "", 
                "Database Files (*.db)"
            )
            
            if filename:
                if self.db.restore_database(filename):
                    QMessageBox.information(self, "Success", "Database restored successfully. Please restart the application.")
                else:
                    QMessageBox.critical(self, "Error", "Failed to restore database")


def main():
    app = QApplication(sys.argv)
    app.setApplicationName("Biometric Attendance Management System")
    
    db = Database()
    biometric = NBioBSPWrapper()
    
    login_window = LoginWindow(db)
    
    if login_window.exec() == QDialog.DialogCode.Accepted:
        main_window = MainWindow(db, biometric, login_window.current_user)
        main_window.show()
        
        exit_code = app.exec()
        
        biometric.cleanup()
        sys.exit(exit_code)
    else:
        biometric.cleanup()
        sys.exit(0)


if __name__ == '__main__':
    main()
